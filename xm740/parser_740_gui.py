#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
.740 工装数据文件解析工具 - GUI界面
解析工装记录的 .740 二进制数据文件，支持查看、绘图、导出Excel。
"""

import struct
import datetime
import os
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path


# ═══════════════════════════════════════════════════════════════
# 解析核心
# ═══════════════════════════════════════════════════════════════

class Parser740:
    """解析 .740 二进制数据文件"""

    RECORD_SIZE = 112       # 每条记录 112 字节
    HEADER_SIZE = 16        # 前 16 字节为头部
    NUM_CHANNELS = 12       # 12 个通道
    NUM_ACCUMULATED = 12    # 12 个累积值

    COLUMN_NAMES = (
        ['时间', '报警']
        + [f'通道{i}' for i in range(1, 13)]
        + [f'累积值{i}' for i in range(1, 13)]
    )

    def parse(self, filepath):
        """
        解析 .740 文件，返回记录列表。
        每条记录为 dict: {time, alarm, channels[12], accumulated[12]}
        """
        with open(filepath, 'rb') as f:
            data = f.read()

        if len(data) % self.RECORD_SIZE != 0:
            raise ValueError(
                f"文件大小 {len(data)} 字节不是 {self.RECORD_SIZE} 的整数倍，"
                "可能不是有效的 .740 文件"
            )

        num_records = len(data) // self.RECORD_SIZE
        records = []

        for i in range(num_records):
            offset = i * self.RECORD_SIZE
            rec = data[offset:offset + self.RECORD_SIZE]

            # 头部校验 (首字节固定为 1a，第2字节为版本号，兼容 03/04 等)
            if rec[0:1] != b'\x1a':
                raise ValueError(f"记录 {i} 头部标识不匹配: {rec[0:4].hex()}")

            # 时间戳 (Unix timestamp, uint32 LE, offset 8)
            timestamp = struct.unpack_from('<I', rec, 8)[0]
            dt = datetime.datetime.fromtimestamp(timestamp)

            # 12 个通道值 (float32 LE, offset 16)
            channels = list(struct.unpack_from('<12f', rec, 16))

            # 12 个累积值 (float32 LE, offset 64)
            accumulated = list(struct.unpack_from('<12f', rec, 64))

            records.append({
                'time': dt,
                'alarm': 0,
                'channels': channels,
                'accumulated': accumulated,
            })

        return records

    def get_active_channels(self, records):
        """检测哪些通道有非零数据，返回通道索引列表 (0-based)"""
        active = set()
        for rec in records:
            for idx, val in enumerate(rec['channels']):
                if val != 0.0:
                    active.add(idx)
        return sorted(active)

    def export_to_excel(self, records, output_path):
        """导出为 Excel 文件（.xlsx）"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '监控数据'

        # 写表头
        headers = self.COLUMN_NAMES
        header_font = Font(name='Arial', size=9, bold=True)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        header_align = Alignment(horizontal='center')

        for col_idx, name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # 写数据
        data_font = Font(name='Arial', size=9)
        for row_idx, rec in enumerate(records, 2):
            ws.cell(row=row_idx, column=1, value=rec['time']).font = data_font
            ws.cell(row=row_idx, column=2, value=rec['alarm']).font = data_font
            for ch_idx, val in enumerate(rec['channels']):
                cell = ws.cell(row=row_idx, column=3 + ch_idx, value=round(val, 2))
                cell.font = data_font
                cell.number_format = '0.00'
            for acc_idx, val in enumerate(rec['accumulated']):
                cell = ws.cell(row=row_idx, column=15 + acc_idx, value=round(val, 2))
                cell.font = data_font
                cell.number_format = '0.00'

        # 设置时间列宽度
        ws.column_dimensions['A'].width = 20

        wb.save(output_path)
        return output_path


# ═══════════════════════════════════════════════════════════════
# GUI 界面
# ═══════════════════════════════════════════════════════════════

class Parser740GUI:
    def __init__(self, root):
        self.root = root
        self.root.title(".740 工装数据解析工具")
        self.root.geometry("1100x750")
        self.root.minsize(900, 600)

        self.parser = Parser740()
        self.current_records = []
        self.current_file = None
        self.chart_window = None

        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 8, "pady": 4}

        # ── 顶部：文件操作区 ──
        file_frame = ttk.LabelFrame(self.root, text="文件操作", padding=8)
        file_frame.pack(fill=tk.X, **pad)

        ttk.Button(file_frame, text="打开 .740 文件", command=self._open_file).pack(side=tk.LEFT, padx=4)
        ttk.Button(file_frame, text="批量解析", command=self._batch_parse).pack(side=tk.LEFT, padx=4)

        ttk.Separator(file_frame, orient=tk.VERTICAL).pack(side=tk.LEFT, fill=tk.Y, padx=8)

        ttk.Button(file_frame, text="导出 Excel", command=self._export_excel).pack(side=tk.LEFT, padx=4)
        ttk.Button(file_frame, text="显示曲线图", command=self._show_chart).pack(side=tk.LEFT, padx=4)

        self.var_file_info = tk.StringVar(value="未加载文件")
        ttk.Label(file_frame, textvariable=self.var_file_info, foreground="gray").pack(side=tk.RIGHT, padx=8)

        # ── 中部：数据表格区 ──
        table_frame = ttk.LabelFrame(self.root, text="数据预览", padding=4)
        table_frame.pack(fill=tk.BOTH, expand=True, **pad)

        # Treeview + 滚动条
        tree_container = ttk.Frame(table_frame)
        tree_container.pack(fill=tk.BOTH, expand=True)

        y_scroll = ttk.Scrollbar(tree_container, orient=tk.VERTICAL)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        x_scroll = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL)
        x_scroll.pack(side=tk.BOTTOM, fill=tk.X)

        self.tree = ttk.Treeview(
            tree_container,
            yscrollcommand=y_scroll.set,
            xscrollcommand=x_scroll.set,
            show='headings',
            height=20,
        )
        self.tree.pack(fill=tk.BOTH, expand=True)
        y_scroll.config(command=self.tree.yview)
        x_scroll.config(command=self.tree.xview)

        # ── 底部：状态信息 ──
        status_frame = ttk.Frame(self.root)
        status_frame.pack(fill=tk.X, **pad)

        self.var_status = tk.StringVar(value="就绪")
        ttk.Label(status_frame, textvariable=self.var_status).pack(side=tk.LEFT)

        self.progress = ttk.Progressbar(status_frame, mode="determinate", maximum=100)
        self.progress.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(8, 0))

    # ── 打开文件 ──

    def _open_file(self):
        filepath = filedialog.askopenfilename(
            title="选择 .740 数据文件",
            filetypes=[("740 数据文件", "*.740"), ("所有文件", "*.*")],
        )
        if not filepath:
            return

        self.var_status.set(f"正在解析: {os.path.basename(filepath)} ...")
        self.progress["value"] = 0
        self.root.update_idletasks()

        try:
            records = self.parser.parse(filepath)
            self.current_records = records
            self.current_file = filepath
            self._populate_table(records)

            active = self.parser.get_active_channels(records)
            active_str = ', '.join(f'通道{i+1}' for i in active) if active else '无'
            start_time = records[0]['time'].strftime('%Y-%m-%d %H:%M:%S') if records else '-'
            end_time = records[-1]['time'].strftime('%Y-%m-%d %H:%M:%S') if records else '-'

            self.var_file_info.set(
                f"{os.path.basename(filepath)}  |  {len(records)} 条记录  |  "
                f"{start_time} ~ {end_time}  |  活跃通道: {active_str}"
            )
            self.var_status.set("解析完成")
            self.progress["value"] = 100

        except Exception as e:
            messagebox.showerror("解析错误", str(e))
            self.var_status.set("解析失败")

    # ── 批量解析 ──

    def _batch_parse(self):
        filepaths = filedialog.askopenfilenames(
            title="选择多个 .740 文件",
            filetypes=[("740 数据文件", "*.740"), ("所有文件", "*.*")],
        )
        if not filepaths:
            return

        output_dir = filedialog.askdirectory(title="选择输出文件夹")
        if not output_dir:
            return

        self.var_status.set(f"批量解析中... 共 {len(filepaths)} 个文件")
        self.progress["value"] = 0
        self.root.update_idletasks()

        # 在后台线程中执行
        thread = threading.Thread(
            target=self._do_batch_parse,
            args=(list(filepaths), output_dir),
            daemon=True,
        )
        thread.start()

    def _do_batch_parse(self, filepaths, output_dir):
        total = len(filepaths)
        success = 0
        failed = 0

        for idx, fp in enumerate(filepaths):
            try:
                records = self.parser.parse(fp)
                base_name = Path(fp).stem
                out_path = os.path.join(output_dir, f"{base_name}.xlsx")
                self.parser.export_to_excel(records, out_path)
                success += 1
            except Exception as e:
                failed += 1
                print(f"解析失败 {fp}: {e}")

            pct = int((idx + 1) / total * 100)
            self.root.after(0, self._update_batch_progress, pct, idx + 1, total)

        msg = f"批量解析完成: 成功 {success} 个, 失败 {failed} 个"
        self.root.after(0, self._on_batch_done, msg, output_dir)

    def _update_batch_progress(self, pct, current, total):
        self.progress["value"] = pct
        self.var_status.set(f"批量解析中... ({current}/{total})")

    def _on_batch_done(self, msg, output_dir):
        self.var_status.set(msg)
        self.progress["value"] = 100
        if messagebox.askyesno("完成", f"{msg}\n\n是否打开输出文件夹？"):
            os.startfile(output_dir)

    # ── 填充表格 ──

    def _populate_table(self, records):
        # 清空
        self.tree.delete(*self.tree.get_children())

        if not records:
            return

        # 检测活跃通道，只显示有数据的列
        active_ch = self.parser.get_active_channels(records)

        # 检测有累积值的通道
        active_acc = set()
        for rec in records:
            for idx, val in enumerate(rec['accumulated']):
                if val != 0.0:
                    active_acc.add(idx)
        active_acc = sorted(active_acc)

        # 构建列
        columns = ['time', 'alarm']
        col_headers = {'time': '时间', 'alarm': '报警'}

        for i in active_ch:
            key = f'ch{i}'
            columns.append(key)
            col_headers[key] = f'通道{i+1}'

        for i in active_acc:
            key = f'acc{i}'
            columns.append(key)
            col_headers[key] = f'累积值{i+1}'

        self.tree['columns'] = columns
        for col_id in columns:
            self.tree.heading(col_id, text=col_headers[col_id])
            width = 160 if col_id == 'time' else 80
            self.tree.column(col_id, width=width, anchor=tk.CENTER)

        # 填充数据
        for rec in records:
            values = [
                rec['time'].strftime('%Y-%m-%d %H:%M:%S'),
                rec['alarm'],
            ]
            for i in active_ch:
                values.append(f"{rec['channels'][i]:.2f}")
            for i in active_acc:
                values.append(f"{rec['accumulated'][i]:.2f}")

            self.tree.insert('', tk.END, values=values)

    # ── 导出 Excel ──

    def _export_excel(self):
        if not self.current_records:
            messagebox.showwarning("提示", "请先打开一个 .740 文件")
            return

        default_name = Path(self.current_file).stem + ".xlsx" if self.current_file else "output.xlsx"
        output_path = filedialog.asksaveasfilename(
            title="保存 Excel 文件",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if not output_path:
            return

        try:
            self.parser.export_to_excel(self.current_records, output_path)
            self.var_status.set(f"已导出: {output_path}")
            if messagebox.askyesno("导出成功", f"已保存到:\n{output_path}\n\n是否打开文件？"):
                os.startfile(output_path)
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    # ── 显示曲线图 ──

    def _show_chart(self):
        if not self.current_records:
            messagebox.showwarning("提示", "请先打开一个 .740 文件")
            return

        try:
            import matplotlib
            matplotlib.use('TkAgg')
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
            import matplotlib.dates as mdates
        except ImportError:
            messagebox.showerror("缺少依赖", "需要安装 matplotlib:\npip install matplotlib")
            return

        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial']
        plt.rcParams['axes.unicode_minus'] = False

        records = self.current_records
        active_ch = self.parser.get_active_channels(records)

        if not active_ch:
            messagebox.showinfo("提示", "没有活跃通道数据")
            return

        times = [r['time'] for r in records]

        # 创建图表窗口
        if self.chart_window and self.chart_window.winfo_exists():
            self.chart_window.destroy()

        self.chart_window = tk.Toplevel(self.root)
        self.chart_window.title(f"曲线图 - {os.path.basename(self.current_file or '')}")
        self.chart_window.geometry("1000x600")

        fig, ax = plt.subplots(figsize=(10, 5))
        fig.subplots_adjust(left=0.08, right=0.95, top=0.92, bottom=0.15)

        colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728',
                  '#9467bd', '#8c564b', '#e377c2', '#7f7f7f',
                  '#bcbd22', '#17becf', '#aec7e8', '#ffbb78']

        for ch_idx in active_ch:
            values = [r['channels'][ch_idx] for r in records]
            ax.plot(times, values, label=f'通道{ch_idx + 1}',
                    color=colors[ch_idx % len(colors)], linewidth=1.2)

        ax.set_xlabel('时间')
        ax.set_ylabel('数值')
        ax.set_title(os.path.basename(self.current_file or '监控曲线'))
        ax.legend(loc='upper right')
        ax.grid(True, alpha=0.3)

        ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        fig.autofmt_xdate(rotation=30)

        canvas = FigureCanvasTkAgg(fig, master=self.chart_window)
        canvas.draw()

        toolbar = NavigationToolbar2Tk(canvas, self.chart_window)
        toolbar.update()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)


# ═══════════════════════════════════════════════════════════════
# 入口
# ═══════════════════════════════════════════════════════════════

def main():
    root = tk.Tk()
    Parser740GUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
