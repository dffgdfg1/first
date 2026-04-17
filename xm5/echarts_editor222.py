"""
ECharts HTML数据编辑器
===================

功能说明：
1. 上传ECharts图表HTML文件
2. 解析并导出数据为Excel格式
3. 用户在Excel中编辑数据
4. 导入修改后的Excel文件
5. 生成保持原样式的exit)(新HTML文件

依赖库：
pip install beautifulsoup4 openpyxl

使用方法：
1. 运行程序
2. 点击"上传HTML文件"选择ECharts图表文件
3. 点击"导出Excel"将数据保存为Excel文件
4. 在Excel中编辑数据
5. 点击"导入Excel"加载修改后的数据
6. 点击"生成新HTML"保存修改后的HTML文件
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import re
from datetime import datetime
from bs4 import BeautifulSoup
import os
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# 导入智能波动算法
try:
    from smart_fluctuation import SmartFluctuationGenerator, analyze_and_generate
    SMART_FLUCTUATION_AVAILABLE = True
except ImportError:
    SMART_FLUCTUATION_AVAILABLE = False


class EChartsEditor:
    """ECharts HTML编辑器主类"""

    def __init__(self, root):
        self.root = root
        self.root.title("ECharts HTML数据编辑器")
        self.root.geometry("1200x700")

        # 数据存储
        self.html_files = []  # 存储多个HTML文件信息 [{'path': '', 'content': '', 'option': {}, 'series': []}]
        self.current_file_index = 0  # 当前选中的文件索引
        self.operation_backup = None  # 全局操作备份
        self.last_operation_params = None  # 存储上一次操作的参数

        # 兼容旧代码的属性
        self.html_content = None  # 原始HTML内容
        self.soup = None  # BeautifulSoup对象
        self.option_data = None  # 解析后的option对象
        self.series_data = []  # 所有series的数据
        self.file_path = None  # 当前文件路径

        # 创建GUI界面
        self.create_widgets()

    def create_widgets(self):
        """创建GUI组件"""
        # 顶部工具栏
        toolbar = ttk.Frame(self.root, padding="5")
        toolbar.pack(side=tk.TOP, fill=tk.X)

        ttk.Button(toolbar, text="1. 批量上传HTML", command=self.load_multiple_files).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="2. 导出Excel", command=self.export_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="3. 导入Excel", command=self.import_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="4. 生成新HTML", command=self.generate_html).pack(side=tk.LEFT, padx=5)

        # 分隔符
        ttk.Separator(toolbar, orient=tk.VERTICAL).pack(side=tk.LEFT, padx=10, fill=tk.Y)

        # 新增功能按钮


        ttk.Button(toolbar, text="退出", command=self.root.quit).pack(side=tk.RIGHT, padx=5)

        # 状态栏
        self.status_var = tk.StringVar(value="请批量上传ECharts HTML文件")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # 主内容区域 - 分为左右两部分
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 左侧：文件列表
        left_frame = ttk.LabelFrame(main_frame, text="已加载的文件", padding=5)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 5))

        # 文件列表框架
        list_container = ttk.Frame(left_frame)
        list_container.pack(fill=tk.BOTH, expand=True)

        # 文件列表
        self.file_listbox = tk.Listbox(list_container, width=30, height=20)
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.file_listbox.bind('<<ListboxSelect>>', self.on_file_select)

        # 文件列表滚动条
        file_scrollbar = ttk.Scrollbar(list_container, orient="vertical", command=self.file_listbox.yview)
        file_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.file_listbox.configure(yscrollcommand=file_scrollbar.set)

        # 文件操作按钮
        file_btn_frame = ttk.Frame(left_frame)
        file_btn_frame.pack(fill=tk.X, pady=(5, 0))
        ttk.Button(file_btn_frame, text="删除选中", command=self.delete_selected_file).pack(side=tk.LEFT, padx=2)
        ttk.Button(file_btn_frame, text="清空列表", command=self.clear_file_list).pack(side=tk.LEFT, padx=2)

        # 右侧：数据显示区域 - 使用Notebook创建标签页
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.notebook = ttk.Notebook(right_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # 创建数据表格标签页（初始为空，加载文件后动态创建）
        self.series_frames = []

    def load_multiple_files(self):
        """批量加载HTML文件"""
        file_paths = filedialog.askopenfilenames(
            title="选择ECharts HTML文件（可多选）",
            filetypes=[("HTML文件", "*.html"), ("所有文件", "*.*")]
        )

        if not file_paths:
            return

        success_count = 0
        fail_count = 0

        for file_path in file_paths:
            try:
                # 读取HTML文件
                with open(file_path, 'r', encoding='utf-8') as f:
                    html_content = f.read()

                # 解析HTML
                soup = BeautifulSoup(html_content, 'html.parser')
                scripts = soup.find_all('script')
                option_script = None

                for script in scripts:
                    if script.string and 'const option' in script.string:
                        option_script = script.string
                        break

                if not option_script:
                    fail_count += 1
                    continue

                # 提取option对象
                pattern = r'const\s+option\s*=\s*(\{[\s\S]*?\});'
                match = re.search(pattern, option_script)

                if not match:
                    fail_count += 1
                    continue

                option_json_str = match.group(1)
                option_data = self.parse_js_object(option_json_str)

                if 'series' not in option_data:
                    fail_count += 1
                    continue

                # 存储文件信息
                file_info = {
                    'path': file_path,
                    'name': os.path.basename(file_path),
                    'content': html_content,
                    'soup': soup,
                    'option': option_data,
                    'series': option_data['series']
                }

                self.html_files.append(file_info)
                success_count += 1

            except Exception as e:
                fail_count += 1
                print(f"加载文件失败 {file_path}: {str(e)}")

        # 更新文件列表显示
        self.update_file_list()

        # 如果有成功加载的文件，选中最后一个（最新上传的）
        if success_count > 0:
            last_index = len(self.html_files) - 1
            self.file_listbox.selection_clear(0, tk.END)
            self.file_listbox.selection_set(last_index)
            self.file_listbox.see(last_index)  # 滚动到可见位置
            self.load_file_by_index(last_index)

        # 显示结果（仅在状态栏显示，不弹窗）
        message = f"成功加载 {success_count} 个文件"
        if fail_count > 0:
            message += f"，失败 {fail_count} 个"

        self.status_var.set(message)

    def update_file_list(self):
        """更新文件列表显示"""
        self.file_listbox.delete(0, tk.END)
        for file_info in self.html_files:
            # 显示完整路径
            self.file_listbox.insert(tk.END, file_info['path'])

    def on_file_select(self, event=None):
        """文件列表选择事件"""
        selection = self.file_listbox.curselection()
        if selection:
            index = selection[0]
            self.load_file_by_index(index)

    def delete_selected_file(self):
        """删除选中的文件"""
        selection = self.file_listbox.curselection()
        if not selection:
            messagebox.showwarning("警告", "请先选择要删除的文件")
            return

        index = selection[0]
        file_info = self.html_files[index]

        # 确认删除
        if messagebox.askyesno("确认删除", f"确定要删除文件吗？\n{file_info['path']}"):
            # 删除文件
            del self.html_files[index]

            # 更新文件列表显示
            self.update_file_list()

            # 如果还有文件，加载第一个文件
            if self.html_files:
                if index >= len(self.html_files):
                    index = len(self.html_files) - 1
                self.load_file_by_index(index)
                self.file_listbox.selection_set(index)
            else:
                # 清空显示
                self.current_file_index = 0
                self.html_content = None
                self.soup = None
                self.option_data = None
                self.series_data = []
                self.file_path = None
                for frame in self.series_frames:
                    frame.destroy()
                self.series_frames.clear()
                self.status_var.set("就绪")

    def clear_file_list(self):
        """清空文件列表"""
        if not self.html_files:
            return

        if messagebox.askyesno("确认清空", "确定要清空所有文件吗？"):
            self.html_files.clear()
            self.update_file_list()

            # 清空显示
            self.current_file_index = 0
            self.html_content = None
            self.soup = None
            self.option_data = None
            self.series_data = []
            self.file_path = None
            for frame in self.series_frames:
                frame.destroy()
            self.series_frames.clear()
            self.status_var.set("就绪")

    def load_file_by_index(self, index):
        """根据索引加载文件数据"""
        if 0 <= index < len(self.html_files):
            self.current_file_index = index
            file_info = self.html_files[index]

            # 更新兼容属性
            self.file_path = file_info['path']
            self.html_content = file_info['content']
            self.soup = file_info['soup']
            self.option_data = file_info['option']
            self.series_data = file_info['series']

            # 显示数据
            self.display_data()
            self.status_var.set(f"当前文件: {file_info['name']}")

    def load_file(self):
        """加载HTML文件"""
        file_path = filedialog.askopenfilename(
            title="选择ECharts HTML文件",
            filetypes=[("HTML文件", "*.html"), ("所有文件", "*.*")]
        )

        if not file_path:
            return

        try:
            # 读取HTML文件
            with open(file_path, 'r', encoding='utf-8') as f:
                self.html_content = f.read()

            self.file_path = file_path

            # 解析HTML
            self.parse_html()

            # 显示数据
            self.display_data()

            self.status_var.set(f"文件加载成功: {os.path.basename(file_path)}")
            messagebox.showinfo("成功", "HTML文件加载成功！")

        except Exception as e:
            messagebox.showerror("错误", f"加载文件失败：{str(e)}")
            self.status_var.set("加载失败")

    def parse_html(self):
        """解析HTML文件，提取ECharts配置"""
        try:
            # 使用BeautifulSoup解析HTML
            self.soup = BeautifulSoup(self.html_content, 'html.parser')

            # 查找包含option的script标签
            scripts = self.soup.find_all('script')
            option_script = None

            for script in scripts:
                if script.string and 'const option' in script.string:
                    option_script = script.string
                    break

            if not option_script:
                raise ValueError("未找到ECharts配置（const option）")

            # 提取option对象的JSON字符串
            # 使用正则表达式匹配 const option = {...};
            pattern = r'const\s+option\s*=\s*(\{[\s\S]*?\});'
            match = re.search(pattern, option_script)

            if not match:
                raise ValueError("无法解析option对象")

            option_json_str = match.group(1)

            # 解析JSON（注意：JavaScript对象可能不是标准JSON，需要处理）
            self.option_data = self.parse_js_object(option_json_str)

            # 提取series数据
            if 'series' not in self.option_data:
                raise ValueError("option中未找到series数组")

            self.series_data = self.option_data['series']

        except Exception as e:
            raise Exception(f"解析HTML失败：{str(e)}")

    def parse_js_object(self, js_str):
        """将JavaScript对象字符串转换为Python字典"""
        try:
            # 尝试直接解析为JSON
            return json.loads(js_str)
        except json.JSONDecodeError:
            # 如果失败，尝试处理JavaScript特性
            try:
                # 1. 移除尾随逗号
                js_str = re.sub(r',(\s*[}\]])', r'\1', js_str)

                # 2. 给JavaScript对象的属性名添加双引号
                # 匹配模式：属性名: 值（属性名不在引号中）
                js_str = re.sub(r'(\s*)([a-zA-Z_$][a-zA-Z0-9_$]*)\s*:', r'\1"\2":', js_str)

                # 3. 处理单引号字符串（转换为双引号）
                # 注意：这个简单的替换可能不完美，但对大多数情况有效
                js_str = re.sub(r"'([^']*)'", r'"\1"', js_str)

                # 4. 再次尝试解析
                return json.loads(js_str)
            except Exception as e:
                raise ValueError(f"无法解析JavaScript对象为JSON: {str(e)}")

    def display_data(self):
        """在GUI中显示数据"""
        # 清除现有的标签页
        for frame in self.series_frames:
            frame.destroy()
        self.series_frames.clear()

        # 清空notebook
        for tab in self.notebook.tabs():
            self.notebook.forget(tab)

        # 为每个series创建一个标签页
        for idx, series in enumerate(self.series_data):
            series_name = series.get('name', f'数据序列 {idx + 1}')
            data = series.get('data', [])

            # 创建标签页框架
            frame = ttk.Frame(self.notebook)
            self.notebook.add(frame, text=series_name)
            self.series_frames.append(frame)

            # 创建表格
            self.create_data_table(frame, data, idx)

    def create_data_table(self, parent, data, series_idx):
        """创建数据表格"""
        # 创建框架容器
        container = ttk.Frame(parent)
        container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 创建Treeview表格
        columns = ('index', 'timestamp', 'datetime', 'value')
        tree = ttk.Treeview(container, columns=columns, show='headings', height=20)

        # 定义列标题
        tree.heading('index', text='序号')
        tree.heading('timestamp', text='时间戳')
        tree.heading('datetime', text='日期时间')
        tree.heading('value', text='数值')

        # 设置列宽
        tree.column('index', width=60, anchor=tk.CENTER)
        tree.column('timestamp', width=150, anchor=tk.CENTER)
        tree.column('datetime', width=180, anchor=tk.CENTER)
        tree.column('value', width=120, anchor=tk.CENTER)

        # 添加滚动条
        vsb = ttk.Scrollbar(container, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(container, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        # 布局
        tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        # 填充数据
        for i, point in enumerate(data):
            if isinstance(point, list) and len(point) >= 2:
                timestamp = point[0]
                # 确保值不是None
                if point[1] is not None:
                    value = round(point[1], 3)  # 保留3位小数
                else:
                    value = 0.0  # 如果是None，使用0

                # 转换时间戳为可读格式
                try:
                    dt = datetime.fromtimestamp(timestamp / 1000)
                    datetime_str = dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    datetime_str = "无效时间"

                tree.insert('', tk.END, values=(i + 1, timestamp, datetime_str, value),
                           tags=(series_idx, i))

        # 绑定双击事件进行编辑
        tree.bind('<Double-1>', lambda e: self.edit_cell(e, tree, series_idx))

        # 绑定右键复制事件
        tree.bind('<Button-3>', lambda e: self.copy_selection(e, tree))
        # 绑定Ctrl+C复制
        tree.bind('<Control-c>', lambda e: self.copy_selection(e, tree))
        # 绑定Ctrl+V粘贴日期
        tree.bind('<Control-v>', lambda e: self.paste_date(e, tree, series_idx))

        # 添加说明标签
        info_label = ttk.Label(container,
                              text="提示：双击日期/数值列可编辑，右键或Ctrl+C复制，Ctrl+V粘贴日期到选中行",
                              foreground='blue')
        info_label.grid(row=2, column=0, columnspan=2, pady=5)

    def copy_selection(self, event=None, tree=None):
        """复制选中的数据到剪贴板"""
        try:
            selection = tree.selection()
            if not selection:
                return

            # 获取选中行的数据
            copied_data = []
            for item in selection:
                values = tree.item(item, 'values')
                # 复制：序号、时间戳、日期时间、数值
                copied_data.append('\t'.join(str(v) for v in values))

            # 复制到剪贴板
            clipboard_text = '\n'.join(copied_data)
            self.root.clipboard_clear()
            self.root.clipboard_append(clipboard_text)

            self.status_var.set(f"已复制 {len(selection)} 行数据到剪贴板")
        except Exception as e:
            print(f"复制失败: {str(e)}")

    def paste_date(self, event=None, tree=None, series_idx=0):
        """粘贴日期到选中的行，同步更新时间戳"""
        try:
            selection = tree.selection()
            if not selection:
                self.status_var.set("请先选择要粘贴日期的行")
                return

            # 获取剪贴板内容
            try:
                clipboard_text = self.root.clipboard_get().strip()
            except tk.TclError:
                self.status_var.set("剪贴板为空")
                return

            if not clipboard_text:
                self.status_var.set("剪贴板为空")
                return

            # 解析剪贴板中的日期（支持多行粘贴）
            clipboard_lines = clipboard_text.split('\n')
            dates_to_paste = []

            for line in clipboard_lines:
                line = line.strip()
                if not line:
                    continue

                # 如果是制表符分隔的数据行（从表格复制的），尝试提取日期列
                parts = line.split('\t')
                date_str = None

                if len(parts) >= 3:
                    # 可能是完整的表格行，日期在第3列（索引2）
                    date_str = parts[2].strip()
                elif len(parts) == 1:
                    # 单独的日期字符串
                    date_str = parts[0].strip()

                if date_str:
                    # 尝试解析日期
                    dt = self._parse_date_string(date_str)
                    if dt:
                        dates_to_paste.append(dt)

            if not dates_to_paste:
                self.status_var.set("剪贴板中未找到有效的日期数据")
                return

            # 将日期粘贴到选中的行
            updated_count = 0
            selection_list = list(selection)

            for i, item in enumerate(selection_list):
                # 如果日期数量不够，循环使用最后一个日期
                date_idx = min(i, len(dates_to_paste) - 1)
                dt = dates_to_paste[date_idx]

                values = tree.item(item, 'values')
                row_idx = int(values[0]) - 1

                # 转换为毫秒时间戳
                new_timestamp = int(dt.timestamp() * 1000)
                new_datetime_str = dt.strftime('%Y-%m-%d %H:%M:%S')

                # 更新series_data中的时间戳
                self.series_data[series_idx]['data'][row_idx][0] = new_timestamp

                # 更新表格显示
                new_values = list(values)
                new_values[1] = new_timestamp
                new_values[2] = new_datetime_str
                tree.item(item, values=new_values)
                updated_count += 1

            self.status_var.set(f"已粘贴日期到 {updated_count} 行")

        except Exception as e:
            self.status_var.set(f"粘贴日期失败: {str(e)}")
            print(f"粘贴日期失败: {str(e)}")

    def _parse_date_string(self, date_str):
        """解析日期字符串，支持多种格式，返回datetime对象或None"""
        formats = (
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%d',
            '%Y/%m/%d %H:%M:%S',
            '%Y/%m/%d %H:%M',
            '%Y/%m/%d',
        )
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        return None

    def edit_cell(self, event, tree, series_idx):
        """编辑单元格（支持编辑日期和数值）"""
        # 获取选中的行和列
        selection = tree.selection()
        if not selection:
            return
        item = selection[0]
        column = tree.identify_column(event.x)

        # 允许编辑日期列(#3)和数值列(#4)
        if column not in ('#3', '#4'):
            return

        # 获取当前值
        values = tree.item(item, 'values')
        row_idx = int(values[0]) - 1

        if column == '#4':
            # 编辑数值
            self._edit_value_cell(tree, item, values, row_idx, series_idx)
        elif column == '#3':
            # 编辑日期
            self._edit_date_cell(tree, item, values, row_idx, series_idx)

    def _edit_value_cell(self, tree, item, values, row_idx, series_idx):
        """编辑数值单元格"""
        current_value = values[3]

        edit_window = tk.Toplevel(self.root)
        edit_window.title("编辑数值")
        edit_window.geometry("300x120")
        edit_window.transient(self.root)
        edit_window.grab_set()

        edit_window.update_idletasks()
        x = (edit_window.winfo_screenwidth() // 2) - (edit_window.winfo_width() // 2)
        y = (edit_window.winfo_screenheight() // 2) - (edit_window.winfo_height() // 2)
        edit_window.geometry(f"+{x}+{y}")

        ttk.Label(edit_window, text="请输入新的数值：").pack(pady=10)

        entry = ttk.Entry(edit_window, width=30)
        entry.insert(0, str(current_value))
        entry.pack(pady=5)
        entry.focus()
        entry.select_range(0, tk.END)

        def save_edit():
            try:
                new_value = float(entry.get())
                self.series_data[series_idx]['data'][row_idx][1] = new_value
                new_values = list(values)
                new_values[3] = new_value
                tree.item(item, values=new_values)
                edit_window.destroy()
                self.status_var.set(f"已更新数据：序列{series_idx + 1}，第{row_idx + 1}行")
            except ValueError:
                messagebox.showerror("错误", "请输入有效的数值！", parent=edit_window)

        def cancel_edit():
            edit_window.destroy()

        btn_frame = ttk.Frame(edit_window)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="保存", command=save_edit).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消", command=cancel_edit).pack(side=tk.LEFT, padx=5)

        entry.bind('<Return>', lambda e: save_edit())
        entry.bind('<Escape>', lambda e: cancel_edit())

    def _edit_date_cell(self, tree, item, values, row_idx, series_idx):
        """编辑日期单元格，修改日期后同步更新时间戳"""
        current_datetime = values[2]

        edit_window = tk.Toplevel(self.root)
        edit_window.title("编辑日期时间")
        edit_window.geometry("380x180")
        edit_window.transient(self.root)
        edit_window.grab_set()

        edit_window.update_idletasks()
        x = (edit_window.winfo_screenwidth() // 2) - (edit_window.winfo_width() // 2)
        y = (edit_window.winfo_screenheight() // 2) - (edit_window.winfo_height() // 2)
        edit_window.geometry(f"+{x}+{y}")

        ttk.Label(edit_window, text="请输入新的日期时间：").pack(pady=(10, 2))
        ttk.Label(edit_window, text="支持格式: YYYY-MM-DD HH:MM:SS 或 YYYY-MM-DD",
                  foreground='gray').pack(pady=(0, 5))

        entry = ttk.Entry(edit_window, width=35)
        entry.insert(0, str(current_datetime))
        entry.pack(pady=5)
        entry.focus()
        entry.select_range(0, tk.END)

        def save_edit():
            date_str = entry.get().strip()
            try:
                # 支持多种日期格式
                dt = None
                for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d',
                            '%Y/%m/%d %H:%M:%S', '%Y/%m/%d %H:%M', '%Y/%m/%d'):
                    try:
                        dt = datetime.strptime(date_str, fmt)
                        break
                    except ValueError:
                        continue

                if dt is None:
                    messagebox.showerror("错误",
                        "无法识别的日期格式！\n请使用: YYYY-MM-DD HH:MM:SS 或 YYYY-MM-DD",
                        parent=edit_window)
                    return

                # 转换为毫秒时间戳
                new_timestamp = int(dt.timestamp() * 1000)
                new_datetime_str = dt.strftime('%Y-%m-%d %H:%M:%S')

                # 更新series_data中的时间戳
                self.series_data[series_idx]['data'][row_idx][0] = new_timestamp

                # 更新表格显示（时间戳和日期时间都更新）
                new_values = list(values)
                new_values[1] = new_timestamp
                new_values[2] = new_datetime_str
                tree.item(item, values=new_values)

                edit_window.destroy()
                self.status_var.set(
                    f"已更新日期：序列{series_idx + 1}，第{row_idx + 1}行 → {new_datetime_str} (时间戳: {new_timestamp})")
            except Exception as e:
                messagebox.showerror("错误", f"更新日期失败：{str(e)}", parent=edit_window)

        def cancel_edit():
            edit_window.destroy()

        btn_frame = ttk.Frame(edit_window)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="保存", command=save_edit).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消", command=cancel_edit).pack(side=tk.LEFT, padx=5)

        entry.bind('<Return>', lambda e: save_edit())
        entry.bind('<Escape>', lambda e: cancel_edit())

    def export_excel(self):
        """导出数据为Excel文件"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("错误", "未安装openpyxl库！\n请运行: pip install openpyxl")
            return

        if not self.series_data:
            messagebox.showwarning("警告", "请先加载HTML文件！")
            return

        try:
            # 生成默认文件名，将标识符移到末尾
            base_name = os.path.basename(self.file_path)
            # 移除.html后缀
            name_without_ext = base_name.replace('.html', '')

            # 检查是否有标识符（如-l, -r等）
            # 如果文件名以 -x 结尾（x为字母或数字），将其移到最后
            match = re.match(r'(.+?)(-[a-zA-Z0-9]+)$', name_without_ext)
            if match:
                # 有标识符，格式：基础名_data标识符.xlsx
                base_part = match.group(1)
                suffix_part = match.group(2)
                default_name = f"{base_part}_data{suffix_part}.xlsx"
            else:
                # 没有标识符，正常添加
                default_name = f"{name_without_ext}_data.xlsx"

            save_path = filedialog.asksaveasfilename(
                title="导出Excel文件",
                initialfile=default_name,
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )

            if not save_path:
                return

            # 创建Excel工作簿
            wb = Workbook()
            wb.remove(wb.active)  # 删除默认工作表

            # 为每个series创建一个工作表
            for idx, series in enumerate(self.series_data):
                series_name = series.get('name', f'数据序列{idx + 1}')
                # Excel工作表名称限制
                sheet_name = series_name[:31]  # Excel工作表名最多31字符
                ws = wb.create_sheet(title=sheet_name)

                # 设置表头
                headers = ['序号', '时间戳', '日期时间', '数值']
                ws.append(headers)

                # 设置表头样式
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # 填充数据
                data = series.get('data', [])
                for i, point in enumerate(data):
                    if isinstance(point, list) and len(point) >= 2:
                        timestamp = point[0]
                        value = point[1]

                        # 转换时间戳为可读格式
                        try:
                            dt = datetime.fromtimestamp(timestamp / 1000)
                            datetime_str = dt.strftime('%Y-%m-%d %H:%M:%S')
                        except:
                            datetime_str = "无效时间"

                        ws.append([i + 1, timestamp, datetime_str, value])

                # 调整列宽
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 20
                ws.column_dimensions['D'].width = 15

                # 设置数据对齐
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            # 保存Excel文件
            wb.save(save_path)

            self.status_var.set(f"Excel文件已导出: {os.path.basename(save_path)}")
            messagebox.showinfo("成功", f"数据已导出至Excel文件：\n{save_path}\n\n请在Excel中编辑数据，然后使用'导入Excel'功能加载修改后的数据。")

        except Exception as e:
            messagebox.showerror("错误", f"导出Excel失败：{str(e)}")
            self.status_var.set("导出失败")

    def import_excel(self):
        """从Excel文件导入数据"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("错误", "未安装openpyxl库！\n请运行: pip install openpyxl")
            return

        if not self.html_content or not self.option_data:
            messagebox.showwarning("警告", "请先加载HTML文件！")
            return

        try:
            # 选择Excel文件
            file_path = filedialog.askopenfilename(
                title="选择Excel文件",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
            )

            if not file_path:
                return

            # 读取Excel文件
            wb = load_workbook(file_path)

            # 检查工作表数量是否匹配
            if len(wb.sheetnames) != len(self.series_data):
                messagebox.showwarning("警告",
                    f"Excel工作表数量({len(wb.sheetnames)})与原始数据序列数量({len(self.series_data)})不匹配！\n将尝试读取匹配的工作表。")

            # 读取每个工作表的数据
            for idx, sheet_name in enumerate(wb.sheetnames):
                if idx >= len(self.series_data):
                    break

                ws = wb[sheet_name]
                new_data = []

                # 跳过表头，从第2行开始读取
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1] is not None and row[3] is not None:  # 时间戳和数值都不为空
                        timestamp = int(row[1])
                        value = float(row[3])
                        new_data.append([timestamp, value])

                # 更新series_data
                if new_data:
                    self.series_data[idx]['data'] = new_data

            # 刷新显示
            self.display_data()

            self.status_var.set(f"Excel数据已导入: {os.path.basename(file_path)}")
            messagebox.showinfo("成功", "Excel数据导入成功！\n现在可以点击'生成新HTML'保存修改后的图表。")

        except Exception as e:
            messagebox.showerror("错误", f"导入Excel失败：{str(e)}")
            self.status_var.set("导入失败")

    def generate_html(self):
        """生成新的HTML文件"""
        if not self.html_files:
            messagebox.showwarning("警告", "请先加载HTML文件！")
            return

        try:
            # 询问用户是保存当前文件还是所有文件
            if len(self.html_files) > 1:
                result = messagebox.askyesnocancel(
                    "保存选项",
                    f"当前有 {len(self.html_files)} 个文件\n\n"
                    "是(Yes): 保存所有文件\n"
                    "否(No): 仅保存当前文件\n"
                    "取消(Cancel): 取消操作"
                )

                if result is None:  # Cancel
                    return
                elif result:  # Yes - 保存所有文件
                    self.generate_all_html()
                    return
                # No - 继续保存当前文件

            # 创建采样率设置窗口
            sample_window = tk.Toplevel(self.root)
            sample_window.title("输出设置")
            sample_window.geometry("400x150")

            frame = ttk.Frame(sample_window, padding=20)
            frame.pack(fill=tk.BOTH, expand=True)

            ttk.Label(frame, text="输出采样率:").grid(row=0, column=0, sticky=tk.W, pady=10)
            sample_rate_entry = ttk.Entry(frame, width=15)
            sample_rate_entry.grid(row=0, column=1, padx=10, pady=10)
            sample_rate_entry.insert(0, "1")

            ttk.Label(frame, text="（1=全部输出，2=每2个取1个，3=每3个取1个...）").grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=5)

            result_sample_rate = [1]  # 使用列表存储结果

            def confirm_sample():
                try:
                    rate = int(sample_rate_entry.get())
                    if rate < 1:
                        messagebox.showerror("错误", "采样率必须大于等于1！", parent=sample_window)
                        return
                    result_sample_rate[0] = rate
                    sample_window.destroy()
                except ValueError:
                    messagebox.showerror("错误", "请输入有效的数字！", parent=sample_window)
                    return

            def cancel_sample():
                result_sample_rate[0] = None
                sample_window.destroy()

            btn_frame = ttk.Frame(frame)
            btn_frame.grid(row=2, column=0, columnspan=2, pady=10)
            ttk.Button(btn_frame, text="确定", command=confirm_sample, width=10).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="取消", command=cancel_sample, width=10).pack(side=tk.LEFT, padx=5)

            sample_window.wait_window()

            if result_sample_rate[0] is None:
                return

            sample_rate = result_sample_rate[0]

            # 保存当前文件
            file_info = self.html_files[self.current_file_index]

            # 生成默认文件名
            base_name = os.path.basename(file_info['path'])
            name_without_ext = base_name.replace('.html', '')

            match = re.match(r'(.+?)(-[a-zA-Z0-9]+)$', name_without_ext)
            if match:
                base_part = match.group(1)
                suffix_part = match.group(2)
                default_name = f"{base_part}{suffix_part}.html"
            else:
                default_name = f"{name_without_ext}.html"

            save_path = filedialog.asksaveasfilename(
                title="保存新HTML文件",
                initialfile=default_name,
                defaultextension=".html",
                filetypes=[("HTML文件", "*.html"), ("所有文件", "*.*")]
            )

            if not save_path:
                return

            # 应用采样率
            if sample_rate > 1:
                import copy
                file_info = copy.deepcopy(file_info)
                for series in file_info['series']:
                    if 'data' in series:
                        series['data'] = [series['data'][idx] for idx in range(0, len(series['data']), sample_rate)]

            # 更新option_data中的series数据
            file_info['option']['series'] = file_info['series']

            # 生成新的HTML内容
            new_html = self.rebuild_html_for_file(file_info)

            # 保存文件
            with open(save_path, 'w', encoding='utf-8') as f:
                f.write(new_html)

            self.status_var.set(f"新HTML文件已保存: {os.path.basename(save_path)}")
            messagebox.showinfo("成功", f"新HTML文件已保存至：\n{save_path}")

        except Exception as e:
            messagebox.showerror("错误", f"生成HTML文件失败：{str(e)}")
            self.status_var.set("生成失败")

    def generate_all_html(self):
        """批量生成所有HTML文件"""
        try:
            # 创建采样率设置窗口
            sample_window = tk.Toplevel(self.root)
            sample_window.title("输出设置")
            sample_window.geometry("400x150")

            frame = ttk.Frame(sample_window, padding=20)
            frame.pack(fill=tk.BOTH, expand=True)

            ttk.Label(frame, text="输出采样率:").grid(row=0, column=0, sticky=tk.W, pady=10)
            sample_rate_entry = ttk.Entry(frame, width=15)
            sample_rate_entry.grid(row=0, column=1, padx=10, pady=10)
            sample_rate_entry.insert(0, "1")

            ttk.Label(frame, text="（1=全部输出，2=每2个取1个，3=每3个取1个...）").grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=5)

            result_sample_rate = [1]  # 使用列表存储结果

            def confirm_sample():
                try:
                    rate = int(sample_rate_entry.get())
                    if rate < 1:
                        messagebox.showerror("错误", "采样率必须大于等于1！", parent=sample_window)
                        return
                    result_sample_rate[0] = rate
                    sample_window.destroy()
                except ValueError:
                    messagebox.showerror("错误", "请输入有效的数字！", parent=sample_window)
                    return

            def cancel_sample():
                result_sample_rate[0] = None
                sample_window.destroy()

            btn_frame = ttk.Frame(frame)
            btn_frame.grid(row=2, column=0, columnspan=2, pady=10)
            ttk.Button(btn_frame, text="确定", command=confirm_sample, width=10).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="取消", command=cancel_sample, width=10).pack(side=tk.LEFT, padx=5)

            sample_window.wait_window()

            if result_sample_rate[0] is None:
                return

            sample_rate = result_sample_rate[0]

            # 选择保存目录
            save_dir = filedialog.askdirectory(title="选择保存目录")
            if not save_dir:
                return

            success_count = 0
            fail_count = 0

            import copy
            for file_info in self.html_files:
                try:
                    # 应用采样率
                    if sample_rate > 1:
                        file_info = copy.deepcopy(file_info)
                        for series in file_info['series']:
                            if 'data' in series:
                                series['data'] = [series['data'][idx] for idx in range(0, len(series['data']), sample_rate)]

                    # 生成文件名
                    base_name = os.path.basename(file_info['path'])
                    name_without_ext = base_name.replace('.html', '')

                    match = re.match(r'(.+?)(-[a-zA-Z0-9]+)$', name_without_ext)
                    if match:
                        base_part = match.group(1)
                        suffix_part = match.group(2)
                        new_name = f"{base_part}{suffix_part}.html"
                    else:
                        new_name = f"{name_without_ext}.html"

                    save_path = os.path.join(save_dir, new_name)

                    # 更新option_data中的series数据
                    file_info['option']['series'] = file_info['series']

                    # 生成新的HTML内容
                    new_html = self.rebuild_html_for_file(file_info)

                    # 保存文件
                    with open(save_path, 'w', encoding='utf-8') as f:
                        f.write(new_html)

                    success_count += 1

                except Exception as e:
                    fail_count += 1
                    print(f"保存文件失败 {file_info['name']}: {str(e)}")

            message = f"成功保存 {success_count} 个文件"
            if fail_count > 0:
                message += f"，失败 {fail_count} 个"

            self.status_var.set(message)
            messagebox.showinfo("批量保存完成", f"{message}\n保存目录：{save_dir}")

        except Exception as e:
            messagebox.showerror("错误", f"批量生成HTML文件失败：{str(e)}")
            self.status_var.set("批量生成失败")

    def rebuild_html_for_file(self, file_info):
        """为指定文件重建HTML内容"""
        try:
            # 重新解析原始HTML
            new_soup = BeautifulSoup(file_info['content'], 'html.parser')

            # 查找包含option的script标签
            scripts = new_soup.find_all('script')
            target_script = None

            for script in scripts:
                if script.string and 'const option' in script.string:
                    target_script = script
                    break

            if not target_script:
                raise ValueError("未找到ECharts配置脚本")

            # 获取原始脚本内容
            original_script = target_script.string

            # 提取option对象的位置
            pattern = r'(const\s+option\s*=\s*)(\{[\s\S]*?\});'
            match = re.search(pattern, original_script)

            if not match:
                raise ValueError("无法定位option对象")

            # 生成新的option JSON字符串（处理电流为0的情况）
            # 使用深拷贝避免修改原始数据
            import copy
            option_copy = copy.deepcopy(file_info['option'])

            # 将电流series中值为0的数据点改为null，避免曲线连接
            # 同时检测电压段切换，在切换处断开电流曲线（修复休眠模式bug）
            # 三重条件：只有在特定电压（14V左右）且电流小时才是休眠模式，才断开

            # 首先获取电压和电流数据
            voltage_data = None
            current_data = None

            for series in option_copy['series']:
                if series.get('name') == '电压':
                    voltage_data = series.get('data', [])
                elif series.get('name') == '电流':
                    current_data = series.get('data', [])

            # 处理电流数据
            if current_data:
                # 智能处理0值：只有连续大范围的0才改为null断开曲线
                # 定义连续0的阈值（连续多少个0才断开）
                zero_threshold = 10  # 连续10个以上的0才断开

                # 第一步：找出所有连续0的区间
                zero_ranges = []  # 存储连续0的起始和结束索引
                zero_start = None

                for i in range(len(current_data)):
                    if isinstance(current_data[i], list) and len(current_data[i]) >= 2:
                        if current_data[i][1] == 0:
                            if zero_start is None:
                                zero_start = i  # 开始一个新的0区间
                        else:
                            if zero_start is not None:
                                # 结束当前0区间
                                zero_length = i - zero_start
                                if zero_length >= zero_threshold:
                                    zero_ranges.append((zero_start, i - 1))
                                zero_start = None

                # 处理最后一个0区间（如果数据以0结尾）
                if zero_start is not None:
                    zero_length = len(current_data) - zero_start
                    if zero_length >= zero_threshold:
                        zero_ranges.append((zero_start, len(current_data) - 1))

                # 第二步：只将大范围连续0改为null
                for start, end in zero_ranges:
                    for i in range(start, end + 1):
                        if isinstance(current_data[i], list) and len(current_data[i]) >= 2:
                            current_data[i][1] = None

            # 不修改Y轴范围，保持原HTML文件的设置

            new_option_json = json.dumps(option_copy, ensure_ascii=False, indent=2)

            # 替换option对象
            new_script_content = original_script[:match.start(2)] + new_option_json + original_script[match.end(2):]

            # 更新script标签内容
            target_script.string = new_script_content

            # 返回完整的HTML
            return str(new_soup)

        except Exception as e:
            raise Exception(f"重建HTML失败：{str(e)}")

    def rebuild_html(self):
        """重建HTML内容，保持原有样式"""
        try:
            # 重新解析原始HTML（避免修改原soup对象）
            new_soup = BeautifulSoup(self.html_content, 'html.parser')

            # 查找包含option的script标签
            scripts = new_soup.find_all('script')
            target_script = None

            for script in scripts:
                if script.string and 'const option' in script.string:
                    target_script = script
                    break

            if not target_script:
                raise ValueError("未找到ECharts配置脚本")

            # 获取原始脚本内容
            original_script = target_script.string

            # 提取option对象的位置
            pattern = r'(const\s+option\s*=\s*)(\{[\s\S]*?\});'
            match = re.search(pattern, original_script)

            if not match:
                raise ValueError("无法定位option对象")

            # 生成新的option JSON字符串（保持格式化）
            new_option_json = json.dumps(self.option_data, ensure_ascii=False, indent=2)

            # 替换option对象
            new_script_content = original_script[:match.start(2)] + new_option_json + original_script[match.end(2):]

            # 更新script标签内容
            target_script.string = new_script_content

            # 返回完整的HTML
            return str(new_soup)

        except Exception as e:
            raise Exception(f"重建HTML失败：{str(e)}")

    def global_operations(self, restore_params=None):
        """全局批量操作"""
        if not self.html_files:
            messagebox.showwarning("警告", "请先加载HTML文件！")
            return

        # 创建全局操作窗口
        op_window = tk.Toplevel(self.root)
        op_window.title("全局批量操作")
        op_window.geometry("800x650")
        op_window.transient(self.root)
        op_window.grab_set()

        # 居中显示
        op_window.update_idletasks()
        x = (op_window.winfo_screenwidth() // 2) - (op_window.winfo_width() // 2)
        y = (op_window.winfo_screenheight() // 2) - (op_window.winfo_height() // 2)
        op_window.geometry(f"+{x}+{y}")

        # 标题和按钮框架
        header_frame = ttk.Frame(op_window)
        header_frame.pack(fill=tk.X, pady=10, padx=20)

        ttk.Label(header_frame, text="对所有文件的所有样品数据进行批量操作",
                 font=('Arial', 11, 'bold')).pack(side=tk.LEFT)

        # 按钮放在右上角
        btn_frame = ttk.Frame(header_frame)
        btn_frame.pack(side=tk.RIGHT)
        ttk.Button(btn_frame, text="应用", command=lambda: apply_operation(), width=12).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消", command=op_window.destroy, width=12).pack(side=tk.LEFT, padx=5)

        # 创建滚动容器
        canvas = tk.Canvas(op_window)
        scrollbar = ttk.Scrollbar(op_window, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda _: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # 绑定鼠标滚轮事件
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # 当窗口关闭时解绑滚轮事件
        def _on_closing():
            canvas.unbind_all("<MouseWheel>")
            op_window.destroy()

        op_window.protocol("WM_DELETE_WINDOW", _on_closing)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(20, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 20))

        # 作用范围选择
        scope_frame = ttk.LabelFrame(scrollable_frame, text="操作范围", padding=10)
        scope_frame.pack(fill=tk.X, padx=0, pady=5)

        scope_var = tk.StringVar(value="current")
        ttk.Radiobutton(scope_frame, text=f"仅当前文件 ({len(self.series_data)} 个样品)",
                       variable=scope_var, value="current").pack(anchor=tk.W)

        total_series = sum(len(f['series']) for f in self.html_files)
        ttk.Radiobutton(scope_frame, text=f"所有文件 ({len(self.html_files)} 个文件，共 {total_series} 个样品)",
                       variable=scope_var, value="all").pack(anchor=tk.W)

        # 操作选项框架
        options_frame = ttk.LabelFrame(scrollable_frame, text="选择操作类型", padding=10)
        options_frame.pack(fill=tk.BOTH, expand=False, padx=0, pady=10)

        # 操作类型变量
        op_type = tk.StringVar(value="replace")

        # 1. 数据替换
        replace_frame = ttk.Frame(options_frame)
        replace_frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(replace_frame, text="数据替换（用指定序号段替换）",
                       variable=op_type, value="replace").pack(side=tk.LEFT)

        # 替换参数子框架
        replace_params_frame = ttk.Frame(options_frame)
        replace_params_frame.pack(fill=tk.X, pady=2, padx=20)

        # 数据类型选择
        ttk.Label(replace_params_frame, text="替换类型:").grid(row=0, column=0, sticky=tk.W, pady=2)
        replace_type_var = tk.StringVar(value="both")
        ttk.Radiobutton(replace_params_frame, text="电压", variable=replace_type_var, value="voltage").grid(row=0, column=1, padx=2)
        ttk.Radiobutton(replace_params_frame, text="电流", variable=replace_type_var, value="current").grid(row=0, column=2, padx=2)
        ttk.Radiobutton(replace_params_frame, text="电压和电流", variable=replace_type_var, value="both").grid(row=0, column=3, padx=2)

        ttk.Label(replace_params_frame, text="源序号段:").grid(row=1, column=0, sticky=tk.W, pady=2)
        source_start_entry = ttk.Entry(replace_params_frame, width=10)
        source_start_entry.grid(row=1, column=1, padx=2)
        source_start_entry.insert(0, "1")
        ttk.Label(replace_params_frame, text="到").grid(row=0, column=2, padx=2)
        source_end_entry = ttk.Entry(replace_params_frame, width=10)
        source_end_entry.grid(row=1, column=3, padx=2)
        source_end_entry.insert(0, "10")

        def fill_source_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                source_start_entry.delete(0, tk.END)
                source_start_entry.insert(0, "1")
                source_end_entry.delete(0, tk.END)
                source_end_entry.insert(0, str(max_len))

        ttk.Button(replace_params_frame, text="全部", command=fill_source_all, width=6).grid(row=1, column=4, padx=2)

        ttk.Label(replace_params_frame, text="目标序号段:").grid(row=2, column=0, sticky=tk.W, pady=2)
        target_start_entry = ttk.Entry(replace_params_frame, width=10)
        target_start_entry.grid(row=2, column=1, padx=2)
        target_start_entry.insert(0, "100")
        ttk.Label(replace_params_frame, text="到").grid(row=1, column=2, padx=2)
        target_end_entry = ttk.Entry(replace_params_frame, width=10)
        target_end_entry.grid(row=2, column=3, padx=2)
        target_end_entry.insert(0, "109")

        def fill_target_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                target_start_entry.delete(0, tk.END)
                target_start_entry.insert(0, "1")
                target_end_entry.delete(0, tk.END)
                target_end_entry.insert(0, str(max_len))

        ttk.Button(replace_params_frame, text="全部", command=fill_target_all, width=6).grid(row=2, column=4, padx=2)

        ttk.Label(replace_params_frame, text="（目标段长度可与源段不同，自动循环补齐或截取）", foreground='gray').grid(row=3, column=0, columnspan=5, sticky=tk.W, pady=2)
        # 2. 周期电流替换
        cycle_replace_frame = ttk.Frame(options_frame)
        cycle_replace_frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(cycle_replace_frame, text="周期电流替换（用周期电流数据替换）",
                       variable=op_type, value="cycle_replace").pack(side=tk.LEFT)

        # 周期替换参数子框架
        cycle_params_frame = ttk.Frame(options_frame)
        cycle_params_frame.pack(fill=tk.X, pady=2, padx=20)

        ttk.Label(cycle_params_frame, text="周期范围:").grid(row=0, column=0, sticky=tk.W, pady=2)
        cycle_start_entry = ttk.Entry(cycle_params_frame, width=10)
        cycle_start_entry.grid(row=0, column=1, padx=2)
        cycle_start_entry.insert(0, "1")
        ttk.Label(cycle_params_frame, text="到").grid(row=0, column=2, padx=2)
        cycle_end_entry = ttk.Entry(cycle_params_frame, width=10)
        cycle_end_entry.grid(row=0, column=3, padx=2)
        cycle_end_entry.insert(0, "100")

        def fill_cycle_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                cycle_start_entry.delete(0, tk.END)
                cycle_start_entry.insert(0, "1")
                cycle_end_entry.delete(0, tk.END)
                cycle_end_entry.insert(0, str(max_len))

        ttk.Button(cycle_params_frame, text="全部", command=fill_cycle_all, width=6).grid(row=0, column=4, padx=2)

        ttk.Label(cycle_params_frame, text="替换周期范围:").grid(row=1, column=0, sticky=tk.W, pady=2)
        replace_cycle_start_entry = ttk.Entry(cycle_params_frame, width=10)
        replace_cycle_start_entry.grid(row=1, column=1, padx=2)
        replace_cycle_start_entry.insert(0, "101")
        ttk.Label(cycle_params_frame, text="到").grid(row=1, column=2, padx=2)
        replace_cycle_end_entry = ttk.Entry(cycle_params_frame, width=10)
        replace_cycle_end_entry.grid(row=1, column=3, padx=2)
        replace_cycle_end_entry.insert(0, "200")

        def fill_replace_cycle_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                replace_cycle_start_entry.delete(0, tk.END)
                replace_cycle_start_entry.insert(0, "1")
                replace_cycle_end_entry.delete(0, tk.END)
                replace_cycle_end_entry.insert(0, str(max_len))

        ttk.Button(cycle_params_frame, text="全部", command=fill_replace_cycle_all, width=6).grid(row=1, column=4, padx=2)

        ttk.Label(cycle_params_frame, text="周期电流数据（每行一个值）:").grid(row=2, column=0, sticky=tk.NW, pady=2, columnspan=4)

        # 周期电流数据文本框
        cycle_current_frame = ttk.Frame(options_frame)
        cycle_current_frame.pack(fill=tk.BOTH, expand=True, pady=2, padx=20)

        cycle_current_text = tk.Text(cycle_current_frame, height=5, width=40)
        cycle_current_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        cycle_current_scroll = ttk.Scrollbar(cycle_current_frame, orient="vertical", command=cycle_current_text.yview)
        cycle_current_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        cycle_current_text.configure(yscrollcommand=cycle_current_scroll.set)
        cycle_current_text.insert('1.0', "100\n200\n300\n400\n500")

        # 3. 周期电压替换
        cycle_voltage_replace_frame = ttk.Frame(options_frame)
        cycle_voltage_replace_frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(cycle_voltage_replace_frame, text="周期电压替换（用周期电压数据替换）",
                       variable=op_type, value="cycle_voltage_replace").pack(side=tk.LEFT)

        # 周期电压替换参数子框架
        cycle_voltage_params_frame = ttk.Frame(options_frame)
        cycle_voltage_params_frame.pack(fill=tk.X, pady=2, padx=20)

        ttk.Label(cycle_voltage_params_frame, text="周期范围:").grid(row=0, column=0, sticky=tk.W, pady=2)
        cycle_voltage_start_entry = ttk.Entry(cycle_voltage_params_frame, width=10)
        cycle_voltage_start_entry.grid(row=0, column=1, padx=2)
        cycle_voltage_start_entry.insert(0, "1")
        ttk.Label(cycle_voltage_params_frame, text="到").grid(row=0, column=2, padx=2)
        cycle_voltage_end_entry = ttk.Entry(cycle_voltage_params_frame, width=10)
        cycle_voltage_end_entry.grid(row=0, column=3, padx=2)
        cycle_voltage_end_entry.insert(0, "100")

        def fill_cycle_voltage_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                cycle_voltage_start_entry.delete(0, tk.END)
                cycle_voltage_start_entry.insert(0, "1")
                cycle_voltage_end_entry.delete(0, tk.END)
                cycle_voltage_end_entry.insert(0, str(max_len))

        ttk.Button(cycle_voltage_params_frame, text="全部", command=fill_cycle_voltage_all, width=6).grid(row=0, column=4, padx=2)

        ttk.Label(cycle_voltage_params_frame, text="替换周期范围:").grid(row=1, column=0, sticky=tk.W, pady=2)
        replace_cycle_voltage_start_entry = ttk.Entry(cycle_voltage_params_frame, width=10)
        replace_cycle_voltage_start_entry.grid(row=1, column=1, padx=2)
        replace_cycle_voltage_start_entry.insert(0, "101")
        ttk.Label(cycle_voltage_params_frame, text="到").grid(row=1, column=2, padx=2)
        replace_cycle_voltage_end_entry = ttk.Entry(cycle_voltage_params_frame, width=10)
        replace_cycle_voltage_end_entry.grid(row=1, column=3, padx=2)
        replace_cycle_voltage_end_entry.insert(0, "200")

        def fill_replace_cycle_voltage_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                replace_cycle_voltage_start_entry.delete(0, tk.END)
                replace_cycle_voltage_start_entry.insert(0, "1")
                replace_cycle_voltage_end_entry.delete(0, tk.END)
                replace_cycle_voltage_end_entry.insert(0, str(max_len))

        ttk.Button(cycle_voltage_params_frame, text="全部", command=fill_replace_cycle_voltage_all, width=6).grid(row=1, column=4, padx=2)

        ttk.Label(cycle_voltage_params_frame, text="周期电压数据（每行一个值）:").grid(row=2, column=0, sticky=tk.NW, pady=2, columnspan=4)

        # 周期电压数据文本框
        cycle_voltage_frame = ttk.Frame(options_frame)
        cycle_voltage_frame.pack(fill=tk.BOTH, expand=True, pady=2, padx=20)

        cycle_voltage_text = tk.Text(cycle_voltage_frame, height=5, width=40)
        cycle_voltage_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        cycle_voltage_scroll = ttk.Scrollbar(cycle_voltage_frame, orient="vertical", command=cycle_voltage_text.yview)
        cycle_voltage_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        cycle_voltage_text.configure(yscrollcommand=cycle_voltage_scroll.set)
        cycle_voltage_text.insert('1.0', "3.7\n3.8\n3.9\n4.0\n4.1")

        # 4. 删除序号
        delete_frame = ttk.Frame(options_frame)
        delete_frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(delete_frame, text="删除序号（删除指定序号段的数据）",
                       variable=op_type, value="delete").pack(side=tk.LEFT)

        # 删除参数子框架
        delete_params_frame = ttk.Frame(options_frame)
        delete_params_frame.pack(fill=tk.X, pady=2, padx=20)

        ttk.Label(delete_params_frame, text="删除序号段:").grid(row=0, column=0, sticky=tk.W, pady=2)
        delete_start_entry = ttk.Entry(delete_params_frame, width=10)
        delete_start_entry.grid(row=0, column=1, padx=2)
        delete_start_entry.insert(0, "1")
        ttk.Label(delete_params_frame, text="到").grid(row=0, column=2, padx=2)
        delete_end_entry = ttk.Entry(delete_params_frame, width=10)
        delete_end_entry.grid(row=0, column=3, padx=2)
        delete_end_entry.insert(0, "10")

        def fill_delete_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                delete_start_entry.delete(0, tk.END)
                delete_start_entry.insert(0, "1")
                delete_end_entry.delete(0, tk.END)
                delete_end_entry.insert(0, str(max_len))

        ttk.Button(delete_params_frame, text="全部", command=fill_delete_all, width=6).grid(row=0, column=4, padx=2)

        # 5. 电流随机波动
        fluctuation_frame = ttk.Frame(options_frame)
        fluctuation_frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(fluctuation_frame, text="电流随机波动（对电流数据添加随机波动）",
                       variable=op_type, value="current_fluctuation").pack(side=tk.LEFT)

        # 波动参数子框架
        fluctuation_params_frame = ttk.Frame(options_frame)
        fluctuation_params_frame.pack(fill=tk.X, pady=2, padx=20)

        ttk.Label(fluctuation_params_frame, text="波动范围:").grid(row=0, column=0, sticky=tk.W, pady=2)
        fluctuation_start_entry = ttk.Entry(fluctuation_params_frame, width=10)
        fluctuation_start_entry.grid(row=0, column=1, padx=2)
        fluctuation_start_entry.insert(0, "1")
        ttk.Label(fluctuation_params_frame, text="到").grid(row=0, column=2, padx=2)
        fluctuation_end_entry = ttk.Entry(fluctuation_params_frame, width=10)
        fluctuation_end_entry.grid(row=0, column=3, padx=2)
        fluctuation_end_entry.insert(0, "100")

        def fill_fluctuation_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                fluctuation_start_entry.delete(0, tk.END)
                fluctuation_start_entry.insert(0, "1")
                fluctuation_end_entry.delete(0, tk.END)
                fluctuation_end_entry.insert(0, str(max_len))

        ttk.Button(fluctuation_params_frame, text="全部", command=fill_fluctuation_all, width=6).grid(row=0, column=4, padx=2)

        # 波动类型选择
        fluctuation_type_var = tk.StringVar(value="percent")
        ttk.Label(fluctuation_params_frame, text="波动类型:").grid(row=1, column=0, sticky=tk.W, pady=2)
        ttk.Radiobutton(fluctuation_params_frame, text="随机波动（±%）", variable=fluctuation_type_var, value="percent").grid(row=1, column=1, padx=2, columnspan=2, sticky=tk.W)
        ttk.Radiobutton(fluctuation_params_frame, text="跳跃波动（±%）", variable=fluctuation_type_var, value="jump").grid(row=1, column=3, padx=2, columnspan=2, sticky=tk.W)
        ttk.Radiobutton(fluctuation_params_frame, text="固定加减", variable=fluctuation_type_var, value="fixed").grid(row=1, column=5, padx=2, columnspan=2, sticky=tk.W)

        ttk.Label(fluctuation_params_frame, text="数值:").grid(row=2, column=0, sticky=tk.W, pady=2)
        fluctuation_percent_entry = ttk.Entry(fluctuation_params_frame, width=10)
        fluctuation_percent_entry.grid(row=2, column=1, padx=2)
        fluctuation_percent_entry.insert(0, "5")

        ttk.Label(fluctuation_params_frame, text="跳跃比例(%):").grid(row=2, column=2, sticky=tk.W, pady=2, padx=(10,0))
        fluctuation_jump_ratio_entry = ttk.Entry(fluctuation_params_frame, width=10)
        fluctuation_jump_ratio_entry.grid(row=2, column=3, padx=2)
        fluctuation_jump_ratio_entry.insert(0, "30")

        ttk.Label(fluctuation_params_frame, text="（随机波动：全部数据±5%；跳跃波动：随机30%的数据±5%；固定加减：所有值+0.1，对0和<0.1A不生效）").grid(row=3, column=0, sticky=tk.W, pady=2, columnspan=7)

        # 5.5. 电压随机波动
        voltage_fluctuation_frame = ttk.Frame(options_frame)
        voltage_fluctuation_frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(voltage_fluctuation_frame, text="电压随机波动（对电压数据添加随机波动）",
                       variable=op_type, value="voltage_fluctuation").pack(side=tk.LEFT)

        # 电压波动参数子框架
        voltage_fluctuation_params_frame = ttk.Frame(options_frame)
        voltage_fluctuation_params_frame.pack(fill=tk.X, pady=2, padx=20)

        ttk.Label(voltage_fluctuation_params_frame, text="波动范围:").grid(row=0, column=0, sticky=tk.W, pady=2)
        voltage_fluctuation_start_entry = ttk.Entry(voltage_fluctuation_params_frame, width=10)
        voltage_fluctuation_start_entry.grid(row=0, column=1, padx=2)
        voltage_fluctuation_start_entry.insert(0, "1")
        ttk.Label(voltage_fluctuation_params_frame, text="到").grid(row=0, column=2, padx=2)
        voltage_fluctuation_end_entry = ttk.Entry(voltage_fluctuation_params_frame, width=10)
        voltage_fluctuation_end_entry.grid(row=0, column=3, padx=2)
        voltage_fluctuation_end_entry.insert(0, "100")

        def fill_voltage_fluctuation_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                voltage_fluctuation_start_entry.delete(0, tk.END)
                voltage_fluctuation_start_entry.insert(0, "1")
                voltage_fluctuation_end_entry.delete(0, tk.END)
                voltage_fluctuation_end_entry.insert(0, str(max_len))

        ttk.Button(voltage_fluctuation_params_frame, text="全部", command=fill_voltage_fluctuation_all, width=6).grid(row=0, column=4, padx=2)

        # 电压波动类型选择
        voltage_fluctuation_type_var = tk.StringVar(value="percent")
        ttk.Label(voltage_fluctuation_params_frame, text="波动类型:").grid(row=1, column=0, sticky=tk.W, pady=2)
        ttk.Radiobutton(voltage_fluctuation_params_frame, text="随机波动（±%）", variable=voltage_fluctuation_type_var, value="percent").grid(row=1, column=1, padx=2, columnspan=2, sticky=tk.W)
        ttk.Radiobutton(voltage_fluctuation_params_frame, text="跳跃波动（±%）", variable=voltage_fluctuation_type_var, value="jump").grid(row=1, column=3, padx=2, columnspan=2, sticky=tk.W)
        ttk.Radiobutton(voltage_fluctuation_params_frame, text="固定加减", variable=voltage_fluctuation_type_var, value="fixed").grid(row=1, column=5, padx=2, columnspan=2, sticky=tk.W)

        ttk.Label(voltage_fluctuation_params_frame, text="数值:").grid(row=2, column=0, sticky=tk.W, pady=2)
        voltage_fluctuation_percent_entry = ttk.Entry(voltage_fluctuation_params_frame, width=10)
        voltage_fluctuation_percent_entry.grid(row=2, column=1, padx=2)
        voltage_fluctuation_percent_entry.insert(0, "5")

        ttk.Label(voltage_fluctuation_params_frame, text="跳跃比例(%):").grid(row=2, column=2, sticky=tk.W, pady=2, padx=(10,0))
        voltage_fluctuation_jump_ratio_entry = ttk.Entry(voltage_fluctuation_params_frame, width=10)
        voltage_fluctuation_jump_ratio_entry.grid(row=2, column=3, padx=2)
        voltage_fluctuation_jump_ratio_entry.insert(0, "30")

        ttk.Label(voltage_fluctuation_params_frame, text="（随机波动：全部数据±5%；跳跃波动：随机30%的数据±5%；固定加减：所有值+0.1）").grid(row=3, column=0, sticky=tk.W, pady=2, columnspan=7)

        # 5.7. 批量跳跃波动（电压电流分开）
        batch_jump_frame = ttk.Frame(options_frame)
        batch_jump_frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(batch_jump_frame, text="批量跳跃波动（电压和电流分别设置跳跃波动参数）",
                       variable=op_type, value="batch_jump_fluctuation").pack(side=tk.LEFT)

        # 批量跳跃波动参数子框架
        batch_jump_params_frame = ttk.Frame(options_frame)
        batch_jump_params_frame.pack(fill=tk.X, pady=2, padx=20)

        # 电压跳跃波动参数
        ttk.Label(batch_jump_params_frame, text="【电压跳跃波动】", font=('Arial', 9, 'bold')).grid(row=0, column=0, sticky=tk.W, pady=5, columnspan=6)

        ttk.Label(batch_jump_params_frame, text="波动范围:").grid(row=1, column=0, sticky=tk.W, pady=2)
        batch_voltage_start_entry = ttk.Entry(batch_jump_params_frame, width=10)
        batch_voltage_start_entry.grid(row=1, column=1, padx=2)
        batch_voltage_start_entry.insert(0, "1")
        ttk.Label(batch_jump_params_frame, text="到").grid(row=1, column=2, padx=2)
        batch_voltage_end_entry = ttk.Entry(batch_jump_params_frame, width=10)
        batch_voltage_end_entry.grid(row=1, column=3, padx=2)
        batch_voltage_end_entry.insert(0, "100")

        def fill_batch_voltage_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                batch_voltage_start_entry.delete(0, tk.END)
                batch_voltage_start_entry.insert(0, "1")
                batch_voltage_end_entry.delete(0, tk.END)
                batch_voltage_end_entry.insert(0, str(max_len))

        ttk.Button(batch_jump_params_frame, text="全部", command=fill_batch_voltage_all, width=6).grid(row=1, column=4, padx=2)

        ttk.Label(batch_jump_params_frame, text="波动幅度(±%):").grid(row=2, column=0, sticky=tk.W, pady=2)
        batch_voltage_percent_entry = ttk.Entry(batch_jump_params_frame, width=10)
        batch_voltage_percent_entry.grid(row=2, column=1, padx=2)
        batch_voltage_percent_entry.insert(0, "5")

        ttk.Label(batch_jump_params_frame, text="跳跃比例(%):").grid(row=2, column=2, sticky=tk.W, pady=2, padx=(10,0))
        batch_voltage_jump_ratio_entry = ttk.Entry(batch_jump_params_frame, width=10)
        batch_voltage_jump_ratio_entry.grid(row=2, column=3, padx=2)
        batch_voltage_jump_ratio_entry.insert(0, "30")

        # 电流跳跃波动参数
        ttk.Label(batch_jump_params_frame, text="【电流跳跃波动】", font=('Arial', 9, 'bold')).grid(row=3, column=0, sticky=tk.W, pady=5, columnspan=6)

        ttk.Label(batch_jump_params_frame, text="波动范围:").grid(row=4, column=0, sticky=tk.W, pady=2)
        batch_current_start_entry = ttk.Entry(batch_jump_params_frame, width=10)
        batch_current_start_entry.grid(row=4, column=1, padx=2)
        batch_current_start_entry.insert(0, "1")
        ttk.Label(batch_jump_params_frame, text="到").grid(row=4, column=2, padx=2)
        batch_current_end_entry = ttk.Entry(batch_jump_params_frame, width=10)
        batch_current_end_entry.grid(row=4, column=3, padx=2)
        batch_current_end_entry.insert(0, "100")

        def fill_batch_current_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                batch_current_start_entry.delete(0, tk.END)
                batch_current_start_entry.insert(0, "1")
                batch_current_end_entry.delete(0, tk.END)
                batch_current_end_entry.insert(0, str(max_len))

        ttk.Button(batch_jump_params_frame, text="全部", command=fill_batch_current_all, width=6).grid(row=4, column=4, padx=2)

        ttk.Label(batch_jump_params_frame, text="波动幅度(±%):").grid(row=5, column=0, sticky=tk.W, pady=2)
        batch_current_percent_entry = ttk.Entry(batch_jump_params_frame, width=10)
        batch_current_percent_entry.grid(row=5, column=1, padx=2)
        batch_current_percent_entry.insert(0, "5")

        ttk.Label(batch_jump_params_frame, text="跳跃比例(%):").grid(row=5, column=2, sticky=tk.W, pady=2, padx=(10,0))
        batch_current_jump_ratio_entry = ttk.Entry(batch_jump_params_frame, width=10)
        batch_current_jump_ratio_entry.grid(row=5, column=3, padx=2)
        batch_current_jump_ratio_entry.insert(0, "30")

        ttk.Label(batch_jump_params_frame, text="（跳跃波动：随机选择指定比例的数据点进行±%波动）", foreground='gray').grid(row=6, column=0, sticky=tk.W, pady=2, columnspan=6)

        # 5.6. 电压电流削峰值
        peak_clipping_frame = ttk.Frame(options_frame)
        peak_clipping_frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(peak_clipping_frame, text="电压电流削峰值（智能替换异常峰值为左右平均值）",
                       variable=op_type, value="peak_clipping").pack(side=tk.LEFT)

        # 削峰参数子框架
        peak_clipping_params_frame = ttk.Frame(options_frame)
        peak_clipping_params_frame.pack(fill=tk.X, pady=2, padx=20)

        # 数据类型选择
        peak_clipping_type_var = tk.StringVar(value="voltage")
        ttk.Label(peak_clipping_params_frame, text="数据类型:").grid(row=0, column=0, sticky=tk.W, pady=2)
        ttk.Radiobutton(peak_clipping_params_frame, text="电压", variable=peak_clipping_type_var, value="voltage").grid(row=0, column=1, padx=2)
        ttk.Radiobutton(peak_clipping_params_frame, text="电流", variable=peak_clipping_type_var, value="current").grid(row=0, column=2, padx=2)

        ttk.Label(peak_clipping_params_frame, text="作用范围:").grid(row=1, column=0, sticky=tk.W, pady=2)
        peak_clipping_start_entry = ttk.Entry(peak_clipping_params_frame, width=10)
        peak_clipping_start_entry.grid(row=1, column=1, padx=2)
        peak_clipping_start_entry.insert(0, "1")
        ttk.Label(peak_clipping_params_frame, text="到").grid(row=1, column=2, padx=2)
        peak_clipping_end_entry = ttk.Entry(peak_clipping_params_frame, width=10)
        peak_clipping_end_entry.grid(row=1, column=3, padx=2)
        peak_clipping_end_entry.insert(0, "100")

        def fill_peak_clipping_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                peak_clipping_start_entry.delete(0, tk.END)
                peak_clipping_start_entry.insert(0, "1")
                peak_clipping_end_entry.delete(0, tk.END)
                peak_clipping_end_entry.insert(0, str(max_len))

        ttk.Button(peak_clipping_params_frame, text="全部", command=fill_peak_clipping_all, width=6).grid(row=1, column=4, padx=2)

        ttk.Label(peak_clipping_params_frame, text="最大阈值:").grid(row=2, column=0, sticky=tk.W, pady=2)
        peak_clipping_max_entry = ttk.Entry(peak_clipping_params_frame, width=10)
        peak_clipping_max_entry.grid(row=2, column=1, padx=2)
        peak_clipping_max_entry.insert(0, "5.0")
        ttk.Label(peak_clipping_params_frame, text="（高于此值将被削峰）").grid(row=2, column=2, sticky=tk.W, pady=2, columnspan=2)

        ttk.Label(peak_clipping_params_frame, text="最小阈值:").grid(row=3, column=0, sticky=tk.W, pady=2)
        peak_clipping_min_entry = ttk.Entry(peak_clipping_params_frame, width=10)
        peak_clipping_min_entry.grid(row=3, column=1, padx=2)
        peak_clipping_min_entry.insert(0, "2.0")
        ttk.Label(peak_clipping_params_frame, text="（低于此值将被削峰）").grid(row=3, column=2, sticky=tk.W, pady=2, columnspan=2)

        # 削峰方式选择
        ttk.Label(peak_clipping_params_frame, text="削峰方式:").grid(row=4, column=0, sticky=tk.W, pady=2)
        peak_clipping_method_var = tk.StringVar(value="average")
        ttk.Radiobutton(peak_clipping_params_frame, text="替换为左右平均值", variable=peak_clipping_method_var, value="average").grid(row=4, column=1, columnspan=2, sticky=tk.W, padx=2)
        ttk.Radiobutton(peak_clipping_params_frame, text="设置为0", variable=peak_clipping_method_var, value="zero").grid(row=4, column=3, columnspan=2, sticky=tk.W, padx=2)

        ttk.Label(peak_clipping_params_frame, text="（异常值处理方式）", foreground='gray').grid(row=5, column=0, columnspan=4, sticky=tk.W, pady=2)

        # 5.7. 智能电压削峰
        smart_peak_frame = ttk.Frame(options_frame)
        smart_peak_frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(smart_peak_frame, text="智能电压削峰（自动识别电压段，按目标电压±偏差范围削峰）",
                       variable=op_type, value="smart_peak_clipping").pack(side=tk.LEFT)

        # 智能削峰参数子框架
        smart_peak_params_frame = ttk.Frame(options_frame)
        smart_peak_params_frame.pack(fill=tk.X, pady=2, padx=20)

        ttk.Label(smart_peak_params_frame, text="作用范围:").grid(row=0, column=0, sticky=tk.W, pady=2)
        smart_peak_start_entry = ttk.Entry(smart_peak_params_frame, width=10)
        smart_peak_start_entry.grid(row=0, column=1, padx=2)
        smart_peak_start_entry.insert(0, "1")
        ttk.Label(smart_peak_params_frame, text="到").grid(row=0, column=2, padx=2)
        smart_peak_end_entry = ttk.Entry(smart_peak_params_frame, width=10)
        smart_peak_end_entry.grid(row=0, column=3, padx=2)
        smart_peak_end_entry.insert(0, "100")

        def fill_smart_peak_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                smart_peak_start_entry.delete(0, tk.END)
                smart_peak_start_entry.insert(0, "1")
                smart_peak_end_entry.delete(0, tk.END)
                smart_peak_end_entry.insert(0, str(max_len))

        ttk.Button(smart_peak_params_frame, text="全部", command=fill_smart_peak_all, width=6).grid(row=0, column=4, padx=2)

        ttk.Label(smart_peak_params_frame, text="允许偏差(%):").grid(row=1, column=0, sticky=tk.W, pady=2)
        smart_peak_tolerance_entry = ttk.Entry(smart_peak_params_frame, width=10)
        smart_peak_tolerance_entry.grid(row=1, column=1, padx=2)
        smart_peak_tolerance_entry.insert(0, "5")
        ttk.Label(smart_peak_params_frame, text="（例如：5表示±5%）").grid(row=1, column=2, sticky=tk.W, pady=2, columnspan=2)

        ttk.Label(smart_peak_params_frame, text="削峰方式:").grid(row=2, column=0, sticky=tk.W, pady=2)
        smart_peak_method_var = tk.StringVar(value="target")
        ttk.Radiobutton(smart_peak_params_frame, text="设为目标电压", variable=smart_peak_method_var, value="target").grid(row=2, column=1, sticky=tk.W, padx=2)
        ttk.Radiobutton(smart_peak_params_frame, text="设为边界值", variable=smart_peak_method_var, value="boundary").grid(row=2, column=2, sticky=tk.W, padx=2)
        ttk.Radiobutton(smart_peak_params_frame, text="左右平均", variable=smart_peak_method_var, value="average").grid(row=2, column=3, sticky=tk.W, padx=2)

        ttk.Label(smart_peak_params_frame, text="（自动识别14V、18V等电压段，超出±偏差范围的异常值将被削峰）", foreground='gray').grid(row=3, column=0, columnspan=5, sticky=tk.W, pady=2)

        # 5.8. 填充0值
        fill_zero_frame = ttk.Frame(options_frame)
        fill_zero_frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(fill_zero_frame, text="填充0值（将0值用线性插值替换，保持曲线连续）",
                       variable=op_type, value="fill_zero").pack(side=tk.LEFT)

        # 填充0值参数子框架
        fill_zero_params_frame = ttk.Frame(options_frame)
        fill_zero_params_frame.pack(fill=tk.X, pady=2, padx=20)

        # 数据类型选择
        fill_zero_type_var = tk.StringVar(value="both")
        ttk.Label(fill_zero_params_frame, text="数据类型:").grid(row=0, column=0, sticky=tk.W, pady=2)
        ttk.Radiobutton(fill_zero_params_frame, text="电压", variable=fill_zero_type_var, value="voltage").grid(row=0, column=1, padx=2)
        ttk.Radiobutton(fill_zero_params_frame, text="电流", variable=fill_zero_type_var, value="current").grid(row=0, column=2, padx=2)
        ttk.Radiobutton(fill_zero_params_frame, text="电压和电流", variable=fill_zero_type_var, value="both").grid(row=0, column=3, padx=2)

        ttk.Label(fill_zero_params_frame, text="作用范围:").grid(row=1, column=0, sticky=tk.W, pady=2)
        fill_zero_start_entry = ttk.Entry(fill_zero_params_frame, width=10)
        fill_zero_start_entry.grid(row=1, column=1, padx=2)
        fill_zero_start_entry.insert(0, "1")
        ttk.Label(fill_zero_params_frame, text="到").grid(row=1, column=2, padx=2)
        fill_zero_end_entry = ttk.Entry(fill_zero_params_frame, width=10)
        fill_zero_end_entry.grid(row=1, column=3, padx=2)
        fill_zero_end_entry.insert(0, "100")

        def fill_fill_zero_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                fill_zero_start_entry.delete(0, tk.END)
                fill_zero_start_entry.insert(0, "1")
                fill_zero_end_entry.delete(0, tk.END)
                fill_zero_end_entry.insert(0, str(max_len))

        ttk.Button(fill_zero_params_frame, text="全部", command=fill_fill_zero_all, width=6).grid(row=1, column=4, padx=2)

        ttk.Label(fill_zero_params_frame, text="（连续的0值将被线性插值替换，保持曲线平滑过渡）", foreground='gray').grid(row=2, column=0, columnspan=5, sticky=tk.W, pady=2)

        # 6. 数据插入
        insert_frame = ttk.Frame(options_frame)
        insert_frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(insert_frame, text="数据插入（在指定位置插入数据）",
                       variable=op_type, value="insert").pack(side=tk.LEFT)

        # 插入参数子框架
        insert_params_frame = ttk.Frame(options_frame)
        insert_params_frame.pack(fill=tk.X, pady=2, padx=20)

        ttk.Label(insert_params_frame, text="插入位置（序号后）:").grid(row=0, column=0, sticky=tk.W, pady=2)
        insert_position_entry = ttk.Entry(insert_params_frame, width=10)
        insert_position_entry.grid(row=0, column=1, padx=2)
        insert_position_entry.insert(0, "10")

        ttk.Label(insert_params_frame, text="数据列（每行一个值）:").grid(row=1, column=0, sticky=tk.NW, pady=2)

        # 创建两个文本框用于分别存储电压和电流数据
        insert_text_frame = ttk.Frame(options_frame)
        insert_text_frame.pack(fill=tk.BOTH, expand=True, pady=2, padx=20)

        # 左侧：电压数据文本框
        voltage_frame = ttk.LabelFrame(insert_text_frame, text="电压数据", padding=5)
        voltage_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

        voltage_text = tk.Text(voltage_frame, height=5, width=20)
        voltage_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        voltage_scroll = ttk.Scrollbar(voltage_frame, orient="vertical", command=voltage_text.yview)
        voltage_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        voltage_text.configure(yscrollcommand=voltage_scroll.set)
        voltage_text.insert('1.0', "3.7\n3.8\n3.9\n4.0\n4.1")

        # 右侧：电流数据文本框
        current_frame = ttk.LabelFrame(insert_text_frame, text="电流数据", padding=5)
        current_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))

        current_text = tk.Text(current_frame, height=5, width=20)
        current_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        current_scroll = ttk.Scrollbar(current_frame, orient="vertical", command=current_text.yview)
        current_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        current_text.configure(yscrollcommand=current_scroll.set)
        current_text.insert('1.0', "0.5\n0.6\n0.7\n0.8\n0.9")

        # 7. 循环复制
        cycle_copy_frame = ttk.Frame(options_frame)
        cycle_copy_frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(cycle_copy_frame, text="循环复制（循环复制周期序列的电压或电流数据）",
                       variable=op_type, value="cycle_copy").pack(side=tk.LEFT)

        # 循环复制参数子框架
        cycle_copy_params_frame = ttk.Frame(options_frame)
        cycle_copy_params_frame.pack(fill=tk.X, pady=2, padx=20)

        # 数据类型选择
        cycle_copy_type_var = tk.StringVar(value="both")
        ttk.Label(cycle_copy_params_frame, text="复制类型:").grid(row=0, column=0, sticky=tk.W, pady=2)
        ttk.Radiobutton(cycle_copy_params_frame, text="仅电压", variable=cycle_copy_type_var, value="voltage").grid(row=0, column=1, padx=2)
        ttk.Radiobutton(cycle_copy_params_frame, text="仅电流", variable=cycle_copy_type_var, value="current").grid(row=0, column=2, padx=2)
        ttk.Radiobutton(cycle_copy_params_frame, text="电压和电流", variable=cycle_copy_type_var, value="both").grid(row=0, column=3, padx=2)

        ttk.Label(cycle_copy_params_frame, text="周期序列范围:").grid(row=1, column=0, sticky=tk.W, pady=2)
        cycle_copy_source_start_entry = ttk.Entry(cycle_copy_params_frame, width=10)
        cycle_copy_source_start_entry.grid(row=1, column=1, padx=2)
        cycle_copy_source_start_entry.insert(0, "1")
        ttk.Label(cycle_copy_params_frame, text="到").grid(row=1, column=2, padx=2)
        cycle_copy_source_end_entry = ttk.Entry(cycle_copy_params_frame, width=10)
        cycle_copy_source_end_entry.grid(row=1, column=3, padx=2)
        cycle_copy_source_end_entry.insert(0, "10")

        def fill_cycle_copy_all():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                cycle_copy_source_start_entry.delete(0, tk.END)
                cycle_copy_source_start_entry.insert(0, "1")
                cycle_copy_source_end_entry.delete(0, tk.END)
                cycle_copy_source_end_entry.insert(0, str(max_len))

        ttk.Button(cycle_copy_params_frame, text="全部", command=fill_cycle_copy_all, width=6).grid(row=1, column=4, padx=2)

        # 开始日期时间选择
        ttk.Label(cycle_copy_params_frame, text="开始日期时间:").grid(row=2, column=0, sticky=tk.W, pady=2)

        start_date_frame = ttk.Frame(cycle_copy_params_frame)
        start_date_frame.grid(row=2, column=1, columnspan=5, sticky=tk.W, pady=2)

        # 年月日选择
        from datetime import datetime, timedelta
        now = datetime.now()

        # 禁用Combobox的鼠标滚轮和上下键事件的函数
        def disable_mousewheel(widget):
            def ignore_event(event):
                return "break"
            widget.bind("<MouseWheel>", ignore_event)
            widget.bind("<Up>", ignore_event)
            widget.bind("<Down>", ignore_event)

        ttk.Label(start_date_frame, text="年:").pack(side=tk.LEFT, padx=2)
        start_year_var = tk.StringVar(value=str(now.year))
        start_year_combo = ttk.Combobox(start_date_frame, textvariable=start_year_var, width=6, state='readonly')
        start_year_combo['values'] = [str(y) for y in range(2020, 2031)]
        start_year_combo.pack(side=tk.LEFT, padx=2)
        disable_mousewheel(start_year_combo)

        ttk.Label(start_date_frame, text="月:").pack(side=tk.LEFT, padx=2)
        start_month_var = tk.StringVar(value=str(now.month).zfill(2))
        start_month_combo = ttk.Combobox(start_date_frame, textvariable=start_month_var, width=4, state='readonly')
        start_month_combo['values'] = [str(m).zfill(2) for m in range(1, 13)]
        start_month_combo.pack(side=tk.LEFT, padx=2)
        disable_mousewheel(start_month_combo)

        ttk.Label(start_date_frame, text="日:").pack(side=tk.LEFT, padx=2)
        start_day_var = tk.StringVar(value=str(now.day).zfill(2))
        start_day_combo = ttk.Combobox(start_date_frame, textvariable=start_day_var, width=4, state='readonly')
        start_day_combo['values'] = [str(d).zfill(2) for d in range(1, 32)]
        start_day_combo.pack(side=tk.LEFT, padx=2)
        disable_mousewheel(start_day_combo)

        ttk.Label(start_date_frame, text="时:").pack(side=tk.LEFT, padx=2)
        start_hour_var = tk.StringVar(value=str(now.hour).zfill(2))
        start_hour_combo = ttk.Combobox(start_date_frame, textvariable=start_hour_var, width=4, state='readonly')
        start_hour_combo['values'] = [str(h).zfill(2) for h in range(0, 24)]
        start_hour_combo.pack(side=tk.LEFT, padx=2)
        disable_mousewheel(start_hour_combo)

        ttk.Label(start_date_frame, text="分:").pack(side=tk.LEFT, padx=2)
        start_minute_var = tk.StringVar(value=str(now.minute).zfill(2))
        start_minute_combo = ttk.Combobox(start_date_frame, textvariable=start_minute_var, width=4, state='readonly')
        start_minute_combo['values'] = [str(m).zfill(2) for m in range(0, 60)]
        start_minute_combo.pack(side=tk.LEFT, padx=2)
        disable_mousewheel(start_minute_combo)

        ttk.Label(start_date_frame, text="秒:").pack(side=tk.LEFT, padx=2)
        start_second_var = tk.StringVar(value=str(now.second).zfill(2))
        start_second_combo = ttk.Combobox(start_date_frame, textvariable=start_second_var, width=4, state='readonly')
        start_second_combo['values'] = [str(s).zfill(2) for s in range(0, 60)]
        start_second_combo.pack(side=tk.LEFT, padx=2)
        disable_mousewheel(start_second_combo)

        def set_start_now():
            now = datetime.now()
            start_year_var.set(str(now.year))
            start_month_var.set(str(now.month).zfill(2))
            start_day_var.set(str(now.day).zfill(2))
            start_hour_var.set(str(now.hour).zfill(2))
            start_minute_var.set(str(now.minute).zfill(2))
            start_second_var.set(str(now.second).zfill(2))

        ttk.Button(start_date_frame, text="现在", command=set_start_now, width=6).pack(side=tk.LEFT, padx=5)

        # 时间跨度
        ttk.Label(cycle_copy_params_frame, text="时间跨度:").grid(row=3, column=0, sticky=tk.W, pady=2)

        span_frame = ttk.Frame(cycle_copy_params_frame)
        span_frame.grid(row=3, column=1, columnspan=5, sticky=tk.W, pady=2)

        time_span_days_entry = ttk.Entry(span_frame, width=8)
        time_span_days_entry.pack(side=tk.LEFT, padx=2)
        time_span_days_entry.insert(0, "1")
        ttk.Label(span_frame, text="天").pack(side=tk.LEFT, padx=2)

        time_span_hours_entry = ttk.Entry(span_frame, width=8)
        time_span_hours_entry.pack(side=tk.LEFT, padx=2)
        time_span_hours_entry.insert(0, "0")
        ttk.Label(span_frame, text="小时").pack(side=tk.LEFT, padx=2)

        # 结束日期时间选择
        ttk.Label(cycle_copy_params_frame, text="结束日期时间:").grid(row=4, column=0, sticky=tk.W, pady=2)

        end_date_frame = ttk.Frame(cycle_copy_params_frame)
        end_date_frame.grid(row=4, column=1, columnspan=5, sticky=tk.W, pady=2)

        end_tomorrow = now + timedelta(days=1)

        ttk.Label(end_date_frame, text="年:").pack(side=tk.LEFT, padx=2)
        end_year_var = tk.StringVar(value=str(end_tomorrow.year))
        end_year_combo = ttk.Combobox(end_date_frame, textvariable=end_year_var, width=6, state='readonly')
        end_year_combo['values'] = [str(y) for y in range(2020, 2031)]
        end_year_combo.pack(side=tk.LEFT, padx=2)
        disable_mousewheel(end_year_combo)

        ttk.Label(end_date_frame, text="月:").pack(side=tk.LEFT, padx=2)
        end_month_var = tk.StringVar(value=str(end_tomorrow.month).zfill(2))
        end_month_combo = ttk.Combobox(end_date_frame, textvariable=end_month_var, width=4, state='readonly')
        end_month_combo['values'] = [str(m).zfill(2) for m in range(1, 13)]
        end_month_combo.pack(side=tk.LEFT, padx=2)
        disable_mousewheel(end_month_combo)

        ttk.Label(end_date_frame, text="日:").pack(side=tk.LEFT, padx=2)
        end_day_var = tk.StringVar(value=str(end_tomorrow.day).zfill(2))
        end_day_combo = ttk.Combobox(end_date_frame, textvariable=end_day_var, width=4, state='readonly')
        end_day_combo['values'] = [str(d).zfill(2) for d in range(1, 32)]
        end_day_combo.pack(side=tk.LEFT, padx=2)
        disable_mousewheel(end_day_combo)

        ttk.Label(end_date_frame, text="时:").pack(side=tk.LEFT, padx=2)
        end_hour_var = tk.StringVar(value=str(end_tomorrow.hour).zfill(2))
        end_hour_combo = ttk.Combobox(end_date_frame, textvariable=end_hour_var, width=4, state='readonly')
        end_hour_combo['values'] = [str(h).zfill(2) for h in range(0, 24)]
        end_hour_combo.pack(side=tk.LEFT, padx=2)
        disable_mousewheel(end_hour_combo)

        ttk.Label(end_date_frame, text="分:").pack(side=tk.LEFT, padx=2)
        end_minute_var = tk.StringVar(value=str(end_tomorrow.minute).zfill(2))
        end_minute_combo = ttk.Combobox(end_date_frame, textvariable=end_minute_var, width=4, state='readonly')
        end_minute_combo['values'] = [str(m).zfill(2) for m in range(0, 60)]
        end_minute_combo.pack(side=tk.LEFT, padx=2)
        disable_mousewheel(end_minute_combo)

        ttk.Label(end_date_frame, text="秒:").pack(side=tk.LEFT, padx=2)
        end_second_var = tk.StringVar(value=str(end_tomorrow.second).zfill(2))
        end_second_combo = ttk.Combobox(end_date_frame, textvariable=end_second_var, width=4, state='readonly')
        end_second_combo['values'] = [str(s).zfill(2) for s in range(0, 60)]
        end_second_combo.pack(side=tk.LEFT, padx=2)
        ttk.Label(end_date_frame, text="秒:").pack(side=tk.LEFT, padx=2)
        end_second_var = tk.StringVar(value=str(end_tomorrow.second).zfill(2))
        end_second_combo = ttk.Combobox(end_date_frame, textvariable=end_second_var, width=4, state='readonly')
        end_second_combo['values'] = [str(s).zfill(2) for s in range(0, 60)]
        end_second_combo.pack(side=tk.LEFT, padx=2)
        disable_mousewheel(end_second_combo)

        # 自动计算结束时间按钮
        def calc_end_time():
            try:
                start_dt = datetime(
                    int(start_year_var.get()),
                    int(start_month_var.get()),
                    int(start_day_var.get()),
                    int(start_hour_var.get()),
                    int(start_minute_var.get()),
                    int(start_second_var.get())
                )

                days = int(time_span_days_entry.get())
                hours = int(time_span_hours_entry.get())

                end_dt = start_dt + timedelta(days=days, hours=hours)

                end_year_var.set(str(end_dt.year))
                end_month_var.set(str(end_dt.month).zfill(2))
                end_day_var.set(str(end_dt.day).zfill(2))
                end_hour_var.set(str(end_dt.hour).zfill(2))
                end_minute_var.set(str(end_dt.minute).zfill(2))
                end_second_var.set(str(end_dt.second).zfill(2))
            except Exception as e:
                messagebox.showerror("错误", f"计算失败：{str(e)}", parent=op_window)

        ttk.Button(end_date_frame, text="自动计算", command=calc_end_time, width=10).pack(side=tk.LEFT, padx=5)

        # 快捷预设按钮
        def set_preset_24h():
            now = datetime.now()
            start_year_var.set(str(now.year))
            start_month_var.set(str(now.month).zfill(2))
            start_day_var.set(str(now.day).zfill(2))
            start_hour_var.set(str(now.hour).zfill(2))
            start_minute_var.set(str(now.minute).zfill(2))
            start_second_var.set(str(now.second).zfill(2))

            end = now + timedelta(days=1)
            end_year_var.set(str(end.year))
            end_month_var.set(str(end.month).zfill(2))
            end_day_var.set(str(end.day).zfill(2))
            end_hour_var.set(str(end.hour).zfill(2))
            end_minute_var.set(str(end.minute).zfill(2))
            end_second_var.set(str(end.second).zfill(2))

        def set_preset_7d():
            now = datetime.now()
            start_year_var.set(str(now.year))
            start_month_var.set(str(now.month).zfill(2))
            start_day_var.set(str(now.day).zfill(2))
            start_hour_var.set(str(now.hour).zfill(2))
            start_minute_var.set(str(now.minute).zfill(2))
            start_second_var.set(str(now.second).zfill(2))

            end = now + timedelta(days=7)
            end_year_var.set(str(end.year))
            end_month_var.set(str(end.month).zfill(2))
            end_day_var.set(str(end.day).zfill(2))
            end_hour_var.set(str(end.hour).zfill(2))
            end_minute_var.set(str(end.minute).zfill(2))
            end_second_var.set(str(end.second).zfill(2))

        preset_frame = ttk.Frame(cycle_copy_params_frame)
        preset_frame.grid(row=5, column=0, columnspan=6, pady=5)
        ttk.Label(preset_frame, text="快捷预设:", foreground='blue').pack(side=tk.LEFT, padx=5)
        ttk.Button(preset_frame, text="24小时", command=set_preset_24h, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(preset_frame, text="7天", command=set_preset_7d, width=8).pack(side=tk.LEFT, padx=2)

        ttk.Label(cycle_copy_params_frame, text="最大数据点数:").grid(row=6, column=0, sticky=tk.W, pady=2)
        cycle_copy_max_points_entry = ttk.Entry(cycle_copy_params_frame, width=15)
        cycle_copy_max_points_entry.grid(row=6, column=1, padx=2)
        cycle_copy_max_points_entry.insert(0, "100000")
        ttk.Label(cycle_copy_params_frame, text="（超过此数量将自动调整时间间隔）", foreground='gray').grid(row=6, column=2, columnspan=4, sticky=tk.W, pady=2)

        ttk.Label(cycle_copy_params_frame, text="（提示：点击下拉框选择日期时间，或使用快捷预设按钮）", foreground='blue').grid(row=7, column=0, columnspan=6, sticky=tk.W, pady=2)

        # 8. Y轴范围设置
        yaxis_range_frame = ttk.Frame(options_frame)
        yaxis_range_frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(yaxis_range_frame, text="Y轴范围设置（设置图表Y轴的最大和最小刻度值）",
                       variable=op_type, value="yaxis_range").pack(side=tk.LEFT)

        # Y轴范围参数子框架
        yaxis_range_params_frame = ttk.Frame(options_frame)
        yaxis_range_params_frame.pack(fill=tk.X, pady=2, padx=20)

        # 数据类型选择
        yaxis_range_type_var = tk.StringVar(value="both")
        ttk.Label(yaxis_range_params_frame, text="应用到:").grid(row=0, column=0, sticky=tk.W, pady=2)
        ttk.Radiobutton(yaxis_range_params_frame, text="电压Y轴", variable=yaxis_range_type_var, value="voltage").grid(row=0, column=1, padx=2)
        ttk.Radiobutton(yaxis_range_params_frame, text="电流Y轴", variable=yaxis_range_type_var, value="current").grid(row=0, column=2, padx=2)
        ttk.Radiobutton(yaxis_range_params_frame, text="两者", variable=yaxis_range_type_var, value="both").grid(row=0, column=3, padx=2)

        ttk.Label(yaxis_range_params_frame, text="最小值:").grid(row=1, column=0, sticky=tk.W, pady=2)
        yaxis_min_entry = ttk.Entry(yaxis_range_params_frame, width=15)
        yaxis_min_entry.grid(row=1, column=1, padx=2)
        yaxis_min_entry.insert(0, "0")
        ttk.Label(yaxis_range_params_frame, text="（留空表示自动）").grid(row=1, column=2, sticky=tk.W, pady=2, columnspan=2)

        ttk.Label(yaxis_range_params_frame, text="最大值:").grid(row=2, column=0, sticky=tk.W, pady=2)
        yaxis_max_entry = ttk.Entry(yaxis_range_params_frame, width=15)
        yaxis_max_entry.grid(row=2, column=1, padx=2)
        yaxis_max_entry.insert(0, "1")
        ttk.Label(yaxis_range_params_frame, text="（留空表示自动）").grid(row=2, column=2, sticky=tk.W, pady=2, columnspan=2)

        ttk.Label(yaxis_range_params_frame, text="（例如：电流值为μA时，可设置最大值为1或更小的值）", foreground='gray').grid(row=3, column=0, columnspan=4, sticky=tk.W, pady=2)

        # 3. 偷天换日（时间戳调整）
        time_shift_frame = ttk.Frame(options_frame)
        time_shift_frame.pack(fill=tk.X, pady=5)
        ttk.Radiobutton(time_shift_frame, text="偷天换日（调整时间戳）",
                       variable=op_type, value="time_shift").pack(side=tk.LEFT)

        # 时间调整参数子框架
        time_shift_params_frame = ttk.Frame(options_frame)
        time_shift_params_frame.pack(fill=tk.X, pady=2, padx=20)

        # 方式1：直接输入目标日期时间
        ttk.Label(time_shift_params_frame, text="方式1 - 设置开始日期时间:").grid(row=0, column=0, sticky=tk.W, pady=2, columnspan=4)

        ttk.Label(time_shift_params_frame, text="目标日期时间:").grid(row=1, column=0, sticky=tk.W, pady=2)
        target_datetime_entry = ttk.Entry(time_shift_params_frame, width=25)
        target_datetime_entry.grid(row=1, column=1, padx=2, columnspan=3)
        target_datetime_entry.insert(0, "2025-01-01 00:00:00")
        ttk.Label(time_shift_params_frame, text="（格式：YYYY-MM-DD HH:MM:SS）", foreground='gray').grid(row=1, column=4, sticky=tk.W, pady=2, columnspan=4)

        ttk.Separator(time_shift_params_frame, orient=tk.HORIZONTAL).grid(row=2, column=0, columnspan=8, sticky='ew', pady=10)

        # 方式2：手动调整时间偏移
        ttk.Label(time_shift_params_frame, text="方式2 - 手动调整偏移量:").grid(row=3, column=0, sticky=tk.W, pady=2, columnspan=4)

        ttk.Label(time_shift_params_frame, text="时间调整（dd:hh:mm:ss）:").grid(row=4, column=0, sticky=tk.W, pady=2)

        ttk.Label(time_shift_params_frame, text="天:").grid(row=4, column=1, padx=2)
        time_shift_days_entry = ttk.Entry(time_shift_params_frame, width=8)
        time_shift_days_entry.grid(row=4, column=2, padx=2)
        time_shift_days_entry.insert(0, "0")

        ttk.Label(time_shift_params_frame, text="时:").grid(row=4, column=3, padx=2)
        time_shift_hours_entry = ttk.Entry(time_shift_params_frame, width=8)
        time_shift_hours_entry.grid(row=4, column=4, padx=2)
        time_shift_hours_entry.insert(0, "0")

        ttk.Label(time_shift_params_frame, text="分:").grid(row=4, column=5, padx=2)
        time_shift_minutes_entry = ttk.Entry(time_shift_params_frame, width=8)
        time_shift_minutes_entry.grid(row=4, column=6, padx=2)
        time_shift_minutes_entry.insert(0, "0")

        ttk.Label(time_shift_params_frame, text="秒:").grid(row=4, column=7, padx=2)
        time_shift_seconds_entry = ttk.Entry(time_shift_params_frame, width=8)
        time_shift_seconds_entry.grid(row=4, column=8, padx=2)
        time_shift_seconds_entry.insert(0, "0")

        ttk.Label(time_shift_params_frame, text="（正数向后，负数向前）", foreground='gray').grid(row=5, column=0, columnspan=9, sticky=tk.W, pady=2)

        # 智能生成按钮框架
        smart_gen_frame = ttk.Frame(options_frame)
        smart_gen_frame.pack(fill=tk.X, pady=5, padx=20)

        ttk.Label(smart_gen_frame, text="智能生成:").grid(row=0, column=0, padx=2, sticky=tk.W)

        ttk.Label(smart_gen_frame, text="目标值:").grid(row=0, column=1, padx=2)
        smart_target_entry = ttk.Entry(smart_gen_frame, width=10)
        smart_target_entry.grid(row=0, column=2, padx=2)
        smart_target_entry.insert(0, "3.7")

        ttk.Label(smart_gen_frame, text="生成数量:").grid(row=0, column=3, padx=2)
        smart_count_entry = ttk.Entry(smart_gen_frame, width=10)
        smart_count_entry.grid(row=0, column=4, padx=2)
        smart_count_entry.insert(0, "14")

        # 选择生成类型（电压或电流）
        series_type_var = tk.StringVar(value="voltage")
        ttk.Label(smart_gen_frame, text="数据类型:").grid(row=1, column=0, padx=2, sticky=tk.W, pady=5)
        ttk.Radiobutton(smart_gen_frame, text="电压", variable=series_type_var, value="voltage").grid(row=1, column=1, padx=2)
        ttk.Radiobutton(smart_gen_frame, text="电流", variable=series_type_var, value="current").grid(row=1, column=2, padx=2)

        def smart_generate():
            """智能生成伪数据 - 使用高级波动算法分析历史数据"""
            try:
                target_value = float(smart_target_entry.get())
                gen_count = int(smart_count_entry.get())
                series_type = series_type_var.get()

                if gen_count <= 0:
                    messagebox.showerror("错误", "生成数量必须大于0！", parent=op_window)
                    return

                # 获取当前文件的对应series数据
                if not self.html_files or not self.html_files[self.current_file_index]['series']:
                    messagebox.showerror("错误", "没有可用的历史数据！", parent=op_window)
                    return

                # 根据类型选择series（假设第一个是电压，第二个是电流）
                current_file_series = self.html_files[self.current_file_index]['series']
                if len(current_file_series) < 2:
                    messagebox.showerror("错误", "数据不完整，需要至少2个series（电压和电流）！", parent=op_window)
                    return

                # 选择对应的series
                if series_type == "voltage":
                    target_series = current_file_series[0]  # 电压
                else:
                    target_series = current_file_series[1]  # 电流

                data = target_series['data']

                if len(data) < 3:
                    messagebox.showerror("错误", "历史数据太少，至少需要3个数据点！", parent=op_window)
                    return

                # 提取所有数值
                values = [point[1] for point in data]

                # 使用智能波动算法生成数据
                if SMART_FLUCTUATION_AVAILABLE:
                    # 使用高级算法
                    generated_values = analyze_and_generate(
                        historical_data=values,
                        target_value=target_value,
                        count=gen_count,
                        mode="balanced"  # 可选: smooth, balanced, volatile
                    )

                    # 获取生成器以显示统计信息
                    generator = SmartFluctuationGenerator(values)
                    info_msg = (
                        f"已生成 {gen_count} 个{'电压' if series_type == 'voltage' else '电流'}数据点\n"
                        f"目标值: {target_value:.4f}\n"
                        f"算法: 智能波动算法\n"
                        f"检测到的趋势: {generator.trend:.6f}\n"
                        f"自相关系数: {generator.autocorr:.4f}\n"
                        f"波动率: {generator.volatility:.4f}"
                    )
                else:
                    # 回退到简单算法
                    import random

                    # 找到目标值附近的数据点
                    tolerance = abs(target_value) * 0.1
                    nearby_values = [v for v in values if abs(v - target_value) <= tolerance]

                    if len(nearby_values) < 5:
                        tolerance = abs(target_value) * 0.2
                        nearby_values = [v for v in values if abs(v - target_value) <= tolerance]

                    if len(nearby_values) < 5:
                        nearby_values = values

                    # 计算统计特征
                    nearby_mean = sum(nearby_values) / len(nearby_values)
                    nearby_variance = sum((x - nearby_mean) ** 2 for x in nearby_values) / len(nearby_values)
                    nearby_std = nearby_variance ** 0.5
                    nearby_range = max(nearby_values) - min(nearby_values)

                    # 简单生成
                    generated_values = []
                    for _ in range(gen_count):
                        noise = random.uniform(-nearby_range * 0.2, nearby_range * 0.2)
                        trend = random.uniform(-nearby_std * 0.1, nearby_std * 0.1)
                        new_value = target_value + noise + trend
                        generated_values.append(round(new_value, 3))

                    info_msg = (
                        f"已生成 {gen_count} 个{'电压' if series_type == 'voltage' else '电流'}数据点\n"
                        f"目标值: {target_value:.4f}\n"
                        f"算法: 基础随机算法\n"
                        f"参考数据: {len(nearby_values)} 个附近点\n"
                        f"波动范围: ±{nearby_range/2:.4f}"
                    )

                # 根据类型填充到对应的文本框
                if series_type == "voltage":
                    voltage_text.delete('1.0', tk.END)
                    voltage_text.insert('1.0', '\n'.join(str(v) for v in generated_values))
                else:
                    current_text.delete('1.0', tk.END)
                    current_text.insert('1.0', '\n'.join(str(v) for v in generated_values))

                messagebox.showinfo("成功", info_msg, parent=op_window)

            except ValueError:
                messagebox.showerror("错误", "请输入有效的数值！", parent=op_window)
            except Exception as e:
                messagebox.showerror("错误", f"生成失败：{str(e)}", parent=op_window)

        ttk.Button(smart_gen_frame, text="生成", command=smart_generate).grid(row=1, column=3, columnspan=2, padx=5, pady=5)

        def undo_operation():
            """撤回上一次操作"""
            if self.operation_backup is None:
                messagebox.showwarning("提示", "没有可撤回的操作！", parent=op_window)
                return

            try:
                import copy
                # 恢复备份数据
                scope = scope_var.get()
                if scope == "current":
                    # 恢复当前文件
                    self.html_files[self.current_file_index] = copy.deepcopy(self.operation_backup[0])
                else:
                    # 恢复所有文件
                    self.html_files = copy.deepcopy(self.operation_backup)

                # 刷新显示
                self.load_file_by_index(self.current_file_index)
                current_file = self.html_files[self.current_file_index]
                self.series_data = current_file['series']
                self.option_data = current_file['option']

                self.status_var.set("已撤回上一次操作")
                messagebox.showinfo("成功", "已撤回上一次操作！", parent=op_window)

            except Exception as e:
                messagebox.showerror("错误", f"撤回失败：{str(e)}", parent=op_window)

        ttk.Button(smart_gen_frame, text="撤回", command=undo_operation).grid(row=1, column=5, columnspan=2, padx=5, pady=5)

        # 如果有恢复参数，则恢复上次的设置
        if restore_params:
            # 恢复操作类型和作用域
            op_type.set(restore_params.get('operation', 'replace'))
            scope_var.set(restore_params.get('scope', 'all'))

            # 恢复数据替换参数
            source_start_entry.delete(0, tk.END)
            source_start_entry.insert(0, restore_params.get('source_start', '1'))
            source_end_entry.delete(0, tk.END)
            source_end_entry.insert(0, restore_params.get('source_end', '10'))
            target_start_entry.delete(0, tk.END)
            target_start_entry.insert(0, restore_params.get('target_start', '11'))
            target_end_entry.delete(0, tk.END)
            target_end_entry.insert(0, restore_params.get('target_end', '20'))

            # 恢复周期电流替换参数
            cycle_start_entry.delete(0, tk.END)
            cycle_start_entry.insert(0, restore_params.get('cycle_start', '1'))
            cycle_end_entry.delete(0, tk.END)
            cycle_end_entry.insert(0, restore_params.get('cycle_end', '10'))
            replace_cycle_start_entry.delete(0, tk.END)
            replace_cycle_start_entry.insert(0, restore_params.get('replace_cycle_start', '11'))
            replace_cycle_end_entry.delete(0, tk.END)
            replace_cycle_end_entry.insert(0, restore_params.get('replace_cycle_end', '20'))
            cycle_current_text.delete('1.0', tk.END)
            cycle_current_text.insert('1.0', restore_params.get('cycle_current_text', ''))

            # 恢复周期电压替换参数
            cycle_voltage_start_entry.delete(0, tk.END)
            cycle_voltage_start_entry.insert(0, restore_params.get('cycle_voltage_start', '1'))
            cycle_voltage_end_entry.delete(0, tk.END)
            cycle_voltage_end_entry.insert(0, restore_params.get('cycle_voltage_end', '10'))
            replace_cycle_voltage_start_entry.delete(0, tk.END)
            replace_cycle_voltage_start_entry.insert(0, restore_params.get('replace_cycle_voltage_start', '11'))
            replace_cycle_voltage_end_entry.delete(0, tk.END)
            replace_cycle_voltage_end_entry.insert(0, restore_params.get('replace_cycle_voltage_end', '20'))
            cycle_voltage_text.delete('1.0', tk.END)
            cycle_voltage_text.insert('1.0', restore_params.get('cycle_voltage_text', ''))

            # 恢复删除参数
            delete_start_entry.delete(0, tk.END)
            delete_start_entry.insert(0, restore_params.get('delete_start', '1'))
            delete_end_entry.delete(0, tk.END)
            delete_end_entry.insert(0, restore_params.get('delete_end', '10'))

            # 恢复电流波动参数
            fluctuation_start_entry.delete(0, tk.END)
            fluctuation_start_entry.insert(0, restore_params.get('fluctuation_start', '1'))
            fluctuation_end_entry.delete(0, tk.END)
            fluctuation_end_entry.insert(0, restore_params.get('fluctuation_end', '100'))
            fluctuation_percent_entry.delete(0, tk.END)
            fluctuation_percent_entry.insert(0, restore_params.get('fluctuation_percent', '5'))
            fluctuation_type_var.set(restore_params.get('fluctuation_type', 'percent'))

            # 恢复电压波动参数
            voltage_fluctuation_start_entry.delete(0, tk.END)
            voltage_fluctuation_start_entry.insert(0, restore_params.get('voltage_fluctuation_start', '1'))
            voltage_fluctuation_end_entry.delete(0, tk.END)
            voltage_fluctuation_end_entry.insert(0, restore_params.get('voltage_fluctuation_end', '100'))
            voltage_fluctuation_percent_entry.delete(0, tk.END)
            voltage_fluctuation_percent_entry.insert(0, restore_params.get('voltage_fluctuation_percent', '5'))
            voltage_fluctuation_type_var.set(restore_params.get('voltage_fluctuation_type', 'percent'))

            # 恢复削峰参数
            peak_clipping_type_var.set(restore_params.get('peak_clipping_type', 'voltage'))
            peak_clipping_start_entry.delete(0, tk.END)
            peak_clipping_start_entry.insert(0, restore_params.get('peak_clipping_start', '1'))
            peak_clipping_end_entry.delete(0, tk.END)
            peak_clipping_end_entry.insert(0, restore_params.get('peak_clipping_end', '100'))
            peak_clipping_max_entry.delete(0, tk.END)
            peak_clipping_max_entry.insert(0, restore_params.get('peak_clipping_max', '5.0'))
            peak_clipping_min_entry.delete(0, tk.END)
            peak_clipping_min_entry.insert(0, restore_params.get('peak_clipping_min', '0.0'))

            # 恢复时间调整参数
            target_datetime_entry.delete(0, tk.END)
            target_datetime_entry.insert(0, restore_params.get('target_datetime', '2025-01-01 00:00:00'))
            time_shift_days_entry.delete(0, tk.END)
            time_shift_days_entry.insert(0, restore_params.get('time_shift_days', '0'))
            time_shift_hours_entry.delete(0, tk.END)
            time_shift_hours_entry.insert(0, restore_params.get('time_shift_hours', '0'))
            time_shift_minutes_entry.delete(0, tk.END)
            time_shift_minutes_entry.insert(0, restore_params.get('time_shift_minutes', '0'))
            time_shift_seconds_entry.delete(0, tk.END)
            time_shift_seconds_entry.insert(0, restore_params.get('time_shift_seconds', '0'))

            # 恢复循环复制参数
            cycle_copy_type_var.set(restore_params.get('cycle_copy_type', 'both'))
            cycle_copy_source_start_entry.delete(0, tk.END)
            cycle_copy_source_start_entry.insert(0, restore_params.get('cycle_copy_source_start', '1'))
            cycle_copy_source_end_entry.delete(0, tk.END)
            cycle_copy_source_end_entry.insert(0, restore_params.get('cycle_copy_source_end', '10'))

            # 恢复开始时间
            start_datetime_str = restore_params.get('cycle_copy_start_datetime', '2025-01-01 00:00:00')
            try:
                start_dt = datetime.strptime(start_datetime_str, "%Y-%m-%d %H:%M:%S")
                start_year_var.set(str(start_dt.year))
                start_month_var.set(str(start_dt.month).zfill(2))
                start_day_var.set(str(start_dt.day).zfill(2))
                start_hour_var.set(str(start_dt.hour).zfill(2))
                start_minute_var.set(str(start_dt.minute).zfill(2))
                start_second_var.set(str(start_dt.second).zfill(2))
            except:
                pass

            # 恢复结束时间
            end_datetime_str = restore_params.get('cycle_copy_end_datetime', '2025-01-02 00:00:00')
            try:
                end_dt = datetime.strptime(end_datetime_str, "%Y-%m-%d %H:%M:%S")
                end_year_var.set(str(end_dt.year))
                end_month_var.set(str(end_dt.month).zfill(2))
                end_day_var.set(str(end_dt.day).zfill(2))
                end_hour_var.set(str(end_dt.hour).zfill(2))
                end_minute_var.set(str(end_dt.minute).zfill(2))
                end_second_var.set(str(end_dt.second).zfill(2))
            except:
                pass

            cycle_copy_max_points_entry.delete(0, tk.END)
            cycle_copy_max_points_entry.insert(0, restore_params.get('cycle_copy_max_points', '100000'))

            # 恢复数据插入参数
            insert_position_entry.delete(0, tk.END)
            insert_position_entry.insert(0, restore_params.get('insert_position', '1'))
            series_type_var.set(restore_params.get('series_type', 'voltage'))
            voltage_text.delete('1.0', tk.END)
            voltage_text.insert('1.0', restore_params.get('voltage_text', ''))
            current_text.delete('1.0', tk.END)
            current_text.insert('1.0', restore_params.get('current_text', ''))

            # 恢复Y轴范围参数
            yaxis_range_type_var.set(restore_params.get('yaxis_range_type', 'both'))
            yaxis_min_entry.delete(0, tk.END)
            yaxis_min_entry.insert(0, restore_params.get('yaxis_min', '0'))
            yaxis_max_entry.delete(0, tk.END)
            yaxis_max_entry.insert(0, restore_params.get('yaxis_max', '1'))

        def apply_operation():
            try:
                operation = op_type.get()
                scope = scope_var.get()

                # 确定要处理的文件列表
                if scope == "current":
                    files_to_process = [self.html_files[self.current_file_index]]
                else:
                    files_to_process = self.html_files

                # 备份当前数据（深拷贝）
                import copy
                self.operation_backup = copy.deepcopy(files_to_process)

                if operation == "replace":
                    # 获取序号范围
                    source_start = int(source_start_entry.get()) - 1  # 转为0索引
                    source_end = int(source_end_entry.get()) - 1
                    target_start = int(target_start_entry.get()) - 1
                    target_end = int(target_end_entry.get()) - 1
                    replace_type = replace_type_var.get()  # 获取替换类型

                    # 验证参数
                    if source_start < 0 or source_end < source_start:
                        messagebox.showerror("错误", "源序号段无效！", parent=op_window)
                        return
                    if target_start < 0 or target_end < target_start:
                        messagebox.showerror("错误", "目标序号段无效！", parent=op_window)
                        return
              # 计算源段和目标段的长度
                    source_length = source_end - source_start + 1
                    target_length = target_end - target_start + 1

                    # 对所有文件的所有series进行数据替换
                    replaced_count = 0
                    extended_count = 0
                    for file_info in files_to_process:
                        for series in file_info['series']:
                            data = series['data']

                            # 检查源序号是否有效
                            if source_end >= len(data):
                                continue

                            # 如果目标序号超出范围，需要扩展数据
                            if target_end >= len(data):
                                # 计算需要扩展的数据点数量
                                extend_count = target_end - len(data) + 1

                                # 获取最后一个数据点的时间戳
                                last_timestamp = data[-1][0]

                                # 计算时间间隔（使用最后两个点的间隔）
                                if len(data) >= 2:
                                    time_interval = data[-1][0] - data[-2][0]
                                else:
                                    time_interval = 1000  # 默认1秒

                                # 扩展数据点（使用最后一个点的值）
                                last_value = data[-1][1]
                                for i in range(extend_count):
                                    new_timestamp = last_timestamp + time_interval * (i + 1)
                                    data.append([new_timestamp, last_value])

                                extended_count += 1

                            # 复制源数据段的数值到目标数据段（保持目标的时间戳）
                            for i in range(source_end - source_start + 1):
                                source_idx = source_start + i
                                target_idx = target_start + i
                                # 只替换数值，保持时间戳不变
                                data[target_idx][1] = data[source_idx][1]

                            replaced_count += 1

                    message = f"已替换 {replaced_count} 个样品的数据（序号 {target_start+1}-{target_end+1}）"
                    if extended_count > 0:
                        message += f"，其中 {extended_count} 个样品自动扩展了数据"

                elif operation == "cycle_replace":
                    # 获取周期参数
                    cycle_start = int(cycle_start_entry.get()) - 1  # 转为0索引
                    cycle_end = int(cycle_end_entry.get()) - 1
                    replace_cycle_start = int(replace_cycle_start_entry.get()) - 1
                    replace_cycle_end = int(replace_cycle_end_entry.get()) - 1

                    # 获取周期电流数据
                    cycle_current_data_text = cycle_current_text.get('1.0', tk.END).strip()
                    if not cycle_current_data_text:
                        messagebox.showerror("错误", "请输入周期电流数据！", parent=op_window)
                        return

                    # 解析周期电流数据
                    try:
                        cycle_current_values = []
                        for line in cycle_current_data_text.split('\n'):
                            line = line.strip()
                            if line:
                                cycle_current_values.append(float(line))

                        if not cycle_current_values:
                            messagebox.showerror("错误", "没有有效的周期电流数据！", parent=op_window)
                            return

                        cycle_length = len(cycle_current_values)
                    except ValueError:
                        messagebox.showerror("错误", "周期电流数据格式错误，请确保每行一个数值！", parent=op_window)
                        return

                    # 验证参数
                    if cycle_start < 0 or cycle_end < cycle_start:
                        messagebox.showerror("错误", "周期范围无效！", parent=op_window)
                        return
                    if replace_cycle_start < 0 or replace_cycle_end < replace_cycle_start:
                        messagebox.showerror("错误", "替换周期范围无效！", parent=op_window)
                        return

                    cycle_data_length = cycle_end - cycle_start + 1

                    # 周期范围长度不再需要等于周期电流数据长度
                    # 周期电流数据将自动循环使用

                    # 对所有文件的电流series进行周期替换
                    replaced_count = 0
                    extended_count = 0
                    for file_info in files_to_process:
                        # 找到电流series（假设第二个是电流）
                        if len(file_info['series']) < 2:
                            continue

                        current_series = file_info['series'][1]  # 电流series
                        data = current_series['data']

                        # 如果替换范围超出数据长度，自动扩展数据
                        if replace_cycle_end >= len(data):
                            # 计算需要扩展的数据点数量
                            extend_count = replace_cycle_end - len(data) + 1

                            # 获取最后一个数据点的时间戳
                            last_timestamp = data[-1][0]

                            # 计算时间间隔（使用最后两个点的间隔）
                            if len(data) >= 2:
                                time_interval = data[-1][0] - data[-2][0]
                            else:
                                time_interval = 1000  # 默认1秒

                            # 扩展数据点（使用最后一个点的值）
                            last_value = data[-1][1]
                            for i in range(extend_count):
                                new_timestamp = last_timestamp + time_interval * (i + 1)
                                data.append([new_timestamp, last_value])

                            extended_count += 1

                        # 循环替换电流数据（不再要求整数倍）
                        for i in range(replace_cycle_start, replace_cycle_end + 1):
                            cycle_index = (i - replace_cycle_start) % cycle_length
                            data[i][1] = cycle_current_values[cycle_index]

                        replaced_count += 1

                    message = f"已对 {replaced_count} 个样品进行周期电流替换（序号 {replace_cycle_start+1}-{replace_cycle_end+1}）"
                    if extended_count > 0:
                        message += f"，其中 {extended_count} 个样品自动扩展了数据"

                elif operation == "cycle_voltage_replace":
                    # 获取周期电压参数
                    cycle_voltage_start = int(cycle_voltage_start_entry.get()) - 1  # 转为0索引
                    cycle_voltage_end = int(cycle_voltage_end_entry.get()) - 1
                    replace_cycle_voltage_start = int(replace_cycle_voltage_start_entry.get()) - 1
                    replace_cycle_voltage_end = int(replace_cycle_voltage_end_entry.get()) - 1

                    # 获取周期电压数据
                    cycle_voltage_data_text = cycle_voltage_text.get('1.0', tk.END).strip()
                    if not cycle_voltage_data_text:
                        messagebox.showerror("错误", "请输入周期电压数据！", parent=op_window)
                        return

                    # 解析周期电压数据
                    try:
                        cycle_voltage_values = []
                        for line in cycle_voltage_data_text.split('\n'):
                            line = line.strip()
                            if line:
                                cycle_voltage_values.append(float(line))

                        if not cycle_voltage_values:
                            messagebox.showerror("错误", "没有有效的周期电压数据！", parent=op_window)
                            return

                        cycle_voltage_length = len(cycle_voltage_values)
                    except ValueError:
                        messagebox.showerror("错误", "周期电压数据格式错误，请确保每行一个数值！", parent=op_window)
                        return

                    # 验证参数
                    if cycle_voltage_start < 0 or cycle_voltage_end < cycle_voltage_start:
                        messagebox.showerror("错误", "周期范围无效！", parent=op_window)
                        return
                    if replace_cycle_voltage_start < 0 or replace_cycle_voltage_end < replace_cycle_voltage_start:
                        messagebox.showerror("错误", "替换周期范围无效！", parent=op_window)
                        return

                    # 对所有文件的电压series进行周期替换
                    replaced_count = 0
                    extended_count = 0
                    for file_info in files_to_process:
                        # 找到电压series（假设第一个是电压）
                        if len(file_info['series']) < 1:
                            continue

                        voltage_series = file_info['series'][0]  # 电压series
                        data = voltage_series['data']

                        # 如果替换范围超出数据长度，自动扩展数据
                        if replace_cycle_voltage_end >= len(data):
                            # 计算需要扩展的数据点数量
                            extend_count = replace_cycle_voltage_end - len(data) + 1

                            # 获取最后一个数据点的时间戳
                            last_timestamp = data[-1][0]

                            # 计算时间间隔（使用最后两个点的间隔）
                            if len(data) >= 2:
                                time_interval = data[-1][0] - data[-2][0]
                            else:
                                time_interval = 1000  # 默认1秒

                            # 扩展数据点（使用最后一个点的值）
                            last_value = data[-1][1]
                            for i in range(extend_count):
                                new_timestamp = last_timestamp + time_interval * (i + 1)
                                data.append([new_timestamp, last_value])

                            extended_count += 1

                        # 循环替换电压数据
                        for i in range(replace_cycle_voltage_start, replace_cycle_voltage_end + 1):
                            cycle_index = (i - replace_cycle_voltage_start) % cycle_voltage_length
                            data[i][1] = cycle_voltage_values[cycle_index]

                        replaced_count += 1

                    message = f"已对 {replaced_count} 个样品进行周期电压替换（序号 {replace_cycle_voltage_start+1}-{replace_cycle_voltage_end+1}）"
                    if extended_count > 0:
                        message += f"，其中 {extended_count} 个样品自动扩展了数据"

                elif operation == "current_fluctuation":
                    # 获取波动参数
                    fluctuation_start = int(fluctuation_start_entry.get()) - 1  # 转为0索引
                    fluctuation_end = int(fluctuation_end_entry.get()) - 1
                    fluctuation_value = float(fluctuation_percent_entry.get())
                    fluctuation_type = fluctuation_type_var.get()

                    # 验证参数
                    if fluctuation_start < 0 or fluctuation_end < fluctuation_start:
                        messagebox.showerror("错误", "波动范围无效！", parent=op_window)
                        return

                    if fluctuation_type == "percent":
                        if fluctuation_value < 0 or fluctuation_value > 100:
                            messagebox.showerror("错误", "波动幅度必须在0-100之间！", parent=op_window)
                            return

                    # 对所有文件的电流series进行随机波动
                    import random
                    fluctuated_count = 0
                    for file_info in files_to_process:
                        # 找到电流series（假设第二个是电流）
                        if len(file_info['series']) < 2:
                            continue

                        current_series = file_info['series'][1]  # 电流series
                        data = current_series['data']

                        # 如果是跳跃波动，先确定哪些点需要波动
                        if fluctuation_type == "jump":
                            jump_ratio = float(fluctuation_jump_ratio_entry.get())
                            if jump_ratio < 0 or jump_ratio > 100:
                                messagebox.showerror("错误", "跳跃比例必须在0-100之间！", parent=op_window)
                                return

                            # 过滤掉电流小于0.02A的点，只从有效点中选择跳跃点
                            valid_indices = []
                            for idx in range(fluctuation_start, min(fluctuation_end + 1, len(data))):
                                if isinstance(data[idx], list) and len(data[idx]) >= 2:
                                    current_value = data[idx][1]
                                    if current_value is not None and current_value >= 0.02:
                                        valid_indices.append(idx)

                            # 从有效的非0点中随机选择需要波动的点
                            if len(valid_indices) > 0:
                                jump_count = int(len(valid_indices) * jump_ratio / 100)
                                jump_indices = set(random.sample(valid_indices, min(jump_count, len(valid_indices))))
                            else:
                                jump_indices = set()
                        else:
                            jump_indices = None

                        # 对指定范围内的电流数据添加波动或固定加减
                        for i in range(fluctuation_start, min(fluctuation_end + 1, len(data))):
                            # 如果是跳跃波动，检查当前索引是否在跳跃列表中
                            if fluctuation_type == "jump" and i not in jump_indices:
                                continue

                            if isinstance(data[i], list) and len(data[i]) >= 2:
                                current_value = data[i][1]

                                # 如果源数据为0或None，跳过不处理
                                if current_value is None or current_value == 0:
                                    continue

                                # 根据波动类型处理
                                if fluctuation_type == "percent" or fluctuation_type == "jump":
                                    # 百分比随机波动
                                    fluctuation_range = current_value * (fluctuation_value / 100)
                                    random_fluctuation = random.uniform(-fluctuation_range, fluctuation_range)
                                    new_value = current_value + random_fluctuation
                                    # 确保结果不为负数
                                    if new_value >= 0:
                                        data[i][1] = new_value
                                else:
                                    # 固定值加减
                                    # 对小于0.1A的电流值不生效
                                    if abs(current_value) < 0.1:
                                        continue

                                    new_value = current_value + fluctuation_value
                                    # 如果结果为负数，跳过不处理
                                    if new_value >= 0:
                                        data[i][1] = new_value

                        fluctuated_count += 1

                    if fluctuation_type == "percent":
                        message = f"已对 {fluctuated_count} 个样品的电流数据添加±{fluctuation_value}%的随机波动（序号 {fluctuation_start+1}-{fluctuation_end+1}）"
                    elif fluctuation_type == "jump":
                        jump_ratio = float(fluctuation_jump_ratio_entry.get())
                        message = f"已对 {fluctuated_count} 个样品的电流数据进行跳跃波动，随机{jump_ratio}%的数据±{fluctuation_value}%（序号 {fluctuation_start+1}-{fluctuation_end+1}）"
                    else:
                        if fluctuation_value >= 0:
                            message = f"已对 {fluctuated_count} 个样品的电流数据统一+{fluctuation_value}（序号 {fluctuation_start+1}-{fluctuation_end+1}）"
                        else:
                            message = f"已对 {fluctuated_count} 个样品的电流数据统一{fluctuation_value}（序号 {fluctuation_start+1}-{fluctuation_end+1}）"

                elif operation == "voltage_fluctuation":
                    # 获取电压波动参数
                    voltage_fluctuation_start = int(voltage_fluctuation_start_entry.get()) - 1  # 转为0索引
                    voltage_fluctuation_end = int(voltage_fluctuation_end_entry.get()) - 1
                    voltage_fluctuation_value = float(voltage_fluctuation_percent_entry.get())
                    voltage_fluctuation_type = voltage_fluctuation_type_var.get()

                    # 验证参数
                    if voltage_fluctuation_start < 0 or voltage_fluctuation_end < voltage_fluctuation_start:
                        messagebox.showerror("错误", "波动范围无效！", parent=op_window)
                        return

                    if voltage_fluctuation_type == "percent":
                        if voltage_fluctuation_value < 0 or voltage_fluctuation_value > 100:
                            messagebox.showerror("错误", "波动幅度必须在0-100之间！", parent=op_window)
                            return

                    # 对所有文件的电压series进行随机波动
                    import random
                    fluctuated_count = 0
                    for file_info in files_to_process:
                        # 找到电压series（假设第一个是电压）
                        if len(file_info['series']) < 1:
                            continue

                        voltage_series = file_info['series'][0]  # 电压series
                        data = voltage_series['data']

                        # 如果是跳跃波动，先确定哪些点需要波动
                        if voltage_fluctuation_type == "jump":
                            voltage_jump_ratio = float(voltage_fluctuation_jump_ratio_entry.get())
                            if voltage_jump_ratio < 0 or voltage_jump_ratio > 100:
                                messagebox.showerror("错误", "跳跃比例必须在0-100之间！", parent=op_window)
                                return

                            # 过滤掉电压小于0.02V的点，只从有效点中选择跳跃点
                            valid_indices = []
                            for idx in range(voltage_fluctuation_start, min(voltage_fluctuation_end + 1, len(data))):
                                if isinstance(data[idx], list) and len(data[idx]) >= 2:
                                    voltage_value = data[idx][1]
                                    if voltage_value is not None and voltage_value >= 0.02:
                                        valid_indices.append(idx)

                            # 从有效的非0点中随机选择需要波动的点
                            if len(valid_indices) > 0:
                                jump_count = int(len(valid_indices) * voltage_jump_ratio / 100)
                                voltage_jump_indices = set(random.sample(valid_indices, min(jump_count, len(valid_indices))))
                            else:
                                voltage_jump_indices = set()
                        else:
                            voltage_jump_indices = None

                        # 对指定范围内的电压数据添加波动或固定加减
                        for i in range(voltage_fluctuation_start, min(voltage_fluctuation_end + 1, len(data))):
                            # 如果是跳跃波动，检查当前索引是否在跳跃列表中
                            if voltage_fluctuation_type == "jump" and i not in voltage_jump_indices:
                                continue

                            if isinstance(data[i], list) and len(data[i]) >= 2:
                                voltage_value = data[i][1]

                                # 如果源数据为0或None，跳过不处理
                                if voltage_value is None or voltage_value == 0:
                                    continue

                                # 根据波动类型处理
                                if voltage_fluctuation_type == "percent" or voltage_fluctuation_type == "jump":
                                    # 百分比随机波动
                                    fluctuation_range = voltage_value * (voltage_fluctuation_value / 100)
                                    random_fluctuation = random.uniform(-fluctuation_range, fluctuation_range)
                                    new_value = voltage_value + random_fluctuation
                                    # 确保结果不为负数
                                    if new_value >= 0:
                                        data[i][1] = new_value
                                else:
                                    # 固定值加减
                                    new_value = voltage_value + voltage_fluctuation_value
                                    # 如果结果为负数，跳过不处理
                                    if new_value >= 0:
                                        data[i][1] = new_value

                        fluctuated_count += 1

                    if voltage_fluctuation_type == "percent":
                        message = f"已对 {fluctuated_count} 个样品的电压数据添加±{voltage_fluctuation_value}%的随机波动（序号 {voltage_fluctuation_start+1}-{voltage_fluctuation_end+1}）"
                    elif voltage_fluctuation_type == "jump":
                        voltage_jump_ratio = float(voltage_fluctuation_jump_ratio_entry.get())
                        message = f"已对 {fluctuated_count} 个样品的电压数据进行跳跃波动，随机{voltage_jump_ratio}%的数据±{voltage_fluctuation_value}%（序号 {voltage_fluctuation_start+1}-{voltage_fluctuation_end+1}）"
                    else:
                        if voltage_fluctuation_value >= 0:
                            message = f"已对 {fluctuated_count} 个样品的电压数据统一+{voltage_fluctuation_value}（序号 {voltage_fluctuation_start+1}-{voltage_fluctuation_end+1}）"
                        else:
                            message = f"已对 {fluctuated_count} 个样品的电压数据统一{voltage_fluctuation_value}（序号 {voltage_fluctuation_start+1}-{voltage_fluctuation_end+1}）"

                elif operation == "batch_jump_fluctuation":
                    # 批量跳跃波动 - 电压和电流分开设置
                    import random

                    # 获取电压跳跃波动参数
                    batch_voltage_start = int(batch_voltage_start_entry.get()) - 1
                    batch_voltage_end = int(batch_voltage_end_entry.get()) - 1
                    batch_voltage_percent = float(batch_voltage_percent_entry.get())
                    batch_voltage_jump_ratio = float(batch_voltage_jump_ratio_entry.get())

                    # 获取电流跳跃波动参数
                    batch_current_start = int(batch_current_start_entry.get()) - 1
                    batch_current_end = int(batch_current_end_entry.get()) - 1
                    batch_current_percent = float(batch_current_percent_entry.get())
                    batch_current_jump_ratio = float(batch_current_jump_ratio_entry.get())

                    # 验证参数
                    if batch_voltage_start < 0 or batch_voltage_end < batch_voltage_start:
                        messagebox.showerror("错误", "电压波动范围无效！", parent=op_window)
                        return
                    if batch_current_start < 0 or batch_current_end < batch_current_start:
                        messagebox.showerror("错误", "电流波动范围无效！", parent=op_window)
                        return
                    if batch_voltage_percent < 0 or batch_voltage_percent > 100:
                        messagebox.showerror("错误", "电压波动幅度必须在0-100之间！", parent=op_window)
                        return
                    if batch_current_percent < 0 or batch_current_percent > 100:
                        messagebox.showerror("错误", "电流波动幅度必须在0-100之间！", parent=op_window)
                        return
                    if batch_voltage_jump_ratio < 0 or batch_voltage_jump_ratio > 100:
                        messagebox.showerror("错误", "电压跳跃比例必须在0-100之间！", parent=op_window)
                        return
                    if batch_current_jump_ratio < 0 or batch_current_jump_ratio > 100:
                        messagebox.showerror("错误", "电流跳跃比例必须在0-100之间！", parent=op_window)
                        return

                    processed_count = 0
                    for file_info in files_to_process:
                        if len(file_info['series']) < 2:
                            continue

                        # 处理电压数据
                        voltage_series = file_info['series'][0]
                        voltage_data = voltage_series['data']

                        # 计算电压需要波动的数据点（排除小于0.02V的点）
                        voltage_valid_indices = []
                        for idx in range(batch_voltage_start, min(batch_voltage_end + 1, len(voltage_data))):
                            if isinstance(voltage_data[idx], list) and len(voltage_data[idx]) >= 2:
                                voltage_value = voltage_data[idx][1]
                                if voltage_value is not None and voltage_value >= 0.02:
                                    voltage_valid_indices.append(idx)

                        # 从有效的非0点中随机选择需要波动的点
                        if len(voltage_valid_indices) > 0:
                            voltage_jump_count = int(len(voltage_valid_indices) * batch_voltage_jump_ratio / 100)
                            voltage_jump_indices = set(random.sample(voltage_valid_indices, min(voltage_jump_count, len(voltage_valid_indices))))
                        else:
                            voltage_jump_indices = set()

                        # 对电压数据进行跳跃波动
                        for i in voltage_jump_indices:
                            if isinstance(voltage_data[i], list) and len(voltage_data[i]) >= 2:
                                voltage_value = voltage_data[i][1]
                                if voltage_value is not None and voltage_value >= 0.02:
                                    fluctuation_range = voltage_value * (batch_voltage_percent / 100)
                                    random_fluctuation = random.uniform(-fluctuation_range, fluctuation_range)
                                    new_value = voltage_value + random_fluctuation
                                    if new_value >= 0:
                                        voltage_data[i][1] = new_value

                        # 处理电流数据
                        current_series = file_info['series'][1]
                        current_data = current_series['data']

                        # 计算电流需要波动的数据点（排除小于0.02A的点）
                        current_valid_indices = []
                        for idx in range(batch_current_start, min(batch_current_end + 1, len(current_data))):
                            if isinstance(current_data[idx], list) and len(current_data[idx]) >= 2:
                                current_value = current_data[idx][1]
                                if current_value is not None and current_value >= 0.02:
                                    current_valid_indices.append(idx)

                        # 从有效的非0点中随机选择需要波动的点
                        if len(current_valid_indices) > 0:
                            current_jump_count = int(len(current_valid_indices) * batch_current_jump_ratio / 100)
                            current_jump_indices = set(random.sample(current_valid_indices, min(current_jump_count, len(current_valid_indices))))
                        else:
                            current_jump_indices = set()

                        # 对电流数据进行跳跃波动
                        for i in current_jump_indices:
                            if isinstance(current_data[i], list) and len(current_data[i]) >= 2:
                                current_value = current_data[i][1]
                                if current_value is not None and current_value >= 0.02:
                                    fluctuation_range = current_value * (batch_current_percent / 100)
                                    random_fluctuation = random.uniform(-fluctuation_range, fluctuation_range)
                                    new_value = current_value + random_fluctuation
                                    if new_value >= 0:
                                        current_data[i][1] = new_value

                        processed_count += 1

                    message = (f"已对 {processed_count} 个样品进行批量跳跃波动\n"
                              f"电压：随机{batch_voltage_jump_ratio}%的数据±{batch_voltage_percent}%（序号 {batch_voltage_start+1}-{batch_voltage_end+1}）\n"
                              f"电流：随机{batch_current_jump_ratio}%的数据±{batch_current_percent}%（序号 {batch_current_start+1}-{batch_current_end+1}）")

                elif operation == "peak_clipping":
                    # 获取削峰参数
                    peak_clipping_type = peak_clipping_type_var.get()
                    peak_clipping_start = int(peak_clipping_start_entry.get()) - 1  # 转为0索引
                    peak_clipping_end = int(peak_clipping_end_entry.get()) - 1
                    peak_clipping_max = float(peak_clipping_max_entry.get())
                    peak_clipping_min = float(peak_clipping_min_entry.get())
                    peak_clipping_method = peak_clipping_method_var.get()  # 获取削峰方式

                    # 验证参数
                    if peak_clipping_start < 0 or peak_clipping_end < peak_clipping_start:
                        messagebox.showerror("错误", "作用范围无效！", parent=op_window)
                        return
                    if peak_clipping_max <= peak_clipping_min:
                        messagebox.showerror("错误", "最大阈值必须大于最小阈值！", parent=op_window)
                        return

                    # 对所有文件的指定series进行削峰处理
                    clipped_count = 0
                    total_clipped_points = 0

                    for file_info in files_to_process:
                        # 根据类型选择series
                        if peak_clipping_type == "voltage":
                            if len(file_info['series']) < 1:
                                continue
                            series = file_info['series'][0]  # 电压series
                        else:  # current
                            if len(file_info['series']) < 2:
                                continue
                            series = file_info['series'][1]  # 电流series

                        data = series['data']
                        clipped_points = 0

                        # 对指定范围内的数据进行削峰处理
                        for i in range(peak_clipping_start, min(peak_clipping_end + 1, len(data))):
                            if isinstance(data[i], list) and len(data[i]) >= 2:
                                current_value = data[i][1]

                                # 检查是否为None
                                if current_value is None:
                                    continue

                                # 检查是否超出阈值
                                if current_value > peak_clipping_max or current_value < peak_clipping_min:
                                    if peak_clipping_method == "zero":
                                        # 方式1：直接设置为0
                                        data[i][1] = 0
                                        clipped_points += 1
                                    else:
                                        # 方式2：计算左右相邻点的平均值
                                        left_value = None
                                        right_value = None

                                        # 获取左侧有效值
                                        if i > 0 and isinstance(data[i-1], list) and len(data[i-1]) >= 2:
                                            left_value = data[i-1][1]

                                        # 获取右侧有效值
                                        if i < len(data) - 1 and isinstance(data[i+1], list) and len(data[i+1]) >= 2:
                                            right_value = data[i+1][1]

                                        # 计算替换值
                                        if left_value is not None and right_value is not None:
                                            # 左右都有值，取平均
                                            replacement_value = (left_value + right_value) / 2
                                        elif left_value is not None:
                                            # 只有左侧有值
                                            replacement_value = left_value
                                        elif right_value is not None:
                                            # 只有右侧有值
                                            replacement_value = right_value
                                        else:
                                            # 左右都没有值，保持原值
                                            continue

                                        # 替换异常值
                                        data[i][1] = replacement_value
                                        clipped_points += 1

                        if clipped_points > 0:
                            clipped_count += 1
                            total_clipped_points += clipped_points

                    type_name = "电压" if peak_clipping_type == "voltage" else "电流"
                    method_name = "设置为0" if peak_clipping_method == "zero" else "替换为左右平均值"
                    message = f"已对 {clipped_count} 个样品的{type_name}数据进行削峰处理（{method_name}），共处理 {total_clipped_points} 个异常点（阈值范围：{peak_clipping_min}-{peak_clipping_max}）"

                elif operation == "smart_peak_clipping":
                    # 获取智能削峰参数
                    smart_peak_start = int(smart_peak_start_entry.get()) - 1  # 转为0索引
                    smart_peak_end = int(smart_peak_end_entry.get()) - 1
                    smart_peak_tolerance = float(smart_peak_tolerance_entry.get())  # 百分比
                    smart_peak_method = smart_peak_method_var.get()

                    # 验证参数
                    if smart_peak_start < 0 or smart_peak_end < smart_peak_start:
                        messagebox.showerror("错误", "作用范围无效！", parent=op_window)
                        return
                    if smart_peak_tolerance <= 0 or smart_peak_tolerance > 100:
                        messagebox.showerror("错误", "允许偏差必须在0-100之间！", parent=op_window)
                        return

                    # 对所有文件的电压数据进行智能削峰处理
                    processed_count = 0
                    total_clipped_points = 0
                    detected_segments = {}  # 记录检测到的电压段

                    for file_info in files_to_process:
                        # 只处理电压数据（第一个series）
                        if len(file_info['series']) < 1:
                            continue

                        series = file_info['series'][0]  # 电压series
                        data = series['data']

                        # 第一步：自动识别电压段
                        # 收集指定范围内的所有有效电压值
                        voltage_values = []
                        for i in range(smart_peak_start, min(smart_peak_end + 1, len(data))):
                            if isinstance(data[i], list) and len(data[i]) >= 2:
                                value = data[i][1]
                                if value is not None and value > 0:  # 排除None和0值
                                    voltage_values.append(value)

                        if not voltage_values:
                            continue

                        # 使用聚类方法识别电压段（简单方法：按整数部分分组）
                        from collections import defaultdict
                        voltage_groups = defaultdict(list)

                        for v in voltage_values:
                            # 四舍五入到最近的整数作为电压段标识
                            segment = round(v)
                            voltage_groups[segment].append(v)

                        # 计算每个电压段的目标值（平均值）
                        segment_targets = {}
                        for segment, values in voltage_groups.items():
                            if len(values) >= 3:  # 至少3个点才认为是有效电压段
                                segment_targets[segment] = sum(values) / len(values)

                        if not segment_targets:
                            continue

                        # 记录检测到的电压段
                        file_name = os.path.basename(file_info['path'])
                        detected_segments[file_name] = segment_targets

                        # 第二步：对每个数据点进行削峰处理
                        clipped_points = 0

                        for i in range(smart_peak_start, min(smart_peak_end + 1, len(data))):
                            if isinstance(data[i], list) and len(data[i]) >= 2:
                                current_value = data[i][1]

                                if current_value is None or current_value <= 0:
                                    continue

                                # 找到最接近的电压段
                                closest_segment = min(segment_targets.keys(),
                                                     key=lambda s: abs(current_value - segment_targets[s]))
                                target_voltage = segment_targets[closest_segment]

                                # 计算允许的范围
                                tolerance_value = target_voltage * (smart_peak_tolerance / 100)
                                min_allowed = target_voltage - tolerance_value
                                max_allowed = target_voltage + tolerance_value

                                # 检查是否超出允许范围
                                if current_value < min_allowed or current_value > max_allowed:
                                    # 根据削峰方式处理
                                    if smart_peak_method == "target":
                                        # 设为目标电压
                                        data[i][1] = target_voltage
                                    elif smart_peak_method == "boundary":
                                        # 设为边界值
                                        if current_value < min_allowed:
                                            data[i][1] = min_allowed
                                        else:
                                            data[i][1] = max_allowed
                                    else:  # average
                                        # 计算左右平均值
                                        left_value = None
                                        right_value = None

                                        if i > 0 and isinstance(data[i-1], list) and len(data[i-1]) >= 2:
                                            left_value = data[i-1][1]

                                        if i < len(data) - 1 and isinstance(data[i+1], list) and len(data[i+1]) >= 2:
                                            right_value = data[i+1][1]

                                        if left_value is not None and right_value is not None:
                                            data[i][1] = (left_value + right_value) / 2
                                        elif left_value is not None:
                                            data[i][1] = left_value
                                        elif right_value is not None:
                                            data[i][1] = right_value
                                        else:
                                            data[i][1] = target_voltage

                                    clipped_points += 1

                        if clipped_points > 0:
                            processed_count += 1
                            total_clipped_points += clipped_points

                    # 生成结果消息
                    method_names = {
                        "target": "设为目标电压",
                        "boundary": "设为边界值",
                        "average": "左右平均值"
                    }
                    method_name = method_names.get(smart_peak_method, "未知方式")

                    # 显示检测到的电压段信息
                    segments_info = []
                    for file_name, segments in list(detected_segments.items())[:3]:  # 只显示前3个文件
                        seg_str = ", ".join([f"{int(s)}V({t:.2f}V)" for s, t in segments.items()])
                        segments_info.append(f"{file_name}: {seg_str}")

                    segments_text = "\n".join(segments_info)
                    if len(detected_segments) > 3:
                        segments_text += f"\n...等共{len(detected_segments)}个文件"

                    message = f"已对 {processed_count} 个样品进行智能电压削峰（{method_name}）\n共处理 {total_clipped_points} 个异常点（允许偏差：±{smart_peak_tolerance}%）\n\n检测到的电压段：\n{segments_text}"

                elif operation == "fill_zero":
                    # 获取填充0值参数
                    fill_zero_type = fill_zero_type_var.get()
                    fill_zero_start = int(fill_zero_start_entry.get()) - 1  # 转为0索引
                    fill_zero_end = int(fill_zero_end_entry.get()) - 1

                    # 验证参数
                    if fill_zero_start < 0 or fill_zero_end < fill_zero_start:
                        messagebox.showerror("错误", "作用范围无效！", parent=op_window)
                        return

                    # 对所有文件进行0值填充处理
                    filled_count = 0
                    total_filled_points = 0

                    for file_info in files_to_process:
                        file_filled_points = 0

                        # 根据类型选择要处理的series
                        series_to_process = []
                        if fill_zero_type == "voltage" or fill_zero_type == "both":
                            if len(file_info['series']) >= 1:
                                series_to_process.append(file_info['series'][0])  # 电压
                        if fill_zero_type == "current" or fill_zero_type == "both":
                            if len(file_info['series']) >= 2:
                                series_to_process.append(file_info['series'][1])  # 电流

                        for series in series_to_process:
                            data = series['data']

                            # 找到所有连续的0值段
                            i = fill_zero_start
                            while i <= min(fill_zero_end, len(data) - 1):
                                if isinstance(data[i], list) and len(data[i]) >= 2:
                                    current_value = data[i][1]

                                    # 如果当前值是0或None，找到连续0值段的起始和结束
                                    if current_value is not None and current_value == 0:
                                        zero_start = i
                                        zero_end = i

                                        # 找到连续0值段的结束位置
                                        while zero_end < min(fill_zero_end, len(data) - 1):
                                            if isinstance(data[zero_end + 1], list) and len(data[zero_end + 1]) >= 2:
                                                next_value = data[zero_end + 1][1]
                                                if next_value is None or next_value == 0:
                                                    zero_end += 1
                                                else:
                                                    break
                                            else:
                                                break

                                        # 找到0值段前后的非0值
                                        left_value = None
                                        right_value = None

                                        # 向左查找非0值
                                        for j in range(zero_start - 1, -1, -1):
                                            if isinstance(data[j], list) and len(data[j]) >= 2:
                                                val = data[j][1]
                                                if val is not None and val != 0:
                                                    left_value = val
                                                    break

                                        # 向右查找非0值
                                        for j in range(zero_end + 1, len(data)):
                                            if isinstance(data[j], list) and len(data[j]) >= 2:
                                                val = data[j][1]
                                                if val is not None and val != 0:
                                                    right_value = val
                                                    break

                                        # 使用线性插值填充0值段
                                        if left_value is not None and right_value is not None:
                                            # 计算插值
                                            num_points = zero_end - zero_start + 1
                                            for k in range(num_points):
                                                # 线性插值公式
                                                ratio = (k + 1) / (num_points + 1)
                                                interpolated_value = left_value + (right_value - left_value) * ratio
                                                data[zero_start + k][1] = interpolated_value
                                                file_filled_points += 1
                                        elif left_value is not None:
                                            # 只有左侧有值，用左侧值填充
                                            for k in range(zero_start, zero_end + 1):
                                                data[k][1] = left_value
                                                file_filled_points += 1
                                        elif right_value is not None:
                                            # 只有右侧有值，用右侧值填充
                                            for k in range(zero_start, zero_end + 1):
                                                data[k][1] = right_value
                                                file_filled_points += 1

                                        # 跳过已处理的0值段
                                        i = zero_end + 1
                                    else:
                                        i += 1
                                else:
                                    i += 1

                        if file_filled_points > 0:
                            filled_count += 1
                            total_filled_points += file_filled_points

                    type_names = {"voltage": "电压", "current": "电流", "both": "电压和电流"}
                    type_name = type_names.get(fill_zero_type, fill_zero_type)
                    message = f"已对 {filled_count} 个样品的{type_name}数据进行0值填充，共填充 {total_filled_points} 个数据点"

                elif operation == "time_shift":
                    # 检查是否使用方式1（目标日期时间）
                    target_datetime_str = target_datetime_entry.get().strip()

                    time_shift_ms = 0

                    if target_datetime_str:
                        # 方式1：使用目标日期时间
                        try:
                            from datetime import datetime

                            # 解析目标日期时间
                            target_dt = datetime.strptime(target_datetime_str, "%Y-%m-%d %H:%M:%S")
                            target_timestamp_ms = int(target_dt.timestamp() * 1000)

                            # 获取第一个文件的第一个数据点的时间戳
                            if files_to_process and files_to_process[0]['series']:
                                first_series = files_to_process[0]['series'][0]
                                if first_series['data']:
                                    first_timestamp = first_series['data'][0][0]

                                    # 计算时间偏移量
                                    time_shift_ms = target_timestamp_ms - first_timestamp
                                else:
                                    messagebox.showerror("错误", "数据为空！", parent=op_window)
                                    return
                            else:
                                messagebox.showerror("错误", "没有可用的数据！", parent=op_window)
                                return

                        except ValueError:
                            messagebox.showerror("错误", "日期时间格式错误！请使用格式：YYYY-MM-DD HH:MM:SS", parent=op_window)
                            return
                    else:
                        # 方式2：使用手动偏移量
                        try:
                            days = int(time_shift_days_entry.get())
                            hours = int(time_shift_hours_entry.get())
                            minutes = int(time_shift_minutes_entry.get())
                            seconds = int(time_shift_seconds_entry.get())
                        except ValueError:
                            messagebox.showerror("错误", "请输入有效的时间数值！", parent=op_window)
                            return

                        # 计算总的时间偏移量（毫秒）
                        time_shift_ms = (days * 24 * 3600 + hours * 3600 + minutes * 60 + seconds) * 1000

                    if time_shift_ms == 0:
                        messagebox.showwarning("提示", "时间调整量为0，无需调整！", parent=op_window)
                        return

                    # 对所有文件的所有series进行时间戳调整
                    adjusted_count = 0
                    for file_info in files_to_process:
                        for series in file_info['series']:
                            data = series['data']
                            # 调整每个数据点的时间戳
                            for point in data:
                                if isinstance(point, list) and len(point) >= 2:
                                    point[0] += time_shift_ms
                            adjusted_count += 1

                    if target_datetime_str:
                        message = f"已将 {adjusted_count} 个样品的开始时间设置为 {target_datetime_str}"
                    else:
                        sign = "+" if time_shift_ms > 0 else ""
                        days = int(time_shift_days_entry.get())
                        hours = int(time_shift_hours_entry.get())
                        minutes = int(time_shift_minutes_entry.get())
                        seconds = int(time_shift_seconds_entry.get())
                        message = f"已对 {adjusted_count} 个样品调整时间戳（{sign}{days}天 {sign}{hours}时 {sign}{minutes}分 {sign}{seconds}秒）"

                elif operation == "cycle_copy":
                    # 获取循环复制参数
                    from datetime import datetime

                    cycle_copy_type = cycle_copy_type_var.get()
                    cycle_copy_source_start = int(cycle_copy_source_start_entry.get()) - 1  # 转为0索引
                    cycle_copy_source_end = int(cycle_copy_source_end_entry.get()) - 1
                    max_points = int(cycle_copy_max_points_entry.get())

                    # 从下拉框获取开始和结束时间
                    start_datetime_str = f"{start_year_var.get()}-{start_month_var.get()}-{start_day_var.get()} {start_hour_var.get()}:{start_minute_var.get()}:{start_second_var.get()}"
                    end_datetime_str = f"{end_year_var.get()}-{end_month_var.get()}-{end_day_var.get()} {end_hour_var.get()}:{end_minute_var.get()}:{end_second_var.get()}"

                    # 验证参数
                    if cycle_copy_source_start < 0 or cycle_copy_source_end < cycle_copy_source_start:
                        messagebox.showerror("错误", "周期序列范围无效！", parent=op_window)
                        return

                    if not start_datetime_str or not end_datetime_str:
                        messagebox.showerror("错误", "请输入开始和结束日期时间！", parent=op_window)
                        return

                    # 解析日期时间
                    start_timestamp_ms = None
                    end_timestamp_ms = None
                    try:
                        start_dt = datetime.strptime(start_datetime_str, "%Y-%m-%d %H:%M:%S")
                        end_dt = datetime.strptime(end_datetime_str, "%Y-%m-%d %H:%M:%S")

                        start_timestamp_ms = int(start_dt.timestamp() * 1000)
                        end_timestamp_ms = int(end_dt.timestamp() * 1000)

                        if end_timestamp_ms <= start_timestamp_ms:
                            messagebox.showerror("错误", "结束时间必须大于开始时间！", parent=op_window)
                            return
                    except ValueError:
                        messagebox.showerror("错误", "日期时间格式错误！请使用格式：YYYY-MM-DD HH:MM:SS", parent=op_window)
                        return

                    # 确保时间戳已正确解析
                    if start_timestamp_ms is None or end_timestamp_ms is None:
                        messagebox.showerror("错误", "时间戳解析失败！", parent=op_window)
                        return

                    # 调试信息：显示解析的时间戳
                    print(f"循环复制 - 开始时间: {start_datetime_str} -> {start_timestamp_ms}")
                    print(f"循环复制 - 结束时间: {end_datetime_str} -> {end_timestamp_ms}")

                    cycle_length = cycle_copy_source_end - cycle_copy_source_start + 1

                    # 根据选择的类型确定要处理的series索引
                    series_indices = []
                    if cycle_copy_type == "voltage":
                        series_indices = [0]  # 只处理电压（第一个series）
                    elif cycle_copy_type == "current":
                        series_indices = [1]  # 只处理电流（第二个series）
                    else:  # both
                        series_indices = [0, 1]  # 处理电压和电流

                    # 对所有文件的指定series进行循环复制
                    copied_count = 0
                    for file_info in files_to_process:
                        print(f"\n处理文件: {file_info['name']}")
                        for series_idx in series_indices:
                            # 检查series是否存在
                            if series_idx >= len(file_info['series']):
                                continue

                            series = file_info['series'][series_idx]
                            series_name = series.get('name', f'Series {series_idx}')
                            print(f"  处理 series[{series_idx}]: {series_name}")
                            data = series['data']

                            # 检查并调整源范围以适应当前series的数据长度
                            actual_source_start = cycle_copy_source_start
                            actual_source_end = cycle_copy_source_end

                            if actual_source_start >= len(data):
                                print(f"  警告：series[{series_idx}] 数据长度 {len(data)} 小于周期序列起始位置 {actual_source_start + 1}，跳过此series")
                                continue

                            if actual_source_end >= len(data):
                                actual_source_end = len(data) - 1
                                print(f"  警告：series[{series_idx}] 数据长度不足，自动调整周期序列范围为 {actual_source_start + 1} 到 {actual_source_end + 1}")

                            actual_cycle_length = actual_source_end - actual_source_start + 1

                            # 提取周期数据的时间间隔（仅用于计算间隔，不使用实际时间戳）
                            if actual_source_start + 1 <= actual_source_end:
                                # 计算周期内的平均时间间隔
                                cycle_time_span = data[actual_source_end][0] - data[actual_source_start][0]
                                avg_time_interval = cycle_time_span / actual_cycle_length
                            else:
                                avg_time_interval = 1000  # 默认1秒

                            # 计算预期生成的数据点数量
                            time_range = end_timestamp_ms - start_timestamp_ms
                            estimated_points = int(time_range / avg_time_interval)

                            # 如果数据点超过设定的最大值，自动调整时间间隔
                            if estimated_points > max_points:
                                avg_time_interval = time_range / max_points
                                print(f"警告：预计生成 {estimated_points} 个数据点，已自动调整时间间隔为 {avg_time_interval:.2f}ms，将生成约 {max_points} 个数据点")

                            # 提取周期数据的数值（仅提取数值，不使用时间戳）
                            cycle_values = []
                            for i in range(actual_source_start, actual_source_end + 1):
                                if isinstance(data[i], list) and len(data[i]) >= 2:
                                    value = data[i][1]
                                    # 确保值不是None
                                    if value is not None:
                                        cycle_values.append(value)
                                    else:
                                        cycle_values.append(0)  # 使用0作为默认值
                                else:
                                    cycle_values.append(0)

                            # 如果没有有效的周期数据，跳过
                            if not cycle_values:
                                continue

                            # 获取周期序列的起始时间戳，用于排除周期序列之前的数据
                            cycle_start_timestamp = data[actual_source_start][0]

                            # 构建新数据：完全使用新的时间范围，不保留任何旧数据
                            new_data = []

                            # 不保留周期序列之前的数据，直接从新的开始时间生成
                            # （如果需要保留，取消下面的注释）
                            # for point in data:
                            #     if point[0] < cycle_start_timestamp:
                            #         new_data.append(point)

                            # 从开始时间循环生成新数据，直到结束时间
                            current_timestamp = start_timestamp_ms
                            cycle_index = 0

                            print(f"开始生成循环数据 - 起始时间戳: {current_timestamp}, 结束时间戳: {end_timestamp_ms}, 时间间隔: {avg_time_interval}")

                            while current_timestamp < end_timestamp_ms:
                                # 循环使用周期数据的数值
                                value = cycle_values[cycle_index % actual_cycle_length]
                                new_data.append([current_timestamp, value])

                                # 移动到下一个时间点
                                current_timestamp += int(avg_time_interval)
                                cycle_index += 1

                            # 确保结束时间点一定被包含
                            if not new_data or new_data[-1][0] != end_timestamp_ms:
                                value = cycle_values[cycle_index % actual_cycle_length]
                                new_data.append([end_timestamp_ms, value])

                            print(f"生成了 {len(new_data)} 个数据点，第一个时间戳: {new_data[0][0] if new_data else 'N/A'}, 最后一个时间戳: {new_data[-1][0] if new_data else 'N/A'}")

                            # 注意：周期序列的原始数据（包括之前选择的范围及其之后的所有数据）
                            # 已被完全排除，不会出现在新数据中，避免污染新生成的时间戳数据

                            # 替换series的数据
                            print(f"  替换前 series[{series_idx}] 数据点数: {len(series['data'])}")
                            series['data'] = new_data
                            print(f"  替换后 series[{series_idx}] 数据点数: {len(series['data'])}, 第一个时间戳: {series['data'][0][0] if series['data'] else 'N/A'}")

                            copied_count += 1

                    # 根据复制类型生成消息
                    type_names = {
                        "voltage": "电压",
                        "current": "电流",
                        "both": "电压和电流"
                    }
                    type_name = type_names.get(cycle_copy_type, "数据")
                    message = f"已对 {copied_count} 个样品的{type_name}进行循环复制（从 {start_datetime_str} 到 {end_datetime_str}，周期序列原始数据已完全删除）"

                elif operation == "yaxis_range":
                    # 获取Y轴范围参数
                    yaxis_range_type = yaxis_range_type_var.get()
                    yaxis_min_str = yaxis_min_entry.get().strip()
                    yaxis_max_str = yaxis_max_entry.get().strip()

                    # 解析最小值和最大值
                    yaxis_min = None
                    yaxis_max = None

                    if yaxis_min_str:
                        try:
                            yaxis_min = float(yaxis_min_str)
                        except ValueError:
                            messagebox.showerror("错误", "最小值必须是有效的数字！", parent=op_window)
                            return

                    if yaxis_max_str:
                        try:
                            yaxis_max = float(yaxis_max_str)
                        except ValueError:
                            messagebox.showerror("错误", "最大值必须是有效的数字！", parent=op_window)
                            return

                    # 验证最小值小于最大值
                    if yaxis_min is not None and yaxis_max is not None and yaxis_min >= yaxis_max:
                        messagebox.showerror("错误", "最小值必须小于最大值！", parent=op_window)
                        return

                    # 对所有文件设置Y轴范围
                    updated_count = 0
                    for file_info in files_to_process:
                        option = file_info['option']

                        # 确保yAxis存在
                        if 'yAxis' not in option:
                            option['yAxis'] = []

                        # 如果yAxis不是列表，转换为列表
                        if not isinstance(option['yAxis'], list):
                            option['yAxis'] = [option['yAxis']]

                        # 确保至少有两个Y轴（电压和电流）
                        while len(option['yAxis']) < 2:
                            option['yAxis'].append({})

                        # 根据类型设置Y轴范围
                        if yaxis_range_type == "voltage" or yaxis_range_type == "both":
                            # 设置电压Y轴（第一个Y轴）
                            if yaxis_min is not None or yaxis_max is not None:
                                option['yAxis'][0]['min'] = yaxis_min
                                option['yAxis'][0]['max'] = yaxis_max

                        if yaxis_range_type == "current" or yaxis_range_type == "both":
                            # 设置电流Y轴（第二个Y轴）
                            if yaxis_min is not None or yaxis_max is not None:
                                option['yAxis'][1]['min'] = yaxis_min
                                option['yAxis'][1]['max'] = yaxis_max

                        updated_count += 1

                    # 生成消息
                    type_names = {
                        "voltage": "电压Y轴",
                        "current": "电流Y轴",
                        "both": "电压和电流Y轴"
                    }
                    type_name = type_names.get(yaxis_range_type, "Y轴")

                    range_info = []
                    if yaxis_min is not None:
                        range_info.append(f"最小值: {yaxis_min}")
                    if yaxis_max is not None:
                        range_info.append(f"最大值: {yaxis_max}")

                    range_str = "，".join(range_info) if range_info else "自动"
                    message = f"已对 {updated_count} 个样品的{type_name}设置范围（{range_str}）"

                elif operation == "delete":
                    # 获取删除序号范围
                    delete_start = int(delete_start_entry.get()) - 1  # 转为0索引
                    delete_end = int(delete_end_entry.get()) - 1

                    # 验证参数
                    if delete_start < 0 or delete_end < delete_start:
                        messagebox.showerror("错误", "删除序号段无效！", parent=op_window)
                        return

                    # 对所有文件的所有series进行数据删除
                    deleted_count = 0
                    for file_info in files_to_process:
                        for series in file_info['series']:
                            data = series['data']

                            # 检查删除范围是否有效
                            if delete_start >= len(data):
                                continue

                            # 调整删除结束位置（不超过数据长度）
                            actual_delete_end = min(delete_end, len(data) - 1)

                            # 删除指定范围的数据
                            del data[delete_start:actual_delete_end + 1]

                            deleted_count += 1

                    message = f"已从 {deleted_count} 个样品中删除序号 {delete_start+1}-{delete_end+1} 的数据"

                elif operation == "insert":
                    # 获取插入参数
                    insert_position = int(insert_position_entry.get()) - 1  # 转为0索引

                    # 获取当前选择的数据类型
                    series_type = series_type_var.get()
                    series_index = 0 if series_type == "voltage" else 1

                    # 根据数据类型从对应的文本框获取数据
                    if series_type == "voltage":
                        insert_data_text = voltage_text.get('1.0', tk.END).strip()
                    else:
                        insert_data_text = current_text.get('1.0', tk.END).strip()

                    if not insert_data_text:
                        messagebox.showerror("错误", "请输入要插入的数据！", parent=op_window)
                        return

                    # 解析数据（每行一个值）
                    try:
                        insert_values = []
                        for line in insert_data_text.split('\n'):
                            line = line.strip()
                            if line:
                                insert_values.append(float(line))

                        if not insert_values:
                            messagebox.showerror("错误", "没有有效的数据！", parent=op_window)
                            return

                        insert_count = len(insert_values)
                    except ValueError:
                        messagebox.showerror("错误", "数据格式错误，请确保每行一个数值！", parent=op_window)
                        return

                    # 验证参数
                    if insert_position < 0:
                        messagebox.showerror("错误", "插入位置必须≥1！", parent=op_window)
                        return

                    # 对所有文件的指定series进行数据插入
                    inserted_count = 0
                    for file_info in files_to_process:
                        # 检查series是否存在
                        if series_index >= len(file_info['series']):
                            continue

                        series = file_info['series'][series_index]
                        data = series['data']

                        # 如果插入位置在末尾或超出范围，在末尾追加
                        if insert_position >= len(data) - 1:
                            # 在末尾插入，根据最后一个时间戳生成新时间戳
                            if len(data) >= 2:
                                time_interval = data[-1][0] - data[-2][0]
                            else:
                                time_interval = 1000  # 默认1秒

                            last_timestamp = data[-1][0]

                            for i, value in enumerate(insert_values):
                                new_timestamp = last_timestamp + time_interval * (i + 1)
                                data.append([new_timestamp, value])

                            inserted_count += 1
                        else:
                            # 在中间插入，需要重新计算所有时间戳
                            # 计算原始的时间间隔
                            if len(data) >= 2:
                                time_interval = data[-1][0] - data[-2][0]
                            else:
                                time_interval = 1000  # 默认1秒

                            # 在指定位置插入数据
                            for i, value in enumerate(insert_values):
                                data.insert(insert_position + 1 + i, [0, value])  # 先插入临时时间戳

                            # 重新计算所有时间戳（从插入位置开始）
                            for j in range(insert_position + 1, len(data)):
                                if j == insert_position + 1:
                                    # 第一个插入点的时间戳 = 前一个点的时间戳 + 时间间隔
                                    data[j][0] = data[insert_position][0] + time_interval
                                else:
                                    # 后续点的时间戳 = 前一个点的时间戳 + 时间间隔
                                    data[j][0] = data[j-1][0] + time_interval

                            inserted_count += 1

                    type_name = "电压" if series_type == "voltage" else "电流"
                    message = f"已在 {inserted_count} 个样品的{type_name}序号 {insert_position+1} 后插入 {insert_count} 个数据点"

                # 刷新当前显示
                self.load_file_by_index(self.current_file_index)

                # 同步更新当前文件的兼容属性
                if scope == "current" or scope == "all":
                    current_file = self.html_files[self.current_file_index]
                    self.series_data = current_file['series']
                    self.option_data = current_file['option']

                # 保存操作参数以便"再来一次"功能使用
                operation_names = {
                    "replace": "数据替换",
                    "cycle_replace": "周期电流替换",
                    "cycle_voltage_replace": "周期电压替换",
                    "current_fluctuation": "电流随机波动",
                    "voltage_fluctuation": "电压随机波动",
                    "batch_jump_fluctuation": "批量跳跃波动",
                    "peak_clipping": "电压电流削峰值",
                    "fill_zero": "填充0值",
                    "time_shift": "偷天换日",
                    "cycle_copy": "循环复制",
                    "delete": "删除序号",
                    "insert": "数据插入",
                    "yaxis_range": "Y轴范围设置"
                }

                # 收集所有输入框的值
                saved_params = {
                    'operation': operation,
                    'operation_name': operation_names.get(operation, operation),
                    'scope': scope,
                    'message': message,
                    # 数据替换参数
                    'source_start': source_start_entry.get() if source_start_entry.get() else '1',
                    'source_end': source_end_entry.get() if source_end_entry.get() else '10',
                    'target_start': target_start_entry.get() if target_start_entry.get() else '11',
                    'target_end': target_end_entry.get() if target_end_entry.get() else '20',
                    # 周期电流替换参数
                    'cycle_start': cycle_start_entry.get() if cycle_start_entry.get() else '1',
                    'cycle_end': cycle_end_entry.get() if cycle_end_entry.get() else '10',
                    'replace_cycle_start': replace_cycle_start_entry.get() if replace_cycle_start_entry.get() else '11',
                    'replace_cycle_end': replace_cycle_end_entry.get() if replace_cycle_end_entry.get() else '20',
                    'cycle_current_text': cycle_current_text.get('1.0', tk.END),
                    # 周期电压替换参数
                    'cycle_voltage_start': cycle_voltage_start_entry.get() if cycle_voltage_start_entry.get() else '1',
                    'cycle_voltage_end': cycle_voltage_end_entry.get() if cycle_voltage_end_entry.get() else '10',
                    'replace_cycle_voltage_start': replace_cycle_voltage_start_entry.get() if replace_cycle_voltage_start_entry.get() else '11',
                    'replace_cycle_voltage_end': replace_cycle_voltage_end_entry.get() if replace_cycle_voltage_end_entry.get() else '20',
                    'cycle_voltage_text': cycle_voltage_text.get('1.0', tk.END),
                    # 删除参数
                    'delete_start': delete_start_entry.get() if delete_start_entry.get() else '1',
                    'delete_end': delete_end_entry.get() if delete_end_entry.get() else '10',
                    # 电流波动参数
                    'fluctuation_start': fluctuation_start_entry.get() if fluctuation_start_entry.get() else '1',
                    'fluctuation_end': fluctuation_end_entry.get() if fluctuation_end_entry.get() else '100',
                    'fluctuation_percent': fluctuation_percent_entry.get() if fluctuation_percent_entry.get() else '5',
                    'fluctuation_type': fluctuation_type_var.get(),
                    # 电压波动参数
                    'voltage_fluctuation_start': voltage_fluctuation_start_entry.get() if voltage_fluctuation_start_entry.get() else '1',
                    'voltage_fluctuation_end': voltage_fluctuation_end_entry.get() if voltage_fluctuation_end_entry.get() else '100',
                    'voltage_fluctuation_percent': voltage_fluctuation_percent_entry.get() if voltage_fluctuation_percent_entry.get() else '5',
                    'voltage_fluctuation_type': voltage_fluctuation_type_var.get(),
                    # 批量跳跃波动参数
                    'batch_voltage_start': batch_voltage_start_entry.get() if batch_voltage_start_entry.get() else '1',
                    'batch_voltage_end': batch_voltage_end_entry.get() if batch_voltage_end_entry.get() else '100',
                    'batch_voltage_percent': batch_voltage_percent_entry.get() if batch_voltage_percent_entry.get() else '5',
                    'batch_voltage_jump_ratio': batch_voltage_jump_ratio_entry.get() if batch_voltage_jump_ratio_entry.get() else '30',
                    'batch_current_start': batch_current_start_entry.get() if batch_current_start_entry.get() else '1',
                    'batch_current_end': batch_current_end_entry.get() if batch_current_end_entry.get() else '100',
                    'batch_current_percent': batch_current_percent_entry.get() if batch_current_percent_entry.get() else '5',
                    'batch_current_jump_ratio': batch_current_jump_ratio_entry.get() if batch_current_jump_ratio_entry.get() else '30',
                    # 削峰参数
                    'peak_clipping_type': peak_clipping_type_var.get(),
                    'peak_clipping_start': peak_clipping_start_entry.get() if peak_clipping_start_entry.get() else '1',
                    'peak_clipping_end': peak_clipping_end_entry.get() if peak_clipping_end_entry.get() else '100',
                    'peak_clipping_max': peak_clipping_max_entry.get() if peak_clipping_max_entry.get() else '5.0',
                    'peak_clipping_min': peak_clipping_min_entry.get() if peak_clipping_min_entry.get() else '0.0',
                    # 填充0值参数
                    'fill_zero_type': fill_zero_type_var.get(),
                    'fill_zero_start': fill_zero_start_entry.get() if fill_zero_start_entry.get() else '1',
                    'fill_zero_end': fill_zero_end_entry.get() if fill_zero_end_entry.get() else '100',
                    # 时间调整参数
                    'target_datetime': target_datetime_entry.get() if target_datetime_entry.get() else '2025-01-01 00:00:00',
                    'time_shift_days': time_shift_days_entry.get() if time_shift_days_entry.get() else '0',
                    'time_shift_hours': time_shift_hours_entry.get() if time_shift_hours_entry.get() else '0',
                    'time_shift_minutes': time_shift_minutes_entry.get() if time_shift_minutes_entry.get() else '0',
                    'time_shift_seconds': time_shift_seconds_entry.get() if time_shift_seconds_entry.get() else '0',
                    # 循环复制参数
                    'cycle_copy_type': cycle_copy_type_var.get(),
                    'cycle_copy_source_start': cycle_copy_source_start_entry.get() if cycle_copy_source_start_entry.get() else '1',
                    'cycle_copy_source_end': cycle_copy_source_end_entry.get() if cycle_copy_source_end_entry.get() else '10',
                    'cycle_copy_start_datetime': f"{start_year_var.get()}-{start_month_var.get()}-{start_day_var.get()} {start_hour_var.get()}:{start_minute_var.get()}:{start_second_var.get()}",
                    'cycle_copy_end_datetime': f"{end_year_var.get()}-{end_month_var.get()}-{end_day_var.get()} {end_hour_var.get()}:{end_minute_var.get()}:{end_second_var.get()}",
                    'cycle_copy_max_points': cycle_copy_max_points_entry.get() if cycle_copy_max_points_entry.get() else '100000',
                    # 数据插入参数
                    'insert_position': insert_position_entry.get() if insert_position_entry.get() else '1',
                    'series_type': series_type_var.get(),
                    'voltage_text': voltage_text.get('1.0', tk.END),
                    'current_text': current_text.get('1.0', tk.END),
                    # Y轴范围参数
                    'yaxis_range_type': yaxis_range_type_var.get(),
                    'yaxis_min': yaxis_min_entry.get() if yaxis_min_entry.get() else '0',
                    'yaxis_max': yaxis_max_entry.get() if yaxis_max_entry.get() else '1'
                }

                self.last_operation_params = saved_params

                self.status_var.set(message)
                messagebox.showinfo("成功", message)
                # 不关闭窗口，保持当前状态

            except ValueError as e:
                messagebox.showerror("错误", "请输入有效的数值！", parent=op_window)
            except Exception as e:
                messagebox.showerror("错误", f"操作失败：{str(e)}", parent=op_window)

    def difference_analysis(self):
        """差异分析"""
        if not self.series_data:
            messagebox.showwarning("警告", "请先加载HTML文件！")
            return

        try:
            # 创建分析窗口
            analysis_window = tk.Toplevel(self.root)
            analysis_window.title("差异分析")
            analysis_window.geometry("900x600")
            analysis_window.transient(self.root)

            # 居中显示
            analysis_window.update_idletasks()
            x = (analysis_window.winfo_screenwidth() // 2) - (analysis_window.winfo_width() // 2)
            y = (analysis_window.winfo_screenheight() // 2) - (analysis_window.winfo_height() // 2)
            analysis_window.geometry(f"+{x}+{y}")

            # 标题
            ttk.Label(analysis_window, text="样品数据差异分析",
                     font=('Arial', 12, 'bold')).pack(pady=10)

            # 创建表格框架
            table_frame = ttk.Frame(analysis_window)
            table_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

            # 创建Treeview表格
            columns = ('series', 'count', 'min', 'max', 'avg', 'std', 'deviation')
            tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)

            # 定义列标题
            tree.heading('series', text='样品名称')
            tree.heading('count', text='数据点数')
            tree.heading('min', text='最小值')
            tree.heading('max', text='最大值')
            tree.heading('avg', text='平均值')
            tree.heading('std', text='标准差')
            tree.heading('deviation', text='与总平均偏差%')

            # 设置列宽
            tree.column('series', width=150, anchor=tk.W)
            tree.column('count', width=80, anchor=tk.CENTER)
            tree.column('min', width=100, anchor=tk.CENTER)
            tree.column('max', width=100, anchor=tk.CENTER)
            tree.column('avg', width=100, anchor=tk.CENTER)
            tree.column('std', width=100, anchor=tk.CENTER)
            tree.column('deviation', width=120, anchor=tk.CENTER)

            # 添加滚动条
            vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=vsb.set)

            # 布局
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            vsb.pack(side=tk.RIGHT, fill=tk.Y)

            # 计算统计数据
            all_values = []  # 所有样品的所有数据值
            stats_list = []  # 每个样品的统计信息

            for series in self.series_data:
                data = series.get('data', [])
                if not data:
                    continue

                values = [point[1] for point in data if isinstance(point, list) and len(point) >= 2]
                if not values:
                    continue

                all_values.extend(values)

                # 计算统计量
                count = len(values)
                min_val = min(values)
                max_val = max(values)
                avg_val = sum(values) / count

                # 计算标准差
                variance = sum((x - avg_val) ** 2 for x in values) / count
                std_val = variance ** 0.5

                stats_list.append({
                    'name': series.get('name', '未命名'),
                    'count': count,
                    'min': min_val,
                    'max': max_val,
                    'avg': avg_val,
                    'std': std_val,
                    'values': values
                })

            # 计算总平均值
            if all_values:
                total_avg = sum(all_values) / len(all_values)
            else:
                total_avg = 0

            # 填充表格数据
            for stats in stats_list:
                # 计算与总平均值的偏差百分比
                if total_avg != 0:
                    deviation = ((stats['avg'] - total_avg) / total_avg) * 100
                else:
                    deviation = 0

                # 根据偏差大小设置标签（用于高亮异常数据）
                tag = 'normal'
                if abs(deviation) > 10:  # 偏差超过10%标记为异常
                    tag = 'abnormal'
                elif abs(deviation) > 5:  # 偏差超过5%标记为警告
                    tag = 'warning'

                tree.insert('', tk.END, values=(
                    stats['name'],
                    stats['count'],
                    f"{stats['min']:.3f}",
                    f"{stats['max']:.3f}",
                    f"{stats['avg']:.3f}",
                    f"{stats['std']:.3f}",
                    f"{deviation:+.2f}%"
                ), tags=(tag,))

            # 设置标签颜色
            tree.tag_configure('normal', background='white')
            tree.tag_configure('warning', background='#FFF3CD')  # 浅黄色
            tree.tag_configure('abnormal', background='#F8D7DA')  # 浅红色

            # 添加总体统计信息
            summary_frame = ttk.LabelFrame(analysis_window, text="总体统计", padding=10)
            summary_frame.pack(fill=tk.X, padx=20, pady=10)

            if all_values:
                total_count = len(all_values)
                total_min = min(all_values)
                total_max = max(all_values)
                total_variance = sum((x - total_avg) ** 2 for x in all_values) / total_count
                total_std = total_variance ** 0.5

                summary_text = f"总数据点数: {total_count}  |  " \
                              f"总平均值: {total_avg:.3f}  |  " \
                              f"总标准差: {total_std:.3f}  |  " \
                              f"范围: [{total_min:.3f}, {total_max:.3f}]"
            else:
                summary_text = "无数据"

            ttk.Label(summary_frame, text=summary_text, font=('Arial', 10)).pack()

            # 添加说明
            info_frame = ttk.Frame(analysis_window)
            info_frame.pack(fill=tk.X, padx=20, pady=5)

            info_text = "说明：偏差 > 10% 显示为红色（异常），偏差 > 5% 显示为黄色（警告）"
            ttk.Label(info_frame, text=info_text, foreground='blue').pack()

            # 关闭按钮
            ttk.Button(analysis_window, text="关闭", command=analysis_window.destroy).pack(pady=10)

            self.status_var.set("差异分析完成")

        except Exception as e:
            messagebox.showerror("错误", f"差异分析失败：{str(e)}")
            self.status_var.set("分析失败")

    def comparison_view(self):
        """对比视图 - 生成包含所有曲线的对比HTML"""
        if not self.series_data:
            messagebox.showwarning("警告", "请先加载HTML文件！")
            return

        try:
            # 生成对比HTML文件
            comparison_html = self.generate_comparison_html()

            # 保存到临时文件
            import tempfile
            temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.html',
                                                    delete=False, encoding='utf-8')
            temp_file.write(comparison_html)
            temp_file.close()

            # 在默认浏览器中打开
            import webbrowser
            webbrowser.open('file://' + temp_file.name)

            self.status_var.set("对比视图已在浏览器中打开")
            messagebox.showinfo("成功", "对比视图已在浏览器中打开！\n可以查看所有样品的曲线对比。")

        except Exception as e:
            messagebox.showerror("错误", f"生成对比视图失败：{str(e)}")
            self.status_var.set("生成失败")

    def repeat_last_operation(self):
        """重复上一次的全局操作 - 打开全局操作窗口并恢复上次的设置"""
        if not self.last_operation_params:
            messagebox.showinfo("提示", "没有可重复的操作！\n\n请先执行一次全局操作。")
            # 直接打开全局操作窗口
            self.global_operations()
            return

        if not self.html_files:
            messagebox.showwarning("警告", "请先加载HTML文件！")
            return

        # 直接打开全局操作窗口，并传入上次的参数
        self.global_operations(restore_params=self.last_operation_params)

    def batch_generate(self):
        """批量生成HTML文件 - 以当前文件为模板，添加电流随机波动"""
        if not self.html_files:
            messagebox.showwarning("警告", "请先加载HTML文件！")
            return

        if self.current_file_index >= len(self.html_files):
            messagebox.showwarning("警告", "当前文件索引无效！")
            return

        # 创建批量生成窗口
        batch_window = tk.Toplevel(self.root)
        batch_window.title("批量生成HTML文件")
        batch_window.geometry("700x650")  # 增加窗口高度，确保所有控件可见

        # 按钮框架（先创建，固定在窗口底部）
        btn_frame = ttk.Frame(batch_window)
        btn_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=10)

        # 创建主框架和滚动条
        main_frame = ttk.Frame(batch_window)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 创建Canvas和滚动条
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 标题
        title_frame = ttk.Frame(scrollable_frame)
        title_frame.pack(fill=tk.X, padx=10, pady=10)
        ttk.Label(title_frame, text="批量生成HTML文件", font=('Arial', 14, 'bold')).pack()
        ttk.Label(title_frame, text=f"当前模板文件：{self.html_files[self.current_file_index]['name']}",
                 foreground='blue').pack(pady=5)

        # 参数设置框架
        params_frame = ttk.LabelFrame(scrollable_frame, text="生成参数", padding=10)
        params_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # 加载保存的参数
        import json
        config_file = os.path.join(os.path.dirname(__file__), 'batch_generate_config.json')
        saved_params = {}
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    saved_params = json.load(f)
            except:
                pass

        # 生成数量
        ttk.Label(params_frame, text="生成数量:").grid(row=0, column=0, sticky=tk.W, pady=5)
        count_entry = ttk.Entry(params_frame, width=15)
        count_entry.grid(row=0, column=1, padx=5, pady=5)
        count_entry.insert(0, saved_params.get("count", "6"))
        ttk.Label(params_frame, text="（生成多少个HTML文件）").grid(row=0, column=2, sticky=tk.W, pady=5)

        # 波动范围
        ttk.Label(params_frame, text="波动范围:").grid(row=1, column=0, sticky=tk.W, pady=5)
        range_start_entry = ttk.Entry(params_frame, width=10)
        range_start_entry.grid(row=1, column=1, padx=5, pady=5, sticky=tk.W)
        range_start_entry.insert(0, saved_params.get("range_start", "1"))
        ttk.Label(params_frame, text="到").grid(row=1, column=2, padx=2)
        range_end_entry = ttk.Entry(params_frame, width=10)
        range_end_entry.grid(row=1, column=3, padx=5, pady=5)

        # 默认填充全部范围（如果没有保存的参数）
        if "range_end" in saved_params:
            range_end_entry.insert(0, saved_params["range_end"])
        elif self.html_files and self.html_files[self.current_file_index]['series']:
            max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
            range_end_entry.insert(0, str(max_len))
        else:
            range_end_entry.insert(0, "100")

        def fill_all_range():
            if self.html_files and self.html_files[self.current_file_index]['series']:
                max_len = max(len(s['data']) for s in self.html_files[self.current_file_index]['series'])
                range_start_entry.delete(0, tk.END)
                range_start_entry.insert(0, "1")
                range_end_entry.delete(0, tk.END)
                range_end_entry.insert(0, str(max_len))

        ttk.Button(params_frame, text="全部", command=fill_all_range, width=6).grid(row=1, column=4, padx=2)

        # 波动对象选择
        data_type_var = tk.StringVar(value=saved_params.get("data_type", "current"))
        ttk.Label(params_frame, text="波动对象:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(params_frame, text="电流", variable=data_type_var,
                       value="current").grid(row=2, column=1, padx=5, sticky=tk.W)
        ttk.Radiobutton(params_frame, text="电压", variable=data_type_var,
                       value="voltage").grid(row=2, column=2, padx=5, sticky=tk.W)
        ttk.Radiobutton(params_frame, text="电压和电流", variable=data_type_var,
                       value="both").grid(row=2, column=3, padx=5, sticky=tk.W, columnspan=2)

        # 波动类型
        fluctuation_type_var = tk.StringVar(value=saved_params.get("fluctuation_type", "jump"))
        ttk.Label(params_frame, text="波动类型:").grid(row=3, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(params_frame, text="随机波动（±%）", variable=fluctuation_type_var,
                       value="percent").grid(row=3, column=1, padx=5, sticky=tk.W)
        ttk.Radiobutton(params_frame, text="跳跃波动（±%）", variable=fluctuation_type_var,
                       value="jump").grid(row=3, column=2, padx=5, sticky=tk.W)
        ttk.Radiobutton(params_frame, text="固定加减", variable=fluctuation_type_var,
                       value="fixed").grid(row=3, column=3, padx=5, sticky=tk.W)

        # 跳跃比例设置（仅跳跃波动时使用）
        ttk.Label(params_frame, text="跳跃比例(%):").grid(row=4, column=0, sticky=tk.W, pady=5)
        jump_ratio_entry = ttk.Entry(params_frame, width=15)
        jump_ratio_entry.grid(row=4, column=1, padx=5, pady=5)
        jump_ratio_entry.insert(0, saved_params.get("jump_ratio", "20"))
        ttk.Label(params_frame, text="（跳跃波动时，随机选择多少%的数据点进行波动）").grid(row=4, column=2, sticky=tk.W, pady=5, columnspan=3)

        # 波动限制
        ttk.Label(params_frame, text="波动上限:").grid(row=5, column=0, sticky=tk.W, pady=5)
        max_limit_entry = ttk.Entry(params_frame, width=15)
        max_limit_entry.grid(row=5, column=1, padx=5, pady=5)
        max_limit_entry.insert(0, saved_params.get("max_limit", ""))
        ttk.Label(params_frame, text="（可选，波动后超过此值则不波动，留空表示不限制）").grid(row=5, column=2, sticky=tk.W, pady=5, columnspan=3)

        # 波动模式
        fluctuation_mode_var = tk.StringVar(value=saved_params.get("fluctuation_mode", "independent"))
        ttk.Label(params_frame, text="波动模式:").grid(row=6, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(params_frame, text="整体偏移", variable=fluctuation_mode_var,
                       value="overall").grid(row=6, column=1, padx=5, sticky=tk.W)

        # 添加说明行
        fluctuation_mode_frame = ttk.Frame(params_frame)
        fluctuation_mode_frame.grid(row=6, column=2, columnspan=3, sticky=tk.W)
        ttk.Radiobutton(fluctuation_mode_frame, text="每点独立", variable=fluctuation_mode_var,
                       value="independent").pack(side=tk.LEFT)
        ttk.Label(fluctuation_mode_frame, text="（整体偏移保持曲线平滑，每点独立会产生毛刺）",
                 foreground='gray', font=('Arial', 8)).pack(side=tk.LEFT, padx=5)

        # 波动幅度
        ttk.Label(params_frame, text="波动幅度:").grid(row=7, column=0, sticky=tk.W, pady=5)
        fluctuation_entry = ttk.Entry(params_frame, width=15)
        fluctuation_entry.grid(row=7, column=1, padx=5, pady=5)
        fluctuation_entry.insert(0, saved_params.get("fluctuation_value", "30"))
        ttk.Label(params_frame, text="（随机波动/跳跃波动：5表示±5%；固定加减：0.1表示+0.1）").grid(row=7, column=2, sticky=tk.W, pady=5, columnspan=3)

        # 输出目录
        ttk.Label(params_frame, text="输出目录:").grid(row=8, column=0, sticky=tk.W, pady=5)
        output_dir_entry = ttk.Entry(params_frame, width=40)
        output_dir_entry.grid(row=8, column=1, padx=5, pady=5, columnspan=3, sticky=tk.W)

        # 默认输出目录为保存的目录或当前文件所在目录
        if "output_dir" in saved_params and os.path.exists(saved_params["output_dir"]):
            output_dir_entry.insert(0, saved_params["output_dir"])
        elif self.file_path:
            default_dir = os.path.dirname(self.file_path)
            output_dir_entry.insert(0, default_dir)

        def browse_output_dir():
            directory = filedialog.askdirectory(title="选择输出目录", parent=batch_window)
            if directory:
                output_dir_entry.delete(0, tk.END)
                output_dir_entry.insert(0, directory)

        ttk.Button(params_frame, text="浏览", command=browse_output_dir, width=8).grid(row=8, column=4, padx=2)

        # 文件名前缀
        ttk.Label(params_frame, text="文件名前缀:").grid(row=9, column=0, sticky=tk.W, pady=5)
        prefix_entry = ttk.Entry(params_frame, width=30)
        prefix_entry.grid(row=9, column=1, padx=5, pady=5, columnspan=2, sticky=tk.W)

        # 默认前缀为原文件名（去掉扩展名）
        if self.html_files:
            original_name = self.html_files[self.current_file_index]['name']
            prefix = os.path.splitext(original_name)[0]
            prefix_entry.insert(0, prefix)

        ttk.Label(params_frame, text="（生成的文件名格式：前缀-1.html, 前缀-2.html...；电流单位为μA时：前缀-1-l.html, 前缀-2-l.html...）").grid(row=9, column=2, sticky=tk.W, pady=5, columnspan=3)

        # 采样率设置
        ttk.Label(params_frame, text="输出采样率:").grid(row=10, column=0, sticky=tk.W, pady=5)
        sample_rate_entry = ttk.Entry(params_frame, width=15)
        sample_rate_entry.grid(row=10, column=1, padx=5, pady=5)
        sample_rate_entry.insert(0, saved_params.get("sample_rate", "1"))
        ttk.Label(params_frame, text="（1=全部输出，2=每2个取1个，3=每3个取1个...）").grid(row=10, column=2, sticky=tk.W, pady=5, columnspan=3)

        # 记住参数复选框
        remember_params_var = tk.BooleanVar(value=saved_params.get("remember", False))
        remember_checkbox = ttk.Checkbutton(params_frame, text="记住我的参数设置（下次打开自动加载）",
                                           variable=remember_params_var)
        remember_checkbox.grid(row=11, column=0, columnspan=5, sticky=tk.W, pady=10)

        def start_batch_generate():
            try:
                # 获取参数
                count = int(count_entry.get())
                range_start = int(range_start_entry.get()) - 1  # 转为0索引
                range_end = int(range_end_entry.get()) - 1
                data_type = data_type_var.get()
                fluctuation_type = fluctuation_type_var.get()
                fluctuation_mode = fluctuation_mode_var.get()
                fluctuation_value = float(fluctuation_entry.get())
                output_dir = output_dir_entry.get().strip()
                prefix = prefix_entry.get().strip()
                sample_rate = int(sample_rate_entry.get())
                jump_ratio = float(jump_ratio_entry.get()) if fluctuation_type == "jump" else 0

                # 获取波动上限（可选）
                max_limit_str = max_limit_entry.get().strip()
                max_limit = float(max_limit_str) if max_limit_str else None

                # 验证参数
                if count <= 0 or count > 1000:
                    messagebox.showerror("错误", "生成数量必须在1-1000之间！", parent=batch_window)
                    return

                if range_start < 0 or range_end < range_start:
                    messagebox.showerror("错误", "波动范围无效！", parent=batch_window)
                    return

                if not output_dir or not os.path.exists(output_dir):
                    messagebox.showerror("错误", "输出目录不存在！", parent=batch_window)
                    return

                if not prefix:
                    messagebox.showerror("错误", "请输入文件名前缀！", parent=batch_window)
                    return

                if sample_rate < 1:
                    messagebox.showerror("错误", "采样率必须大于等于1！", parent=batch_window)
                    return

                # 确认生成
                if not messagebox.askyesno("确认", f"将生成 {count} 个HTML文件到：\n{output_dir}\n\n确定继续吗？", parent=batch_window):
                    return

                # 如果用户勾选了"记住参数"，保存参数到配置文件
                if remember_params_var.get():
                    params_to_save = {
                        "count": count_entry.get(),
                        "range_start": range_start_entry.get(),
                        "range_end": range_end_entry.get(),
                        "data_type": data_type,
                        "fluctuation_type": fluctuation_type,
                        "fluctuation_mode": fluctuation_mode,
                        "fluctuation_value": fluctuation_entry.get(),
                        "jump_ratio": jump_ratio_entry.get(),
                        "max_limit": max_limit_entry.get(),
                        "output_dir": output_dir,
                        "sample_rate": sample_rate_entry.get(),
                        "remember": True
                    }
                    try:
                        with open(config_file, 'w', encoding='utf-8') as f:
                            json.dump(params_to_save, f, ensure_ascii=False, indent=2)
                    except:
                        pass
                else:
                    # 如果取消勾选，删除配置文件
                    if os.path.exists(config_file):
                        try:
                            os.remove(config_file)
                        except:
                            pass

                # 获取当前文件的数据作为模板
                template_file = self.html_files[self.current_file_index]

                import random
                import copy

                # 检测电流单位（从option的yAxis中获取）
                current_unit = "A"  # 默认为A
                if 'option' in template_file and 'yAxis' in template_file['option']:
                    yaxis_list = template_file['option']['yAxis']
                    if isinstance(yaxis_list, list) and len(yaxis_list) >= 2:
                        # 第二个Y轴通常是电流
                        current_yaxis = yaxis_list[1]
                        if 'name' in current_yaxis:
                            yaxis_name = current_yaxis['name']
                            # 检查是否包含μA或uA
                            if 'μA' in yaxis_name or 'uA' in yaxis_name or 'µA' in yaxis_name:
                                current_unit = "uA"

                generated_count = 0

                for i in range(count):
                    # 每次都从模板深拷贝（避免累积效应）
                    new_file_data = copy.deepcopy(template_file)

                    # 根据选择的数据类型添加波动
                    if data_type == "current" or data_type == "both":
                        # 对电流数据添加波动
                        if len(new_file_data['series']) >= 2:
                            current_series = new_file_data['series'][1]  # 电流series
                            data = current_series['data']

                            # 如果是跳跃波动，先确定哪些点需要波动
                            if fluctuation_type == "jump":
                                # 计算需要波动的数据点数量
                                total_points = min(range_end + 1, len(data)) - range_start
                                jump_count = int(total_points * jump_ratio / 100)
                                # 随机选择需要波动的索引
                                all_indices = list(range(range_start, min(range_end + 1, len(data))))
                                jump_indices = set(random.sample(all_indices, jump_count)) if jump_count > 0 else set()
                            else:
                                jump_indices = None

                            # 根据波动模式处理
                            if fluctuation_mode == "overall":
                                # 整体偏移模式：为整个文件生成一个固定的偏移量
                                if fluctuation_type == "percent" or fluctuation_type == "jump":
                                    # 计算范围内的平均值
                                    valid_values = [data[j][1] for j in range(range_start, min(range_end + 1, len(data)))
                                                   if isinstance(data[j], list) and len(data[j]) >= 2 and data[j][1] is not None and data[j][1] != 0]
                                    if valid_values:
                                        avg_value = sum(valid_values) / len(valid_values)
                                        fluctuation_range = avg_value * (fluctuation_value / 100)
                                        overall_offset = random.uniform(-fluctuation_range, fluctuation_range)
                                    else:
                                        overall_offset = 0
                                else:
                                    overall_offset = fluctuation_value

                                # 对指定范围内的所有数据点应用相同的偏移
                                for j in range(range_start, min(range_end + 1, len(data))):
                                    # 如果是跳跃波动，检查当前索引是否在跳跃列表中
                                    if fluctuation_type == "jump" and j not in jump_indices:
                                        continue

                                    if isinstance(data[j], list) and len(data[j]) >= 2:
                                        current_value = data[j][1]
                                        if current_value is not None and current_value != 0:
                                            if fluctuation_type == "percent" or fluctuation_type == "jump":
                                                new_value = current_value + overall_offset
                                            else:
                                                new_value = current_value + overall_offset
                                            # 检查是否超过上限
                                            if new_value >= 0:
                                                if max_limit is None or new_value <= max_limit:
                                                    data[j][1] = new_value
                            else:
                                # 每点独立模式：每个数据点独立随机波动
                                for j in range(range_start, min(range_end + 1, len(data))):
                                    # 如果是跳跃波动，检查当前索引是否在跳跃列表中
                                    if fluctuation_type == "jump" and j not in jump_indices:
                                        continue

                                    if isinstance(data[j], list) and len(data[j]) >= 2:
                                        current_value = data[j][1]

                                        # 如果源数据为0或None，跳过
                                        if current_value is None or current_value == 0:
                                            continue

                                        # 根据波动类型处理
                                        if fluctuation_type == "percent" or fluctuation_type == "jump":
                                            # 百分比随机波动
                                            fluctuation_range = current_value * (fluctuation_value / 100)
                                            random_fluctuation = random.uniform(-fluctuation_range, fluctuation_range)
                                            new_value = current_value + random_fluctuation
                                            # 检查是否超过上限
                                            if new_value >= 0:
                                                if max_limit is None or new_value <= max_limit:
                                                    data[j][1] = new_value
                                        else:
                                            # 固定值加减
                                            new_value = current_value + fluctuation_value
                                            # 检查是否超过上限
                                            if new_value >= 0:
                                                if max_limit is None or new_value <= max_limit:
                                                    data[j][1] = new_value

                    if data_type == "voltage" or data_type == "both":
                        # 对电压数据添加波动
                        if len(new_file_data['series']) >= 1:
                            voltage_series = new_file_data['series'][0]  # 电压series
                            data = voltage_series['data']

                            # 如果是跳跃波动，先确定哪些点需要波动
                            if fluctuation_type == "jump":
                                # 计算需要波动的数据点数量
                                total_points = min(range_end + 1, len(data)) - range_start
                                jump_count = int(total_points * jump_ratio / 100)
                                # 随机选择需要波动的索引
                                all_indices = list(range(range_start, min(range_end + 1, len(data))))
                                jump_indices = set(random.sample(all_indices, jump_count)) if jump_count > 0 else set()
                            else:
                                jump_indices = None

                            # 根据波动模式处理
                            if fluctuation_mode == "overall":
                                # 整体偏移模式：为整个文件生成一个固定的偏移量
                                if fluctuation_type == "percent" or fluctuation_type == "jump":
                                    # 计算范围内的平均值
                                    valid_values = [data[j][1] for j in range(range_start, min(range_end + 1, len(data)))
                                                   if isinstance(data[j], list) and len(data[j]) >= 2 and data[j][1] is not None and data[j][1] != 0]
                                    if valid_values:
                                        avg_value = sum(valid_values) / len(valid_values)
                                        fluctuation_range = avg_value * (fluctuation_value / 100)
                                        overall_offset = random.uniform(-fluctuation_range, fluctuation_range)
                                    else:
                                        overall_offset = 0
                                else:
                                    overall_offset = fluctuation_value

                                # 对指定范围内的所有数据点应用相同的偏移
                                for j in range(range_start, min(range_end + 1, len(data))):
                                    # 如果是跳跃波动，检查当前索引是否在跳跃列表中
                                    if fluctuation_type == "jump" and j not in jump_indices:
                                        continue

                                    if isinstance(data[j], list) and len(data[j]) >= 2:
                                        voltage_value = data[j][1]
                                        if voltage_value is not None and voltage_value != 0:
                                            if fluctuation_type == "percent" or fluctuation_type == "jump":
                                                new_value = voltage_value + overall_offset
                                            else:
                                                new_value = voltage_value + overall_offset
                                            # 检查是否超过上限
                                            if new_value >= 0:
                                                if max_limit is None or new_value <= max_limit:
                                                    data[j][1] = new_value
                            else:
                                # 每点独立模式：每个数据点独立随机波动
                                for j in range(range_start, min(range_end + 1, len(data))):
                                    # 如果是跳跃波动，检查当前索引是否在跳跃列表中
                                    if fluctuation_type == "jump" and j not in jump_indices:
                                        continue

                                    if isinstance(data[j], list) and len(data[j]) >= 2:
                                        voltage_value = data[j][1]

                                        # 如果源数据为0或None，跳过
                                        if voltage_value is None or voltage_value == 0:
                                            continue

                                        # 根据波动类型处理
                                        if fluctuation_type == "percent" or fluctuation_type == "jump":
                                            # 百分比随机波动
                                            fluctuation_range = voltage_value * (fluctuation_value / 100)
                                            random_fluctuation = random.uniform(-fluctuation_range, fluctuation_range)
                                            new_value = voltage_value + random_fluctuation
                                            # 检查是否超过上限
                                            if new_value >= 0:
                                                if max_limit is None or new_value <= max_limit:
                                                    data[j][1] = new_value
                                        else:
                                            # 固定值加减
                                            new_value = voltage_value + fluctuation_value
                                            # 检查是否超过上限
                                            if new_value >= 0:
                                                if max_limit is None or new_value <= max_limit:
                                                    data[j][1] = new_value

                    # 应用采样率
                    if sample_rate > 1:
                        for series in new_file_data['series']:
                            if 'data' in series:
                                # 对数据进行采样，每sample_rate个取1个
                                series['data'] = [series['data'][idx] for idx in range(0, len(series['data']), sample_rate)]

                    # 生成新的HTML文件
                    # 文件名格式：电流单位为uA时添加-l后缀，单位为A时不添加
                    if current_unit == "uA":
                        output_filename = f"{prefix}-{i+1}-l.html"
                    else:
                        output_filename = f"{prefix}-{i+1}.html"
                    output_path = os.path.join(output_dir, output_filename)

                    # 构建新的HTML内容
                    soup = BeautifulSoup(new_file_data['content'], 'html.parser')
                    script_tag = soup.find('script', string=lambda text: text and ('option' in text or 'chart' in text))

                    if script_tag:
                        # 更新option数据
                        new_file_data['option']['series'] = new_file_data['series']
                        option_json = json.dumps(new_file_data['option'], ensure_ascii=False, indent=2)

                        # 保留原始script内容，只替换option的JSON部分
                        original_script = script_tag.string
                        if original_script:
                            # 使用正则表达式找到option = {...}; 的部分并替换
                            import re
                            # 匹配 const option = {...}; 或 var option = {...};
                            pattern = r'(const|var)\s+option\s*=\s*\{[\s\S]*?\};'

                            # 使用函数替换，避免反向引用问题
                            def replace_option(match):
                                keyword = match.group(1)
                                return f'{keyword} option = {option_json};'

                            new_script = re.sub(pattern, replace_option, original_script)
                            script_tag.string = new_script

                    # 保存文件（无论是否找到script标签都保存）
                    with open(output_path, 'w', encoding='utf-8') as f:
                        f.write(str(soup))

                    generated_count += 1

                batch_window.destroy()
                messagebox.showinfo("成功", f"成功生成 {generated_count} 个HTML文件！\n\n保存位置：{output_dir}")

            except ValueError as e:
                messagebox.showerror("错误", f"参数格式错误：{str(e)}", parent=batch_window)
            except Exception as e:
                messagebox.showerror("错误", f"批量生成失败：{str(e)}", parent=batch_window)

        ttk.Button(btn_frame, text="开始生成", command=start_batch_generate, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消", command=batch_window.destroy, width=15).pack(side=tk.LEFT, padx=5)

    def _execute_operation(self, params):
        """执行操作的核心逻辑（预留接口）"""
        pass

    def _do_replace(self, files_to_process, params):
        """数据替换操作（预留接口）"""
        pass

    def _do_cycle_replace(self, files_to_process, params):
        """周期电流替换操作（预留接口）"""
        pass

    def _do_cycle_voltage_replace(self, files_to_process, params):
        """周期电压替换操作（预留接口）"""
        pass

    def _do_current_fluctuation(self, files_to_process, params):
        """电流随机波动操作（预留接口）"""
        pass

    def _do_voltage_fluctuation(self, files_to_process, params):
        """电压随机波动操作（预留接口）"""
        pass

    def _do_peak_clipping(self, files_to_process, params):
        """削峰值操作（预留接口）"""
        pass

    def _do_time_shift(self, files_to_process, params):
        """时间戳调整操作（预留接口）"""
        pass

    def _do_cycle_copy(self, files_to_process, params):
        """循环复制操作（预留接口）"""
        pass

    def _do_delete(self, files_to_process, params):
        """删除操作（预留接口）"""
        pass

    def _do_insert(self, files_to_process, params):
        """插入操作（预留接口）"""
        pass

    def generate_comparison_html(self):
        """生成对比视图的HTML内容"""
        # 准备series数据（所有曲线）
        series_json = json.dumps(self.series_data, ensure_ascii=False, indent=2)

        # 生成HTML模板
        html_template = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>样品数据对比视图</title>
    <script src="https://cdn.jsdelivr.net/npm/echarts@5.4.3/dist/echarts.min.js"></script>
    <style>
        body {{
            margin: 0;
            padding: 20px;
            font-family: Arial, sans-serif;
            background-color: #f5f5f5;
        }}
        .header {{
            text-align: center;
            margin-bottom: 20px;
        }}
        .header h1 {{
            color: #333;
            margin: 0;
        }}
        .header p {{
            color: #666;
            margin: 5px 0;
        }}
        #chart-container {{
            width: 95%;
            height: 700px;
            margin: 0 auto;
            background-color: white;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>样品数据对比视图</h1>
        <p>所有样品曲线叠加显示，便于对比分析</p>
    </div>
    <div id="chart-container"></div>

    <script>
        const chartDom = document.getElementById('chart-container');
        const myChart = echarts.init(chartDom);

        const seriesData = {series_json};

        const option = {{
            title: {{
                text: '多样品数据对比',
                left: 'center',
                top: 10,
                textStyle: {{
                    fontSize: 18,
                    fontWeight: 'bold'
                }}
            }},
            tooltip: {{
                trigger: 'axis',
                axisPointer: {{
                    type: 'cross',
                    label: {{
                        backgroundColor: '#6a7985'
                    }}
                }},
                formatter: function(params) {{
                    let result = new Date(params[0].axisValue).toLocaleString() + '<br/>';
                    params.forEach(param => {{
                        result += param.marker + param.seriesName + ': ' +
                                 param.value[1].toFixed(4) + '<br/>';
                    }});
                    return result;
                }}
            }},
            legend: {{
                data: seriesData.map(s => s.name),
                top: 40,
                type: 'scroll',
                orient: 'horizontal'
            }},
            grid: {{
                left: '3%',
                right: '4%',
                bottom: '3%',
                top: 100,
                containLabel: true
            }},
            toolbox: {{
                feature: {{
                    dataZoom: {{
                        yAxisIndex: 'none'
                    }},
                    restore: {{}},
                    saveAsImage: {{}}
                }},
                right: 20,
                top: 40
            }},
            xAxis: {{
                type: 'time',
                boundaryGap: false,
                axisLabel: {{
                    formatter: function(value) {{
                        const date = new Date(value);
                        return date.toLocaleTimeString();
                    }}
                }}
            }},
            yAxis: {{
                type: 'value',
                name: '数值',
                axisLabel: {{
                    formatter: '{{value}}'
                }}
            }},
            dataZoom: [
                {{
                    type: 'inside',
                    start: 0,
                    end: 100
                }},
                {{
                    start: 0,
                    end: 100,
                    handleIcon: 'M10.7,11.9v-1.3H9.3v1.3c-4.9,0.3-8.8,4.4-8.8,9.4c0,5,3.9,9.1,8.8,9.4v1.3h1.3v-1.3c4.9-0.3,8.8-4.4,8.8-9.4C19.5,16.3,15.6,12.2,10.7,11.9z M13.3,24.4H6.7V23h6.6V24.4z M13.3,19.6H6.7v-1.4h6.6V19.6z',
                    handleSize: '80%',
                    handleStyle: {{
                        color: '#fff',
                        shadowBlur: 3,
                        shadowColor: 'rgba(0, 0, 0, 0.6)',
                        shadowOffsetX: 2,
                        shadowOffsetY: 2
                    }}
                }}
            ],
            series: seriesData.map((s, index) => ({{
                name: s.name,
                type: 'line',
                smooth: true,
                symbol: 'none',
                sampling: 'lttb',
                lineStyle: {{
                    width: 2
                }},
                emphasis: {{
                    focus: 'series'
                }},
                data: s.data
            }}))
        }};

        myChart.setOption(option);

        // 响应窗口大小变化
        window.addEventListener('resize', function() {{
            myChart.resize();
        }});
    </script>
</body>
</html>"""

        return html_template


def main():
    """主函数"""
    root = tk.Tk()
    app = EChartsEditor(root)
    root.mainloop()


if __name__ == "__main__":
    main()
