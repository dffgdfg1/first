import os
import re
import json
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

FONT_NAME = "Arial"
FONT_SIZE = 11
SESSION_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".last_session.json")

# 温度文件夹名关键字映射
TEMP_MAP = {
    "23": "T_RT", "rt": "T_RT",
    "-40": "T_min", "tmin": "T_min",
    "85": "T_max", "tmax": "T_max",
}
TEMP_ORDER = ["T_min", "T_RT", "T_max"]
VOLT_ORDER = ["8V", "14V", "16V"]


def parse_folder_name(name):
    """从文件夹名解析温度标签和电压，如 '-40 8V' -> ('T_min', '8V')"""
    name_lower = name.lower()
    temp_label = None
    for key, val in TEMP_MAP.items():
        if key in name_lower:
            temp_label = val
            break
    volt_match = re.search(r'(\d+)[vV]', name)
    volt = volt_match.group(1) + "V" if volt_match else name
    return temp_label, volt


def find_csv_folder(root):
    """在 root 下递归找到含CSV文件的最深目录"""
    for dirpath, _, filenames in os.walk(root):
        csvs = [f for f in filenames if f.endswith('.csv')]
        if csvs:
            return dirpath
    return None


def read_csvs(folder):
    """读取文件夹中所有CSV，返回按camera编号排序的 [(camera_num, fname, avg), ...]"""
    files = [f for f in os.listdir(folder) if f.endswith('.csv')]
    results = []
    for fname in files:
        m = re.search(r'camera_(\d+)', fname)
        num = int(m.group(1)) if m else 999
        fpath = os.path.join(folder, fname)
        df = pd.read_csv(fpath, header=0)
        avg = pd.to_numeric(df.iloc[:, 3], errors='coerce').mean()
        results.append((num, fname, avg))
    results.sort(key=lambda x: x[0])
    return results


def auto_split_led_poc(entries):
    """按均值自动分组：小6个=LED，大6个=POC，返回 (led_list, poc_list) 各含 (camera_num, fname, avg)"""
    sorted_by_avg = sorted(entries, key=lambda x: x[2])
    led = sorted(sorted_by_avg[:6], key=lambda x: x[0])
    poc = sorted(sorted_by_avg[6:], key=lambda x: x[0])
    return led, poc


def save_session(before_data, after_data):
    data = {
        "before": {f"{k[0]}|{k[1]}": v["csv_folder"] for k, v in before_data.items()},
        "after":  {f"{k[0]}|{k[1]}": v["csv_folder"] for k, v in after_data.items()},
    }
    try:
        with open(SESSION_FILE, "w", encoding="utf-8") as f:
            import json; json.dump(data, f, ensure_ascii=False)
    except Exception:
        pass


def load_session_folders():
    try:
        with open(SESSION_FILE, "r", encoding="utf-8") as f:
            import json; return json.load(f)
    except Exception:
        return None


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("电流变化分析工具")
        self.geometry("860x620")
        self.resizable(True, True)

        # {(temp_label, volt): {"csv_folder": str, "led": [...], "poc": [...]}}
        self.before_data = {}
        self.after_data = {}

        self._build_ui()

    def _build_ui(self):
        pad = dict(padx=10, pady=5)

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, **pad)

        self.tab_before = tk.Frame(nb)
        self.tab_after = tk.Frame(nb)
        nb.add(self.tab_before, text="试验前数据（必选）")
        nb.add(self.tab_after, text="试验后数据（可选，用于对比）")

        self._build_tab(self.tab_before, "before")
        self._build_tab(self.tab_after, "after")

        tk.Button(self, text="生成 Excel", command=self.export_excel,
                  bg="#1565C0", fg="white",
                  font=(FONT_NAME, 12, "bold")).pack(side="left", padx=10, pady=8)
        tk.Button(self, text="恢复上次文件", command=self._restore_session,
                  font=(FONT_NAME, FONT_SIZE)).pack(side="left", pady=8)

    def _build_tab(self, tab, kind):
        pad = dict(padx=10, pady=5)

        frm_btn = tk.Frame(tab)
        frm_btn.pack(fill="x", **pad)
        tk.Button(frm_btn, text="选择根目录（自动扫描9个子文件夹）",
                  font=(FONT_NAME, FONT_SIZE),
                  command=lambda: self._load_root(kind)).pack(side="left", **pad)
        tk.Button(frm_btn, text="清空",
                  font=(FONT_NAME, FONT_SIZE),
                  command=lambda: self._clear(kind)).pack(side="left", **pad)

        cols = ("温度", "电压", "CSV文件夹", "LED通道", "POC通道")
        tree = ttk.Treeview(tab, columns=cols, show="headings", height=12)
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=80, anchor="center")
        tree.column("CSV文件夹", width=340, anchor="w")
        tree.pack(fill="both", expand=True, **pad)

        sb = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=sb.set)

        if kind == "before":
            self.tree_before = tree
        else:
            self.tree_after = tree

    def _load_root(self, kind):
        root = filedialog.askdirectory(title="选择根目录（含9个子文件夹）")
        if not root:
            return

        data_dict = self.before_data if kind == "before" else self.after_data
        tree = self.tree_before if kind == "before" else self.tree_after

        subdirs = sorted([d for d in os.listdir(root)
                          if os.path.isdir(os.path.join(root, d))])
        loaded = 0
        skipped = []
        for sub in subdirs:
            temp_label, volt = parse_folder_name(sub)
            if not temp_label:
                skipped.append(sub)
                continue

            csv_folder = find_csv_folder(os.path.join(root, sub))
            if not csv_folder:
                skipped.append(sub)
                continue

            entries = read_csvs(csv_folder)
            if len(entries) < 12:
                skipped.append(f"{sub}(仅{len(entries)}个CSV)")
                continue

            led, poc = auto_split_led_poc(entries)
            key = (temp_label, volt)
            data_dict[key] = {"csv_folder": csv_folder, "led": led, "poc": poc}

            iid = f"{kind}|{temp_label}|{volt}"
            if tree.exists(iid):
                tree.delete(iid)
            tree.insert("", "end", iid=iid,
                        values=(temp_label, volt, csv_folder,
                                ",".join(str(x[0]) for x in led),
                                ",".join(str(x[0]) for x in poc)))
            loaded += 1

        save_session(self.before_data, self.after_data)

    def _clear(self, kind):
        if kind == "before":
            self.before_data.clear()
            for item in self.tree_before.get_children():
                self.tree_before.delete(item)
        else:
            self.after_data.clear()
            for item in self.tree_after.get_children():
                self.tree_after.delete(item)
        save_session(self.before_data, self.after_data)

    def _restore_session(self):
        session = load_session_folders()
        if not session:
            messagebox.showinfo("提示", "没有找到上次的记录")
            return
        for kind, folders in [("before", session.get("before", {})),
                               ("after",  session.get("after",  {}))]:
            data_dict = self.before_data if kind == "before" else self.after_data
            tree = self.tree_before if kind == "before" else self.tree_after
            for key_str, csv_folder in folders.items():
                if not os.path.isdir(csv_folder):
                    continue
                temp_label, volt = key_str.split("|", 1)
                try:
                    entries = read_csvs(csv_folder)
                    if len(entries) < 12:
                        continue
                    led, poc = auto_split_led_poc(entries)
                except Exception:
                    continue
                key = (temp_label, volt)
                data_dict[key] = {"csv_folder": csv_folder, "led": led, "poc": poc}
                iid = f"{kind}|{temp_label}|{volt}"
                if tree.exists(iid):
                    tree.delete(iid)
                tree.insert("", "end", iid=iid,
                            values=(temp_label, volt, csv_folder,
                                    ",".join(str(x[0]) for x in led),
                                    ",".join(str(x[0]) for x in poc)))

    def export_excel(self):
        if not self.before_data:
            messagebox.showwarning("提示", "请先加载试验前数据")
            return
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="output",
            title="保存Excel文件")
        if not save_path:
            return
        try:
            self._write_excel(save_path)
            if messagebox.askyesno("完成", f"已保存到:\n{save_path}\n\n是否立即打开？"):
                os.startfile(save_path)
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def _write_excel(self, path):
        wb = Workbook()
        has_after = bool(self.after_data)

        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_mid = Alignment(horizontal="left", vertical="center", wrap_text=True)
        fill_hdr = PatternFill("solid", fgColor="D9E1F2")
        fill_sub = PatternFill("solid", fgColor="F2F2F2")
        fill_led = PatternFill("solid", fgColor="E8F5E9")
        fill_poc = PatternFill("solid", fgColor="E3F2FD")
        thin = Side(style="thin")
        bd = Border(left=thin, right=thin, top=thin, bottom=thin)

        def cell(ws, r, c, val="", bold=False, fill=None, align=center, num_fmt=None, color=None):
            cl = ws.cell(r, c, val)
            cl.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=bold,
                           color=color if color else "000000")
            cl.alignment = align
            cl.border = bd
            if fill:
                cl.fill = fill
            if num_fmt:
                cl.number_format = num_fmt
            return cl

        # 所有电压放同一个sheet
        volts = sorted(set(v for _, v in self.before_data.keys()),
                       key=lambda x: VOLT_ORDER.index(x) if x in VOLT_ORDER else 99)

        ws = wb.active
        ws.title = "电流数据"
        n_samples = 6

        def write_block(ws, start_row, block_label, data_dict, _volts=volts):
            # 标题行
            total_cols = 2 + n_samples
            ws.merge_cells(start_row=start_row, start_column=1,
                           end_row=start_row, end_column=total_cols)
            cell(ws, start_row, 1, block_label, bold=True, fill=fill_hdr, align=left_mid)
            # 表头
            h = start_row + 1
            cell(ws, h, 1, "工作条件", bold=True, fill=fill_sub)
            cell(ws, h, 2, "类型", bold=True, fill=fill_sub)
            for s in range(n_samples):
                cell(ws, h, 3 + s, f"样品 {s+1}", bold=True, fill=fill_sub)
            row = h + 1
            for v in _volts:
                temps = sorted(set(t for t, vv in data_dict.keys() if vv == v),
                               key=lambda x: TEMP_ORDER.index(x) if x in TEMP_ORDER else 99)
                for temp in temps:
                    key = (temp, v)
                    entry = data_dict.get(key)
                    ws.merge_cells(start_row=row, start_column=1,
                                   end_row=row + 1, end_column=1)
                    cell(ws, row, 1, temp, bold=True, fill=fill_hdr)
                    for r_off, (ch_type, ch_key, fill_type) in enumerate([
                        ("POC", "poc", fill_poc), (f"LED {v}", "led", fill_led)
                    ]):
                        cur_row = row + r_off
                        cell(ws, cur_row, 2, ch_type, bold=True, fill=fill_type)
                        channels = entry[ch_key] if entry else []
                        for s in range(n_samples):
                            val = channels[s][2] if s < len(channels) else None
                            cell(ws, cur_row, 3 + s,
                                 f"{round(val, 2)}mA" if val is not None else "N/A",
                                 fill=fill_type)
                    row += 2
            return row

        def write_block_combined(ws, start_row, block_label, _volts=volts):
            ws.merge_cells(start_row=start_row, start_column=1,
                           end_row=start_row, end_column=6)
            cell(ws, start_row, 1, block_label, bold=True, fill=fill_hdr, align=left_mid)
            h = start_row + 1
            for ci, hdr in enumerate(["工作条件", "类型", "样品#", "试验前(mA)", "试验后(mA)", "变化%", "results"]):
                cell(ws, h, ci + 1, hdr, bold=True, fill=fill_sub)
            row = h + 1
            for v in _volts:
                temps = sorted(set(t for t, vv in self.before_data.keys() if vv == v),
                               key=lambda x: TEMP_ORDER.index(x) if x in TEMP_ORDER else 99)
                for temp in temps:
                    key = (temp, v)
                    b = self.before_data.get(key)
                    a = self.after_data.get(key)
                    total_rows = n_samples * 2
                    ws.merge_cells(start_row=row, start_column=1,
                                   end_row=row + total_rows - 1, end_column=1)
                    cell(ws, row, 1, temp, bold=True, fill=fill_hdr)
                    for ch_type, ch_key, fill_type in [
                        ("POC", "poc", fill_poc), (f"LED {v}", "led", fill_led)
                    ]:
                        b_ch = b[ch_key] if b else []
                        a_ch = a[ch_key] if a else []
                        ws.merge_cells(start_row=row, start_column=2,
                                       end_row=row + n_samples - 1, end_column=2)
                        cell(ws, row, 2, ch_type, bold=True, fill=fill_type)
                        for s in range(n_samples):
                            bv = b_ch[s][2] if s < len(b_ch) else None
                            av = a_ch[s][2] if s < len(a_ch) else None
                            cell(ws, row + s, 3, f"{s+1}#", fill=fill_type)
                            cell(ws, row + s, 4,
                                 f"{round(bv, 2)}mA" if bv is not None else "N/A", fill=fill_type)
                            cell(ws, row + s, 5,
                                 f"{round(av, 2)}mA" if av is not None else "N/A", fill=fill_type)
                            if bv is not None and av is not None and bv != 0:
                                pct = (av - bv) / bv * 100
                                cell(ws, row + s, 6, f"{round(pct, 2)}%",
                                     fill=fill_type)
                                result = "PASS" if abs(pct) <= 20 else "FAIL"
                            else:
                                cell(ws, row + s, 6, "N/A", fill=fill_type)
                                result = "N/A"
                            cell(ws, row + s, 7, result, fill=fill_type)
                        row += n_samples
            return row

        cur_row = 1
        if has_after:
            cur_row = write_block(ws, cur_row, "试验后 (mA)", self.after_data)
            cur_row += 1
            write_block_combined(ws, cur_row, "对比（试验前 / 试验后 / 变化%）")
        else:
            write_block(ws, cur_row, "试验后 (mA)", self.before_data)

        # 列宽
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 8
        for c in ["D", "E", "F", "G", "H", "I"]:
            ws.column_dimensions[c].width = 12

        wb.save(path)


if __name__ == "__main__":
    app = App()
    app.mainloop()
